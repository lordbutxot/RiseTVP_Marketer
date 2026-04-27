import argparse
import datetime
import json
import os
import re
from collections import defaultdict

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import CondPageBreak, Image, KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


NEWSPAPER_NAME = "The Vieneo Index"
DEFAULT_IMAGE_OUTPUT_DIR = "TVI_Output"
DEFAULT_OUTPUT = os.path.join(DEFAULT_IMAGE_OUTPUT_DIR, "The_Vieneo_Index.pdf")
ADS_STATE_FILE = ".tvi_ads_state.json"


def build_config_data(selected_ship=None, origin=None, ship_capacity=None, budget=None, status=None):
    return {
        "ship": selected_ship or "Unspecified",
        "origin": origin or "Unspecified",
        "ship_capacity": ship_capacity,
        "budget": budget,
        "status": status,
    }


def _fmt_int(value):
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return "0"


def _fmt_cr(value):
    return f"{_fmt_int(value)} CR"


def _safe_get(row, index, default=None):
    if not row:
        return default
    if index < len(row):
        return row[index]
    return default


def _load_header_image(header_image):
    if header_image and os.path.exists(header_image):
        return header_image

    fallback = os.path.join("images", "TVI_Index.jpg")
    if os.path.exists(fallback):
        return fallback
    return None


def _find_ad_library():
    image_dir = "images"
    if not os.path.isdir(image_dir):
        return []
    ads = []
    for name in sorted(os.listdir(image_dir)):
        upper = name.upper()
        if upper.startswith("TVI_AD") and upper.endswith((".PNG", ".JPG", ".JPEG", ".WEBP")):
            ads.append(os.path.join(image_dir, name))
    return ads


def _load_rotating_ads(slot_count=3):
    ads = _find_ad_library()
    if not ads:
        return [None] * slot_count

    offset = 0
    if os.path.exists(ADS_STATE_FILE):
        try:
            with open(ADS_STATE_FILE, "r", encoding="utf-8") as handle:
                offset = int(json.load(handle).get("offset", 0))
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            offset = 0

    chosen = [ads[(offset + idx) % len(ads)] for idx in range(slot_count)]

    try:
        with open(ADS_STATE_FILE, "w", encoding="utf-8") as handle:
            json.dump({"offset": (offset + 1) % len(ads)}, handle)
    except OSError:
        pass

    return chosen


def _export_pdf_pages_as_images(pdf_path, output_dir=DEFAULT_IMAGE_OUTPUT_DIR, zoom=2.0):
    import fitz

    os.makedirs(output_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]

    doc = fitz.open(pdf_path)
    try:
        matrix = fitz.Matrix(zoom, zoom)
        for page_index in range(len(doc)):
            page = doc.load_page(page_index)
            pix = page.get_pixmap(matrix=matrix, alpha=False)
            out_path = os.path.join(output_dir, f"{base_name}_page_{page_index + 1:02d}.png")
            pix.save(out_path)
    finally:
        doc.close()


def _newspaper_styles():
    styles = getSampleStyleSheet()
    return {
        "masthead": ParagraphStyle(
            "Masthead",
            parent=styles["Title"],
            fontName="Times-Bold",
            fontSize=28,
            leading=30,
            alignment=1,
            textColor=colors.HexColor("#101010"),
            spaceAfter=2,
        ),
        "edition": ParagraphStyle(
            "Edition",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=10,
            alignment=1,
            textColor=colors.HexColor("#505050"),
            spaceAfter=2,
        ),
        "banner": ParagraphStyle(
            "Banner",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            alignment=1,
            textColor=colors.white,
        ),
        "headline": ParagraphStyle(
            "Headline",
            parent=styles["Heading1"],
            fontName="Times-Bold",
            fontSize=20,
            leading=23,
            textColor=colors.HexColor("#111111"),
            spaceAfter=4,
        ),
        "subheadline": ParagraphStyle(
            "Subheadline",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=13,
            textColor=colors.HexColor("#303030"),
            spaceAfter=6,
        ),
        "section": ParagraphStyle(
            "Section",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=10.5,
            leading=12,
            textColor=colors.HexColor("#8A1538"),
            spaceBefore=2,
            spaceAfter=2,
        ),
        "body": ParagraphStyle(
            "Body",
            parent=styles["BodyText"],
            fontName="Times-Roman",
            fontSize=8.8,
            leading=11,
            textColor=colors.HexColor("#1C1C1C"),
            spaceAfter=2,
        ),
        "brief": ParagraphStyle(
            "Brief",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=colors.HexColor("#1C1C1C"),
            spaceAfter=2,
        ),
        "table_title": ParagraphStyle(
            "TableTitle",
            parent=styles["Heading3"],
            fontName="Helvetica-Bold",
            fontSize=8.2,
            leading=10,
            textColor=colors.HexColor("#101010"),
            spaceAfter=2,
        ),
        "quote": ParagraphStyle(
            "Quote",
            parent=styles["BodyText"],
            fontName="Times-Italic",
            fontSize=8.3,
            leading=10,
            alignment=1,
            textColor=colors.HexColor("#404040"),
            spaceAfter=1,
        ),
        "footer": ParagraphStyle(
            "Footer",
            parent=styles["Normal"],
            fontName="Helvetica-Oblique",
            fontSize=7.5,
            leading=9,
            alignment=1,
            textColor=colors.HexColor("#666666"),
        ),
    }


def _hr(widths, color="#101010", padding=(0, 0, 0, 0)):
    table = Table([[""]], colWidths=widths)
    table.setStyle(
        TableStyle(
            [
                ("LINEABOVE", (0, 0), (-1, -1), 1, colors.HexColor(color)),
                ("TOPPADDING", (0, 0), (-1, -1), padding[0]),
                ("BOTTOMPADDING", (0, 0), (-1, -1), padding[1]),
                ("LEFTPADDING", (0, 0), (-1, -1), padding[2]),
                ("RIGHTPADDING", (0, 0), (-1, -1), padding[3]),
            ]
        )
    )
    return table


def _market_summary_story(styles, macro_data, opportunities, config_data):
    top_city = next(iter(macro_data.get("top_cities_by_profit", [])), None)
    best_trip = next(iter(sorted(opportunities or [], key=lambda item: item.get("profit_per_mt", 0), reverse=True)), None)

    headline_parts = []
    if top_city:
        headline_parts.append(
            f"{top_city['city']} leads the Rise economy with {_fmt_cr(top_city['total_profit'])} in mapped opportunity."
        )
    if best_trip:
        headline_parts.append(
            f"The strongest route on file is {best_trip['source']} to {best_trip['destination']} in {best_trip['commodity']}, returning {_fmt_cr(best_trip.get('profit_per_mt', 0))} per MT."
        )
    text = " ".join(headline_parts) or "The current issue tracks the best commercial openings across The Vineo Province."
    return [
        Paragraph("Market Front Page", styles["section"]),
        Paragraph("Capital rotates toward the highest-yield corridors", styles["headline"]),
        Paragraph(text, styles["subheadline"]),
    ]


def _build_opinion_opening(macro_data, opportunities):
    city_summary = macro_data.get("city_summary", [])
    top_city = city_summary[0] if city_summary else None
    runner_up = city_summary[1] if len(city_summary) > 1 else None
    best_route = next(iter(sorted(opportunities or [], key=lambda item: item.get("profit_per_mt", 0), reverse=True)), None)

    lines = []
    if top_city and runner_up:
        gap = max((top_city.get("total_profit", 0) - runner_up.get("total_profit", 0)), 0)
        lines.append(
            f"{top_city['city']} opens the day {_fmt_cr(gap)} ahead of the next city in mapped margin, a sign of broad local depth rather than one isolated route."
        )
    elif top_city:
        lines.append(
            f"{top_city['city']} sets the tone for the day, leading the economic board with {_fmt_cr(top_city.get('total_profit', 0))} in mapped margin."
        )

    if best_route:
        lines.append(
            f"The sharpest signal remains {best_route['commodity']} from {best_route['source']} to {best_route['destination']}, yielding {_fmt_cr(best_route.get('profit_per_mt', 0))} per MT."
        )

    return " ".join(lines) or "Capital remains selective across The Vineo Province, rewarding traders who follow unit spread rather than noise."


def _build_opinion_mid(macro_data):
    commodity_rows = _make_commodity_analysis_rows(macro_data, limit=3)
    city_summary = macro_data.get("city_summary", [])
    active_cities = [item for item in city_summary if item.get("num_lucrative", 0) >= 3]

    lines = []
    if commodity_rows:
        lead = commodity_rows[0]
        lines.append(
            f"{lead[0]} leads the commodity board with a {lead[1]} spread, keeping attention on focused cargo rather than diluted routing."
        )
    if active_cities:
        lines.append(
            f"{len(active_cities)} cities are carrying at least three lucrative goods, pointing to distributed opportunity across the province."
        )

    return " ".join(lines) or "The daily board suggests a market with enough breadth to reward planning and enough dispersion to punish lazy routing."


def _build_opinion_closing(macro_data, opportunities):
    best_route = next(iter(sorted(opportunities or [], key=lambda item: item.get("profit_per_mt", 0), reverse=True)), None)
    top_city = next(iter(macro_data.get("top_cities_by_profit", [])), None)

    if best_route and top_city:
        variants = [
            f"Closing desk: when {top_city['city']} leads the board and {best_route['commodity']} still clears {_fmt_cr(best_route.get('profit_per_mt', 0))} per MT, the market is reminding everyone that price discipline beats cargo volume.",
            f"Closing desk: today's board argues for focus over noise. {top_city['city']} holds the broadest edge, while {best_route['commodity']} keeps proving that the best unit margin usually tells the truth first.",
            f"Closing desk: capital rarely needs a loud signal. A leading city like {top_city['city']} and a clean {best_route['commodity']} spread are usually enough to define the day."
        ]
        selector = (len(top_city["city"]) + len(best_route["commodity"]) + int(best_route.get("profit_per_mt", 0))) % len(variants)
        return variants[selector]
    return "Closing desk: the market rarely rewards the loudest route; it rewards the cleanest spread."


def _build_quote_of_day(macro_data, opportunities):
    best_route = next(iter(sorted(opportunities or [], key=lambda item: item.get("profit_per_mt", 0), reverse=True)), None)
    top_city = next(iter(macro_data.get("top_cities_by_profit", [])), None)
    if best_route:
        quotes = [
            f'"A clean margin per MT is worth more than a crowded hold without pricing power." Today\'s board still proves it in {best_route["commodity"]}.',
            f'"Markets whisper through spreads before they shout through volume." {best_route["commodity"]} is the whisper to watch today.',
            f'"The disciplined trader reads the unit price before the cargo hold." {top_city["city"] if top_city else "The province"} is rewarding that habit again.',
            f'"Distance is only expensive when the spread is weak." Today\'s routes keep making that case.'
        ]
        selector = (datetime.datetime.now().timetuple().tm_yday + len(best_route["commodity"])) % len(quotes)
        return quotes[selector]
    return '"Trade follows discipline before it follows distance."'


def _make_table(title, headers, rows, col_widths, header_fill="#E9E2D0", accent="#8A1538"):
    data = [[Paragraph(f"<b>{h}</b>", getSampleStyleSheet()["BodyText"]) for h in headers]]
    data.extend(rows)
    table = Table(data, colWidths=col_widths, repeatRows=1, splitByRow=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(header_fill)),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#101010")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.4),
                ("LEADING", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#BEB7A4")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFAF6")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("LINEABOVE", (0, 0), (-1, 0), 1, colors.HexColor(accent)),
                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor(accent)),
            ]
        )
    )
    return KeepTogether([Paragraph(title, _newspaper_styles()["table_title"]), Spacer(1, 1), table])


def _ad_slot(styles, image_path, slot_label):
    if image_path and os.path.exists(image_path):
        block = [
            Paragraph(slot_label, styles["table_title"]),
            Image(image_path, width=186 * mm, height=28 * mm),
            Spacer(1, 3),
        ]
        return KeepTogether(block)

    placeholder = Table(
        [[Paragraph("Reserved advertising space", styles["brief"])]],
        colWidths=[186 * mm],
        rowHeights=[28 * mm],
    )
    placeholder.setStyle(
        TableStyle(
            [
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#8A1538")),
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F7F1E4")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    return KeepTogether([Paragraph(slot_label, styles["table_title"]), placeholder, Spacer(1, 3)])


def _section_block(styles, title, elements, min_space_mm=24):
    block = [CondPageBreak(min_space_mm * mm), Paragraph(title, styles["section"])]
    block.extend(elements)
    return block


def _draw_page_chrome(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor("#8A1538"))
    canvas.setLineWidth(0.6)
    canvas.line(doc.leftMargin, A4[1] - 8 * mm, A4[0] - doc.rightMargin, A4[1] - 8 * mm)
    canvas.line(doc.leftMargin, 8 * mm, A4[0] - doc.rightMargin, 8 * mm)
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(colors.HexColor("#555555"))
    canvas.drawString(doc.leftMargin, 4.8 * mm, NEWSPAPER_NAME)
    canvas.drawRightString(A4[0] - doc.rightMargin, 4.8 * mm, f"Page {canvas.getPageNumber()}")
    canvas.restoreState()


def _route_groups(opportunities):
    bands = {"Short Haul": [], "Mid Range": [], "Long Haul": []}
    for op in opportunities or []:
        travel_time = op.get("travel_time", 999)
        if travel_time < 60:
            bands["Short Haul"].append(op)
        elif travel_time < 120:
            bands["Mid Range"].append(op)
        else:
            bands["Long Haul"].append(op)
    return bands


def _make_route_rows(items, limit=4):
    ordered = sorted(items, key=lambda item: item.get("profit_per_mt", 0), reverse=True)[:limit]
    rows = []
    for op in ordered:
        rows.append(
            [
                op.get("commodity", "-"),
                f"{op.get('source', '-')} -> {op.get('destination', '-')}",
                _fmt_cr(op.get("profit_per_mt", 0)),
                f"{op.get('travel_time', 0)} min",
            ]
        )
    return rows


def _make_city_rows(city_summary, limit=8):
    rows = []
    for item in (city_summary or [])[:limit]:
        rows.append(
            [
                item.get("city", "-"),
                _fmt_cr(item.get("total_profit", 0)),
                str(item.get("num_commodities", 0)),
                str(item.get("num_lucrative", 0)),
            ]
        )
    return rows


def _make_commodity_rows(macro_data, limit=8):
    rows = []
    for commodity, sellers in macro_data.get("commodity_best_sellers", {}).items():
        buyers = macro_data.get("commodity_best_buyers", {}).get(commodity, [])
        if not sellers or not buyers:
            continue
        spread = buyers[0]["price"] - sellers[0]["price"]
        rows.append(
            [
                commodity,
                f"{sellers[0]['city']} @ {_fmt_int(sellers[0]['price'])}",
                f"{buyers[0]['city']} @ {_fmt_int(buyers[0]['price'])}",
                _fmt_cr(spread),
            ]
        )
    rows.sort(key=lambda item: int(item[3].replace(",", "").split()[0]), reverse=True)
    return rows[:limit]


def _make_city_analysis_rows(city_summary, limit=10):
    rows = []
    for item in (city_summary or [])[:limit]:
        lucrative = item.get("num_lucrative", 0)
        tracked = item.get("num_commodities", 0)
        coverage = (lucrative / tracked * 100) if tracked else 0
        rows.append(
            [
                item.get("city", "-"),
                _fmt_cr(item.get("total_profit", 0)),
                str(lucrative),
                f"{coverage:.1f}%",
                "Expansion" if coverage >= 60 else "Selective",
            ]
        )
    return rows


def _make_commodity_analysis_rows(macro_data, limit=10):
    rows = []
    for commodity, sellers in macro_data.get("commodity_best_sellers", {}).items():
        buyers = macro_data.get("commodity_best_buyers", {}).get(commodity, [])
        if not sellers or not buyers:
            continue
        best_sell = sellers[0]
        best_buy = buyers[0]
        spread = best_buy["price"] - best_sell["price"]
        intensity = "Hot" if spread >= 1000 else "Active" if spread >= 500 else "Watch"
        rows.append(
            [
                commodity,
                _fmt_cr(spread),
                best_sell["city"],
                best_buy["city"],
                intensity,
            ]
        )
    rows.sort(key=lambda item: int(item[1].replace(",", "").split()[0]), reverse=True)
    return rows[:limit]


def _make_profitable_commodity_leaderboard(opportunities, limit=8):
    grouped = {}
    for op in opportunities or []:
        commodity = op.get("commodity")
        if not commodity:
            continue
        bucket = grouped.setdefault(
            commodity,
            {
                "count": 0,
                "total_margin": 0,
                "best_margin": 0,
                "best_route": "",
            },
        )
        margin = op.get("profit_per_mt", 0) or 0
        bucket["count"] += 1
        bucket["total_margin"] += margin
        if margin > bucket["best_margin"]:
            bucket["best_margin"] = margin
            bucket["best_route"] = f"{op.get('source', '-')} -> {op.get('destination', '-')}"

    rows = []
    for commodity, data in grouped.items():
        avg_margin = data["total_margin"] / data["count"] if data["count"] else 0
        rows.append(
            [
                commodity,
                _fmt_cr(round(avg_margin)),
                _fmt_cr(data["best_margin"]),
                str(data["count"]),
                data["best_route"],
            ]
        )

    rows.sort(key=lambda item: int(item[2].replace(",", "").split()[0]), reverse=True)
    return rows[:limit]


def _make_trade_chain_rows(routes, limit=5):
    rows = []
    for route in (routes or [])[:limit]:
        rows.append(
            [
                " -> ".join(route.get("stops", [])),
                _fmt_cr(route.get("total_profit", 0)),
                f"{route.get('roi', 0):.1f}%",
                f"{route.get('total_time', 0)} min",
            ]
        )
    return rows


def generate_news_pdf(
    macro_data,
    opportunities=None,
    routes=None,
    config_data=None,
    output_file=DEFAULT_OUTPUT,
    header_image=None,
    ads_images=None,
    image_output_dir=DEFAULT_IMAGE_OUTPUT_DIR,
):
    styles = _newspaper_styles()
    rotating_ads = ads_images if ads_images is not None else _load_rotating_ads(3)
    doc = SimpleDocTemplate(
        output_file,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    story = []
    header_image = _load_header_image(header_image)
    today = datetime.datetime.now().strftime("%d %B %Y | %H:%M")

    if header_image:
        story.append(Image(header_image, width=186 * mm, height=34 * mm))
        story.append(Spacer(1, 2))

    story.append(Paragraph(NEWSPAPER_NAME, styles["masthead"]))
    story.append(Paragraph(f"Economic Edition | Rise: The Vineo Province | {today}", styles["edition"]))

    banner_text = "Professional market intelligence for pilots, traders and route planners"
    banner = Table([[Paragraph(banner_text, styles["banner"])]], colWidths=[186 * mm])
    banner.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#111111")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(banner)
    story.append(Spacer(1, 3))
    story.append(_hr([186 * mm], color="#8A1538"))
    story.append(Spacer(1, 2))

    story.extend(_market_summary_story(styles, macro_data, opportunities or [], config_data or {}))
    story.append(Paragraph("Daily Opinion", styles["section"]))
    story.append(Paragraph(_build_opinion_opening(macro_data, opportunities or []), styles["body"]))
    story.append(Paragraph(_build_opinion_mid(macro_data), styles["body"]))

    top_city = next(iter(macro_data.get("top_cities_by_profit", [])), None)
    city_line = (
        f"<b>Lead Province:</b> {top_city['city']} with {_fmt_cr(top_city['total_profit'])} in total mapped margin."
        if top_city
        else "<b>Lead Province:</b> Market scan active."
    )
    story.append(Paragraph(city_line, styles["body"]))

    route_bands = _route_groups(opportunities or [])
    route_tables = []
    for label, items in route_bands.items():
        rows = _make_route_rows(items)
        if rows:
            route_tables.append(
                _make_table(
                    title=label,
                    headers=["Commodity", "Route", "Margin / MT", "Travel"],
                    rows=rows,
                    col_widths=[45 * mm, 78 * mm, 32 * mm, 24 * mm],
                )
            )
            route_tables.append(Spacer(1, 3))
    if route_tables:
        story.extend(_section_block(styles, "Route Desk", route_tables))
        story.append(_ad_slot(styles, rotating_ads[0] if len(rotating_ads) > 0 else None, "Sponsored Placement I"))

    city_rows = _make_city_rows(macro_data.get("city_summary", []))
    commodity_rows = _make_commodity_rows(macro_data)
    ledger_elements = []

    if city_rows:
        ledger_elements.append(_make_table(
            title="City Economy Index",
            headers=["City", "Total Profit", "Tracked Goods", "Lucrative Goods"],
            rows=city_rows,
            col_widths=[46 * mm, 46 * mm, 42 * mm, 52 * mm],
        ))
        ledger_elements.append(Spacer(1, 3))

    city_analysis_rows = _make_city_analysis_rows(macro_data.get("city_summary", []))
    if city_analysis_rows:
        ledger_elements.append(Paragraph(
            "City Analysis",
            styles["table_title"],
        ))
        ledger_elements.append(Paragraph(
            "A city-level reading of total mapped margin, count of lucrative goods and how broad each market's profitable coverage looks right now.",
            styles["body"],
        ))
        ledger_elements.append(_make_table(
            title="City Analysis",
            headers=["City", "Total Margin", "Lucrative Goods", "Coverage", "Market Tone"],
            rows=city_analysis_rows,
            col_widths=[40 * mm, 40 * mm, 34 * mm, 28 * mm, 44 * mm],
        ))
        ledger_elements.append(Spacer(1, 3))

    if commodity_rows:
        ledger_elements.append(_make_table(
            title="Commodity Spread Monitor",
            headers=["Commodity", "Best Seller", "Best Buyer", "Spread"],
            rows=commodity_rows,
            col_widths=[52 * mm, 48 * mm, 48 * mm, 38 * mm],
        ))
        ledger_elements.append(Spacer(1, 3))

    if ledger_elements:
        ledger_elements.append(_ad_slot(styles, rotating_ads[1] if len(rotating_ads) > 1 else None, "Sponsored Placement II"))
        story.extend(_section_block(styles, "Provincial Ledger", ledger_elements))

    commodity_analysis_rows = _make_commodity_analysis_rows(macro_data)
    if commodity_analysis_rows:
        commodity_analysis_elements = [
            Paragraph(
                "A commodity-by-commodity view of unit spread, strongest accumulation point and strongest sell-side outlet across the province.",
                styles["body"],
            ),
            _make_table(
                title="Commodity Analysis",
                headers=["Commodity", "Spread / MT", "Best Buy Origin", "Best Sell Destination", "Signal"],
                rows=commodity_analysis_rows,
                col_widths=[46 * mm, 30 * mm, 40 * mm, 44 * mm, 26 * mm],
            ),
            Spacer(1, 3),
            _ad_slot(styles, rotating_ads[2] if len(rotating_ads) > 2 else None, "Sponsored Placement III"),
        ]
        story.extend(_section_block(styles, "Commodity Analysis", commodity_analysis_elements))

    if opportunities:
        limited = []
        commodity_limit = defaultdict(int)
        for op in sorted(opportunities, key=lambda item: item.get("profit_per_mt", 0), reverse=True):
            commodity = op.get("commodity", "")
            if commodity_limit[commodity] >= 3:
                continue
            commodity_limit[commodity] += 1
            limited.append(
                [
                    commodity,
                    f"{op.get('source', '-')} -> {op.get('destination', '-')}",
                    _fmt_cr(op.get("profit_per_mt", 0)),
                    f"{op.get('source_selling', 0)} -> {op.get('destination_buying', 0)}",
                ]
            )
            if len(limited) == 15:
                break
        if limited:
            exchange_elements = [
                _make_table(
                title="Top 15 Market Opportunities",
                headers=["Commodity", "Route", "Margin / MT", "Price Window"],
                rows=limited,
                col_widths=[48 * mm, 76 * mm, 34 * mm, 28 * mm],
                ),
                Spacer(1, 3),
            ]
            story.extend(_section_block(styles, "Exchange Floor", exchange_elements))

    commodity_leaderboard_rows = _make_profitable_commodity_leaderboard(opportunities or [])
    if commodity_leaderboard_rows:
        leaderboard_elements = [
            Paragraph(
                "A final profitability board ranking the commodities that are producing the strongest trading conditions across the widest set of routes.",
                styles["body"],
            ),
            _make_table(
                title="Profitable Commodity Leaderboard",
                headers=["Commodity", "Avg Margin / MT", "Best Margin / MT", "Live Routes", "Best Route"],
                rows=commodity_leaderboard_rows,
                col_widths=[34 * mm, 32 * mm, 32 * mm, 22 * mm, 66 * mm],
            ),
            Spacer(1, 3),
        ]
        story.extend(_section_block(styles, "Commodity Leaderboard", leaderboard_elements))

    closing = _build_opinion_closing(macro_data, opportunities or [])
    story.extend(_section_block(styles, "Closing Opinion", [Paragraph(closing, styles["body"])]))
    story.append(Paragraph("Quote of the Day", styles["table_title"]))
    story.append(Paragraph(_build_quote_of_day(macro_data, opportunities or []), styles["quote"]))
    story.append(_hr([186 * mm], color="#8A1538"))
    story.append(Spacer(1, 2))
    story.append(Paragraph("Editorial desk: refreshed from live provincial trade sheets.", styles["footer"]))

    doc.build(story, onFirstPage=_draw_page_chrome, onLaterPages=_draw_page_chrome)
    _export_pdf_pages_as_images(output_file, output_dir=image_output_dir)
    print(f"The Vieneo Index generated: {output_file}")


def _find_sheet(workbook, prefix):
    for name in workbook.sheetnames:
        if name.lower().startswith(prefix.lower()):
            return workbook[name]
    return None


def build_news_from_workbook(workbook_path, output_file=DEFAULT_OUTPUT, header_image=None, image_output_dir=DEFAULT_IMAGE_OUTPUT_DIR):
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    config_ws = wb["Config"] if "Config" in wb.sheetnames else None
    opp_ws = wb["Opportunities"] if "Opportunities" in wb.sheetnames else None
    macro_ws = wb["MACRO"] if "MACRO" in wb.sheetnames else None
    route_ws = _find_sheet(wb, "Trade Routes ")

    config_map = {}
    if config_ws:
        for row in config_ws.iter_rows(min_row=2, values_only=True):
            key = _safe_get(row, 0)
            value = _safe_get(row, 1)
            if key:
                config_map[str(key)] = value

    opportunities = []
    if opp_ws and opp_ws.max_row >= 2:
        for row in opp_ws.iter_rows(min_row=2, values_only=True):
            commodity = _safe_get(row, 1)
            source = _safe_get(row, 2)
            destination = _safe_get(row, 4)
            if not commodity or not source or not destination:
                continue
            opportunities.append(
                {
                    "grade": _safe_get(row, 0),
                    "commodity": commodity,
                    "source": source,
                    "source_selling": _safe_get(row, 3) or 0,
                    "destination": destination,
                    "destination_buying": _safe_get(row, 5) or 0,
                    "profit_per_mt": _safe_get(row, 6) or 0,
                    "source_available": _safe_get(row, 7) or 0,
                    "destination_capacity": _safe_get(row, 8) or 0,
                    "_qty_trip": _safe_get(row, 9) or 0,
                    "_cost_trip": _safe_get(row, 10) or 0,
                    "_profit_trip": _safe_get(row, 11) or 0,
                    "_roi": _safe_get(row, 12) or 0,
                    "travel_time": _safe_get(row, 13) or 0,
                }
            )

    macro_data = {
        "city_summary": [],
        "commodity_best_sellers": defaultdict(list),
        "commodity_best_buyers": defaultdict(list),
        "top_cities_by_profit": [],
        "cities_by_lucrative_commodities": [],
    }

    if macro_ws and macro_ws.max_row > 2:
        mode = None
        current_commodity = None
        commodity_mode = None
        for row in macro_ws.iter_rows(values_only=True):
            first = _safe_get(row, 0)
            if first == "SUMMARY BY CITY":
                mode = "city_summary"
                continue
            if first == "TOP 10 CITIES BY TOTAL PROFIT":
                mode = "top_city"
                continue
            if first == "TOP 10 CITIES BY NUMBER OF LUCRATIVE COMMODITIES":
                mode = "lucrative_city"
                continue
            if first == "BEST SELLERS AND BUYERS BY COMMODITY":
                mode = "commodity"
                continue
            if first in (None, "", "City", "Commodity"):
                continue

            if mode == "city_summary":
                macro_data["city_summary"].append(
                    {
                        "city": first,
                        "total_profit": _safe_get(row, 1) or 0,
                        "num_commodities": _safe_get(row, 2) or 0,
                        "num_lucrative": _safe_get(row, 3) or 0,
                    }
                )
            elif mode == "top_city":
                macro_data["top_cities_by_profit"].append(
                    {
                        "city": first,
                        "total_profit": _safe_get(row, 1) or 0,
                        "num_commodities": 0,
                        "num_lucrative": _safe_get(row, 2) or 0,
                    }
                )
            elif mode == "lucrative_city":
                macro_data["cities_by_lucrative_commodities"].append(
                    {
                        "city": first,
                        "total_profit": _safe_get(row, 2) or 0,
                        "num_commodities": 0,
                        "num_lucrative": _safe_get(row, 1) or 0,
                    }
                )
            elif mode == "commodity":
                if isinstance(first, str) and first.startswith("--- ") and first.endswith(" ---"):
                    current_commodity = first.replace("--- ", "").replace(" ---", "").title()
                    if current_commodity == "Rare/Precious":
                        current_commodity = "Rare/Precious"
                    continue
                if isinstance(first, str) and "Best Sellers" in first:
                    commodity_mode = "seller"
                    continue
                if isinstance(first, str) and "Best Buyers" in first:
                    commodity_mode = "buyer"
                    continue
                if _safe_get(row, 1) is None:
                    continue
                commodity = current_commodity
                if not commodity:
                    continue
                if commodity_mode == "seller":
                    macro_data["commodity_best_sellers"][commodity].append(
                        {"city": first, "price": _safe_get(row, 1) or 0, "capacity": _safe_get(row, 2) or 0}
                    )
                elif commodity_mode == "buyer":
                    macro_data["commodity_best_buyers"][commodity].append(
                        {"city": first, "price": _safe_get(row, 1) or 0, "capacity": _safe_get(row, 2) or 0}
                    )

    if not macro_data["top_cities_by_profit"]:
        macro_data["top_cities_by_profit"] = macro_data["city_summary"][:10]

    routes = []
    if route_ws and route_ws.max_row >= 1:
        for row in route_ws.iter_rows(values_only=True):
            summary = _safe_get(row, 0)
            if not isinstance(summary, str) or not summary.startswith("Route #"):
                continue
            parts = [part.strip() for part in summary.split("|")]
            route_path = parts[1] if len(parts) > 1 else ""
            profit_match = re.search(r"Profit:\s*([\d,]+)", summary)
            roi_match = re.search(r"ROI:\s*([\d.]+)", summary)
            time_match = re.search(r"Time:\s*([\d.]+)", summary)
            routes.append(
                {
                    "stops": [part.strip() for part in re.split(r"\s*(?:->|→)\s*", route_path) if part.strip()],
                    "total_profit": int((profit_match.group(1) if profit_match else "0").replace(",", "")),
                    "roi": float(roi_match.group(1)) if roi_match else 0,
                    "total_time": float(time_match.group(1)) if time_match else 0,
                }
            )

    config_data = build_config_data(
        selected_ship=config_map.get("Ship"),
        origin=config_map.get("Origin"),
        ship_capacity=config_map.get("Total Capacity (MT)"),
        budget=config_map.get("Initial Budget"),
        status=config_map.get("Status"),
    )

    generate_news_pdf(
        macro_data=macro_data,
        opportunities=opportunities,
        routes=routes,
        config_data=config_data,
        output_file=output_file,
        header_image=header_image,
        image_output_dir=image_output_dir,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate The Vieneo Index PDF.")
    parser.add_argument("--workbook", default="final_trade.xlsx", help="Workbook used to build the newspaper.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="PDF output path.")
    parser.add_argument("--header-image", default=None, help="Optional masthead image path.")
    parser.add_argument("--image-output-dir", default=DEFAULT_IMAGE_OUTPUT_DIR, help="Folder where PDF pages will also be exported as images.")
    args = parser.parse_args()

    build_news_from_workbook(
        workbook_path=args.workbook,
        output_file=args.output,
        header_image=args.header_image,
        image_output_dir=args.image_output_dir,
    )
