import argparse
import itertools
import json
import math
import os
import platform
import re
import time
from dataclasses import dataclass, field
from typing import Optional
import logging
import sys
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
from pytesseract import Output
import pytesseract

debug_folder = "debug_images"
if not os.path.exists(debug_folder):
    os.makedirs(debug_folder)


# --- TESSERACT CONFIG ---
if platform.system() == "Windows":
    for _cand in [r"C:\Program Files\Tesseract-OCR\tesseract.exe", r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"]:
        if os.path.isfile(_cand):
            pytesseract.pytesseract.tesseract_cmd = _cand
            break

# --- CONSTANTS ---
COMMODITY_CATEGORIES = ["Rare/Precious", "Foodstuffs", "Natural Materials", "Fuel Ore", "Consumer Goods", "Fabricated Material", "Refined Fuel"]
COMMODITY_ORDER = COMMODITY_CATEGORIES

GRADE_STYLES = {
    "A": {"fill": "00B050", "font": "FFFFFF"},
    "B": {"fill": "92D050", "font": "000000"},
    "C": {"fill": "FFC000", "font": "000000"},
    "D": {"fill": "A6A6A6", "font": "FFFFFF"},
}

SHIPS = {
    "AIR AND SPACE": {
        "E-10 Saint": {"cargo_base": 7, "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "E-11 Saint": {"cargo_base": 7, "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "P-13 Prowler": {"cargo_base": 1, "max_containers": 0, "container_mt": 17, "rental_cost_per_day": None},
        "W-6 Manx": {"cargo_base": 7, "max_containers": 3, "container_mt": 17, "rental_cost_per_day": None},
    },
    "ONLY AIR": {
        "A-4 Wanderer": {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
        "T-19 Stratomaster": {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
    },
}

CITIES = {
    "Alphaville": {"x": 420, "y": 180}, 
    "Comstock": {"x": 380, "y": 210}, 
    "Deadwood": {"x": 340, "y": 230},
    "Ederar": {"x": 300, "y": 200}, 
    "Erie": {"x": 460, "y": 250},
    "Freedom": {"x": 200, "y": 320},
    "Gettysburg": {"x": 400, "y": 290},
    "Kansas": {"x": 250, "y": 300}, 
    "Lancaster": {"x": 350, "y": 350}, 
    "Pimlico": {"x": 270, "y": 280}, 
    "SovietUnion": {"x": 500, "y": 150}, 
    "Terrazul": {"x": 480, "y": 320},
    "Sharney 1": {"x": 320, "y": 400},
    "Bethleham": {"x": 0, "y": 0},
    "Deois":     {"x": 310, "y": 260},
    "Outer D":   {"x": 0, "y": 0},
    "Papanui":   {"x": 0, "y": 0},
    "Sharney":   {"x": 0, "y": 0},
    "Solaris":   {"x": 0, "y": 0},
    "Vegapoint": {"x": 0, "y": 0},
    "CapeW":     {"x": 0, "y": 0},
}
CITY_LIST = sorted(CITIES.keys())
TAKEOFF_TIME, LANDING_TIME, SPEED_UNITS_PER_MIN = 5, 10, 5.0

CITY_ROW_KEYWORDS = {
    "Rare/Precious": ("rare", "precious"), "Foodstuffs": ("food", "stuff"),
    "Natural Materials": ("natural", "mater"), "Fuel Ore": ("fuel ore", "ore"),
    "Consumer Goods": ("consumer", "goods"), "Fabricated Material": ("fabricated", "fabric"),
    "Refined Fuel": ("refined", "fuel"),
}

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[logging.FileHandler("debug_log.txt", mode='w', encoding='utf-8'), logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# FORZAR stdout a UTF-8
sys.stdout.reconfigure(encoding='utf-8')
logger = logging.getLogger(__name__)

# --- LAYER 1: OCR & DATA STRUCTURES ---
@dataclass
class CommodityRow:
    commodity: str
    quantity_mt: int = 0
    reserve_mt: int = 0
    sell_price: Optional[int] = None  # Price YOU pay to BUY from city (city is selling)
    buy_price: Optional[int] = None   # Price city pays YOU when you SELL to them (city is buying)
    maximum_mt: int = 0
    sell_locked: bool = False
    buy_locked: bool = False
    
    @property
    def sell_capacity(self) -> int: 
        """MT available for YOU to BUY from this city"""
        return max(self.quantity_mt - self.reserve_mt, 0)

    @property
    def buy_capacity(self) -> int:
        """MT this city will BUY from YOU"""
        return max(self.maximum_mt - self.quantity_mt, 0)

@dataclass
class CitySnapshot:
    city_name: str
    commodities: list[CommodityRow] = field(default_factory=list)
    footer: Optional[any] = None

def clean_ocr_number(text: str) -> int:
    """Limpia el ruido del OCR y maneja correctamente separadores de miles y errores como ',.' """
    if not text: return 0
    # Eliminar espacios y carácteres no numéricos (excepto separadores)
    clean_str = re.sub(r'[^\d,.]', '', text)
    # Si detecta el patrón de error ',.' o simplemente miles, eliminamos los separadores
    clean_str = clean_str.replace(',.', '').replace(',', '').replace('.', '')
    
    try:
        return int(clean_str) if clean_str else 0
    except ValueError:
        return 0


def save_debug_image(image_obj, commodity_name, field_type):
    """
    image_obj: El recorte de la imagen (objeto PIL o OpenCV)
    commodity_name: Nombre del producto (ej. 'Rare/Precious')
    field_type: Qué estamos leyendo (ej. 'Qty' o 'Price')
    """
    # LIMPIEZA CRÍTICA: Reemplazamos / y espacios por guiones bajos
    clean_name = commodity_name.replace("/", "_").replace(" ", "_").lower()
    
    filename = f"{clean_name}_{field_type}.png"
    filepath = os.path.join(debug_folder, filename)
    
    # Si usas PIL (como parece en tu traceback)
    image_obj.save(filepath)
    
    # logger.debug(f"DEBUG: Imagen guardada en {filepath}") # Opcional
    return filepath

def _is_red(img_cv, x, y, w, h):
    """Detecta si el precio está en rojo (bloqueado)."""
    # Extraemos la región de interés (el precio)
    roi = img_cv[max(0, y-2):y+h+2, max(0, x-2):x+w+2]
    if roi.size == 0: return False
    
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    # Rangos de rojo en el espectro HSV
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([170, 70, 50])
    upper_red2 = np.array([180, 255, 255])
    
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    mask = cv2.bitwise_or(mask1, mask2)
    
    # Si más del 10% de los píxeles son rojos, asumimos que está bloqueado
    pixel_ratio = np.count_nonzero(mask) / (roi.shape[0] * roi.shape[1])
    return pixel_ratio > 0.1

def preprocess_for_ocr(img_cv):
    """Preprocesamiento ultra-sensible para categorías difíciles."""
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    
    # Aumentamos el contraste antes de binarizar
    enhanced = cv2.convertScaleAbs(gray, alpha=1.5, beta=0) 
    
    # Binarización más permisiva (umbral 150 en lugar de adaptativo extremo)
    _, thresh = cv2.threshold(enhanced, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Invertir si el fondo es claro (opcional, según tu UI)
    # thresh = cv2.bitwise_not(thresh) 
    
    return thresh

def extract_number_from_region(img_processed, img_original, x1, y1, x2, y2, region_name=""):
    """
    CRITICAL FIX: Detecta si está bloqueado y aplica un umbral en el canal rojo
    para que Tesseract no devuelva None.
    """
    h, w = img_original.shape[:2]
    
    # Reducimos el área de lectura 2px para no tocar los bordes
    x1, y1, x2, y2 = x1+2, y1+2, x2-2, y2-2
    x1, y1, x2, y2 = max(0, x1), max(0, y1), min(w, x2), min(h, y2)
    
    # Usamos siempre la imagen a color original para esta zona
    roi_color = img_original[y1:y2, x1:x2]
    
    if roi_color.size == 0: return None, False

    # 1. Detección de color rojo (Bloqueado)
    hsv = cv2.cvtColor(roi_color, cv2.COLOR_BGR2HSV)
    mask = cv2.bitwise_or(cv2.inRange(hsv, (0, 70, 50), (10, 255, 255)), 
                          cv2.inRange(hsv, (170, 70, 50), (180, 255, 255)))
    is_locked = (np.count_nonzero(mask) / roi_color.size) > 0.05

    # 2. Preprocesamiento adaptativo según el color
    if is_locked:
        # Si es rojo, separamos canales y usamos el Rojo (r) para máximo contraste
        b, g, r = cv2.split(roi_color)
        _, roi_ready = cv2.threshold(r, 150, 255, cv2.THRESH_BINARY_INV)
    else:
        # Si es normal, pasamos a gris y binarizamos
        gray = cv2.cvtColor(roi_color, cv2.COLOR_BGR2GRAY)
        enhanced = cv2.convertScaleAbs(gray, alpha=1.5, beta=0) 
        _, roi_ready = cv2.threshold(enhanced, 150, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # 3. Intentamos leer con OCR
    configs = ['--psm 7', '--psm 6']
    text = ""
    
    for cfg in configs:
        full_cfg = f'{cfg} -c tessedit_char_whitelist=0123456789.,'
        text = pytesseract.image_to_string(roi_ready, config=full_cfg).strip()
        if text: break # Si encontró algo, paramos

    num = clean_ocr_number(text)
    
    # Devuelve el número (int) para no romper las matemáticas y el estado locked
    return num if num > 0 else None, is_locked

def parse_city_image_fixed_regions(img_cv, city_name: str) -> CitySnapshot:
    """
    Parser con coordenadas FIJAS calibradas para layout 1100x630.
    FIXED: Now correctly returns values even when locked.
    """
    snapshot = CitySnapshot(city_name=city_name)
    h, w = img_cv.shape[:2]

    logger.info(f"\n{'='*60}")
    logger.info(f"Procesando: {city_name} (Ajuste de precisión)")
    logger.info(f"Tamaño: {w}x{h}")
    logger.info(f"{'='*60}")

    # =========================
    # 📐 CONFIG RE-CALIBRADA (PIXELS)
    # =========================
    COLS = {
        "quantity": 332,   # Estable
        "reserve":  449,   # Estable
        "sell":     554,   # Estable (precio que pagas TÚ para COMPRAR de la ciudad)
        "buy":      663,   # Ajustado +1/3 derecha (precio que la ciudad te PAGA cuando vendes)
        "maximum":  788,   # Ajustado +1/5 derecha
    }

    # Ajuste vertical
    ROWS_Y = [
        ("Rare/Precious",        92), 
        ("Foodstuffs",           173), 
        ("Natural Materials",    246),
        ("Fuel Ore",             321), 
        ("Consumer Goods",       385),
        ("Fabricated Material",  457), 
        ("Refined Fuel",         532), 
    ]
    CELL_WIDTH = 110   
    CELL_HEIGHT = 48
    
    def get_region(x, y):
        return (
            int(x - CELL_WIDTH // 2),
            int(y - CELL_HEIGHT // 2),
            int(x + CELL_WIDTH // 2),
            int(y + CELL_HEIGHT // 2),
        )

    # =========================
    # 🔍 OCR LOOP
    # =========================
    for commodity, row_y in ROWS_Y:
        logger.info(f"\nProcesando commodity: {commodity} (Y={row_y})")

        # 1. Quantity MT
        x1, y1, x2, y2 = get_region(COLS["quantity"], row_y)
        crop_qty = Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)).crop((x1, y1, x2, y2))
        save_debug_image(crop_qty, commodity, "Qty")
        quantity_mt, _ = extract_number_from_region(img_cv, img_cv, x1, y1, x2, y2, f"{commodity} - Quantity")

        # 2. Reserve MT
        x1, y1, x2, y2 = get_region(COLS["reserve"], row_y)
        save_debug_image(Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)).crop((x1, y1, x2, y2)), commodity, "Reserve")
        reserve_mt, _ = extract_number_from_region(img_cv, img_cv, x1, y1, x2, y2, f"{commodity} - Reserve")

        # 3. Sell Price (what YOU pay to BUY from city)
        x1, y1, x2, y2 = get_region(COLS["sell"], row_y)
        save_debug_image(Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)).crop((x1, y1, x2, y2)), commodity, "Sell")
        sell_price, sell_locked = extract_number_from_region(img_cv, img_cv, x1, y1, x2, y2, f"{commodity} - Sell Price")

        # 4. Buy Price (what city PAYS YOU when you sell)
        x1, y1, x2, y2 = get_region(COLS["buy"], row_y)
        save_debug_image(Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)).crop((x1, y1, x2, y2)), commodity, "Buy")
        buy_price, buy_locked = extract_number_from_region(img_cv, img_cv, x1, y1, x2, y2, f"{commodity} - Buy Price")

        # 5. Maximum MT
        x1, y1, x2, y2 = get_region(COLS["maximum"], row_y)
        save_debug_image(Image.fromarray(cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)).crop((x1, y1, x2, y2)), commodity, "Max")
        maximum_mt, _ = extract_number_from_region(img_cv, img_cv, x1, y1, x2, y2, f"{commodity} - Maximum")

        # --- DATA ASSEMBLY ---
        row = CommodityRow(
            commodity=commodity,
            quantity_mt=int(quantity_mt) if quantity_mt else 0,
            reserve_mt=int(reserve_mt) if reserve_mt else 0,
            sell_price=int(sell_price) if sell_price else None,  # Keep value even if locked
            buy_price=int(buy_price) if buy_price else None,     # Keep value even if locked
            maximum_mt=int(maximum_mt) if maximum_mt else 0,
            sell_locked=sell_locked,
            buy_locked=buy_locked
        )
        snapshot.commodities.append(row)

        logger.info(
            f"  OK {commodity}: "
            f"Qty={row.quantity_mt}, "
            f"Sell={row.sell_price}{'🔒' if sell_locked else ''}, "
            f"Buy={row.buy_price}{'🔒' if buy_locked else ''}, "
            f"Max={row.maximum_mt}"
        )

    return snapshot

def parse_city_image(img_cv, city_name: str) -> CitySnapshot:
    """Función principal que usa el parser mejorado de regiones fijas."""
    return parse_city_image_fixed_regions(img_cv, city_name)

def parse_easydock_image(image_path: str, location_name: str) -> list[dict]:
    """
    EasyDock parser — commodity names not fixed-order; uses text prefix + positional columns.
    Returns raw row dicts for backward compatibility with build_trade_catalog.
    """
    text = pytesseract.image_to_string(Image.open(image_path)).strip()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rows = []

    DOCK_COL_MT = 0; DOCK_COL_BUY_MT = 1; DOCK_COL_BUY_CR = 2
    DOCK_COL_SELL_MT = 3; DOCK_COL_SELL_CR = 4

    def _extract_numbers(line):
        return [int(n.replace(",","")) for n in re.findall(r"[\d,]+", line)
                if n.replace(",","").isdigit()]

    for line in lines:
        low = line.lower()
        if any(kw in low for kw in _SKIP_KEYWORDS):
            continue
        nums = _extract_numbers(line)
        if len(nums) < 2:
            continue
        first_num = re.search(r"[\d,]+", line)
        if not first_num:
            continue
        name = line[:first_num.start()].strip()
        if not name:
            continue

        def _n(idx):
            return nums[idx] if len(nums) > idx else 0

        sell_cr = _n(DOCK_COL_SELL_CR)
        buy_cr  = _n(DOCK_COL_BUY_CR)
        if sell_cr == 0 and buy_cr == 0:
            continue

        rows.append({
            "location":   location_name,
            "category":   _infer_category(name),
            "name":       name,
            "mt":         _n(DOCK_COL_MT),
            "buying_mt":  _n(DOCK_COL_BUY_MT),
            "buying_cr":  buy_cr,
            "selling_mt": _n(DOCK_COL_SELL_MT),
            "selling_cr": sell_cr,
        })
    return rows


def _infer_category(raw: str) -> str:
    text = re.sub(r"^[^A-Za-z]*", "", str(raw or "")).strip()
    for cat in COMMODITY_CATEGORIES:
        if cat.lower() in text.lower():
            return cat
    return text

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 2 — CATALOG BUILDER  (snapshots → unified price dict)
# ─────────────────────────────────────────────────────────────────────────────

def build_trade_catalog(
    city_snapshots: list[CitySnapshot],
    easydock_rows:  list[dict],
) -> dict[str, dict[str, dict]]:
    """
    CRITICAL FIX: Properly labels buy/sell prices:
    - sell_price: What YOU pay to BUY from city (city selling price)
    - buy_price: What city PAYS YOU when you sell (city buying price)
    
    Returns:
        catalog[city_name][commodity] = {
            "sell_price":    int | None,  # Price to BUY from city
            "buy_price":     int | None,  # Price city pays when you SELL
            "sell_capacity": int,         # MT available for purchase
            "buy_capacity":  int,         # MT city will accept
        }

    Trade logic:
        Source city  → use sell_price (you BUY from source)
        Destination  → use buy_price  (you SELL to destination)
        Profit/MT    = destination.buy_price - source.sell_price
    """
    catalog: dict = {}

    # ── City screenshots ──────────────────────────────────────────────────────
    for snap in city_snapshots:
        city = catalog.setdefault(snap.city_name, {})
        for row in snap.commodities:
            city[row.commodity] = {
                "sell_price": row.sell_price,      # What YOU pay to BUY
                "buy_price":  row.buy_price,       # What city PAYS YOU
                "sell_capacity": row.sell_capacity, # MT you can buy
                "buy_capacity":  row.buy_capacity,  # MT city will buy
                "sell_locked": row.sell_locked,
                "buy_locked": row.buy_locked
            }

    # ── EasyDock rows ─────────────────────────────────────────────────────────
    for row in easydock_rows:
        location  = row["location"]
        commodity = row["category"]
        city = catalog.setdefault(location, {})
        city[commodity] = {
            "sell_price":    row.get("selling_cr") or None,
            "buy_price":     row.get("buying_cr")  or None,
            "sell_capacity": row.get("selling_mt", 0),
            "buy_capacity":  row.get("buying_mt",  0),
        }

    return catalog

# ─────────────────────────────────────────────────────────────────────────────
# FLIGHT TIME
# ─────────────────────────────────────────────────────────────────────────────

def _haversine(lat1, lon1, lat2, lon2) -> float:
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a  = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return PLANET_RADIUS * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def _pixel_distance(a: str, b: str) -> float:
    ca, cb = CITIES.get(a), CITIES.get(b)
    if not ca or not cb:
        return 0.0
    return math.sqrt((ca["x"]-cb["x"])**2 + (ca["y"]-cb["y"])**2)


# Asegúrate de tener estas constantes definidas arriba
SPEED_UNITS_PER_MIN = 5.0 # Ajusta según el juego

def get_flight_time(origin: str, destination: str) -> float:
    if origin == destination:
        return 0.0
    
    # Intentar obtener por coordenadas de píxeles (que sí tienes en CITIES)
    dist = _pixel_distance(origin, destination)
    if dist > 0:
        # Fórmula: Tiempo base + (distancia / velocidad) + aterrizaje
        return round(TAKEOFF_TIME + dist / SPEED_UNITS_PER_MIN + LANDING_TIME, 1)
    
    # Valor por defecto si no encuentra la ciudad
    return 120.0

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 3 — OPTIMIZER
# ─────────────────────────────────────────────────────────────────────────────

def find_trade_opportunities(catalog, ship_capacity, budget=None):
    """
    CRITICAL FIX: Correct profit calculation
    Profit = destination.buy_price - source.sell_price
    (You BUY at source, SELL at destination)
    """
    opps = []
    for s_name, s_data in catalog.items():
        for d_name, d_data in catalog.items():
            if s_name == d_name: continue
            for comm, s_info in s_data.items():
                if comm not in d_data: continue
                d_info = d_data[comm]
                
                # CRITICAL FIX: Correct price assignment
                source_sell = s_info["sell_price"]  # What you PAY to buy from source
                dest_buy = d_info["buy_price"]      # What destination PAYS you
                
                if source_sell is None or source_sell <= 0:
                    continue
                if dest_buy is None or dest_buy <= 0:
                    continue
                if dest_buy <= source_sell:  # No profit
                    continue
                
                profit_mt = dest_buy - source_sell  # FIXED: destination pays - source cost
                avail = min(s_info["sell_capacity"], d_info["buy_capacity"], ship_capacity)
                
                if budget: 
                    avail = min(avail, budget // source_sell)
                
                if avail > 0:
                    opps.append({
                        "commodity": comm, 
                        "source": s_name, 
                        "destination": d_name,
                        "source_selling": source_sell,      # You BUY at this price
                        "destination_buying": dest_buy,     # Destination PAYS this
                        "profit_per_mt": profit_mt,
                        "source_available": s_info["sell_capacity"],
                        "destination_capacity": d_info["buy_capacity"],
                        "travel_time": round(TAKEOFF_TIME + 50 / SPEED_UNITS_PER_MIN + LANDING_TIME, 1)
                    })
    return opps

def assign_grades(opportunities, ship_capacity, budget=None, is_rental=False, rental_cost_per_day=None):
    for op in opportunities:
        qty = min(ship_capacity, op["source_available"], op["destination_capacity"])
        if budget: qty = min(qty, budget // op["source_selling"])
        
        op["_qty_trip"] = qty
        op["_cost_trip"] = qty * op["source_selling"]
        op["_profit_trip"] = qty * op["profit_per_mt"]
        op["total_profit"] = op["_profit_trip"] 
        op["_roi"] = (op["_profit_trip"] / op["_cost_trip"] * 100) if op["_cost_trip"] > 0 else 0
        op["grade"] = "D"
    
    opportunities.sort(key=lambda x: x["_profit_trip"], reverse=True)
    for i, op in enumerate([o for o in opportunities if o["_profit_trip"] > 0]):
        pct = i / len(opportunities)
        op["grade"] = "A" if pct < 0.2 else "B" if pct < 0.5 else "C"
    return opportunities


def compute_trade_routes(catalog, origin, allowed_commodities,
                         ship_capacity, budget, max_hops=5, top_n=20) -> list:
    """
    Multi-hop chained route optimizer.
    FIXED: Uses correct buy/sell price logic.
    """
    locations = [loc for loc in catalog if loc != origin]
    all_routes, seen = [], set()

    def _best_leg(src, dst):
        travel_time = get_flight_time(src, dst)
        best = None
        for commodity in allowed_commodities:
            s_data = catalog.get(src, {}).get(commodity)
            d_data = catalog.get(dst, {}).get(commodity)
            if not s_data or not d_data:
                continue
            
            # CRITICAL FIX: Correct price extraction
            source_sell = s_data.get("sell_price")  # What you pay to BUY
            dest_buy = d_data.get("buy_price")      # What dest PAYS you
            
            if not source_sell or source_sell <= 0: continue
            if not dest_buy or dest_buy <= 0: continue
            
            ppm = dest_buy - source_sell  # FIXED: Correct profit calc
            if ppm <= 0: continue
            
            avail = s_data.get("sell_capacity", 0)
            cap   = d_data.get("buy_capacity",  0)
            if avail <= 0 or cap <= 0: continue
            
            max_by_budget = int(budget / source_sell) if budget and source_sell > 0 else float("inf")
            qty    = max(min(ship_capacity, avail, cap, max_by_budget), 0)
            if qty <= 0: continue
            
            cost   = qty * source_sell
            profit = qty * ppm
            if best is None or profit > best["profit"]:
                best = dict(commodity=commodity, src=src, dst=dst,
                            buy_price=source_sell,   # You BUY at source
                            sell_price=dest_buy,     # You SELL at destination
                            profit_per_mt=ppm,
                            qty=qty, cost=cost, profit=profit,
                            roi=(profit/cost*100) if cost else 0,
                            flight_min=travel_time)
                            
        return best or dict(commodity="— empty —", src=src, dst=dst,
                            buy_price=0, sell_price=0, profit_per_mt=0,
                            qty=0, cost=0, profit=0, roi=0, flight_min=travel_time)

    for n_stops in range(1, max_hops+1):
        for stops in itertools.permutations(locations, n_stops):
            seq = [origin] + list(stops) + [origin]
            legs, total_profit, total_cost, total_time = [], 0, 0, 0
            remaining, valid = budget, True
            for i in range(len(seq)-1):
                leg = _best_leg(seq[i], seq[i+1])
                if leg["profit"] == 0 and i < len(seq)-2:
                    valid = False; break
                legs.append(leg)
                total_profit += leg["profit"]
                total_cost   += leg["cost"]
                total_time   += leg["flight_min"]
                if remaining is not None:
                    remaining = remaining - leg["cost"] + leg["profit"]
            if not valid or total_profit <= 0:
                continue
            roi         = (total_profit / total_cost * 100) if total_cost else 0
            cr_per_hour = (total_profit / total_time * 60) if total_time else 0
            key = tuple(seq)
            if key not in seen:
                seen.add(key)
                all_routes.append(dict(stops=seq, legs=legs, n_hops=len(legs),
                                       total_profit=total_profit, total_cost=total_cost,
                                       total_time=total_time, roi=roi, cr_per_hour=cr_per_hour))
            if len(all_routes) >= top_n * 5:
                break
    return sorted(all_routes, key=lambda x: x["total_profit"], reverse=True)[:top_n]

# ─────────────────────────────────────────────────────────────────────────────
# MACRO DATA
# ─────────────────────────────────────────────────────────────────────────────

def calculate_macro_data(catalog: dict, opportunities: list) -> dict:
    city_profits, city_commodities, city_lucrative = {}, {}, {}
    for op in opportunities:
        src = op["source"]
        p = op.get("total_profit", 0) 
        
        city_profits.setdefault(src, 0)
        city_commodities.setdefault(src, set())
        city_lucrative.setdefault(src, set())
        
        city_profits[src] += p
        city_commodities[src].add(op["commodity"])
        if p > 0:
            city_lucrative[src].add(op["commodity"])

    city_summary = sorted(
        [{"city": c,
          "total_profit":    city_profits.get(c, 0),
          "num_commodities": len(city_commodities.get(c, set())),
          "num_lucrative":   len(city_lucrative.get(c, set()))}
         for c in catalog],
        key=lambda x: x["total_profit"], reverse=True,
    )

    commodity_best_sellers, commodity_best_buyers = {}, {}
    for commodity in COMMODITY_CATEGORIES:
        sellers, buyers = [], []
        for city, data in catalog.items():
            if commodity not in data:
                continue
            item = data[commodity]
            
            sp = item.get("sell_price")
            bp = item.get("buy_price")
            
            if sp and sp > 0 and item.get("sell_capacity", 0) > 0:
                sellers.append({"city": city, "price": sp, "capacity": item["sell_capacity"]})
            if bp and bp > 0 and item.get("buy_capacity", 0) > 0:
                buyers.append({"city": city, "price": bp, "capacity": item["buy_capacity"]})
        
        commodity_best_sellers[commodity] = sorted(sellers, key=lambda x: x["price"])[:5]
        commodity_best_buyers[commodity]  = sorted(buyers,  key=lambda x: x["price"], reverse=True)[:5]

    return {
        "city_summary":                    city_summary,
        "commodity_best_sellers":          commodity_best_sellers,
        "commodity_best_buyers":           commodity_best_buyers,
        "top_cities_by_profit":            city_summary[:10],
        "cities_by_lucrative_commodities": sorted(city_summary,
                                                   key=lambda x: x["num_lucrative"],
                                                   reverse=True)[:10],
    }

# ─────────────────────────────────────────────────────────────────────────────
# COLOR REGISTRY
# ─────────────────────────────────────────────────────────────────────────────

import hashlib
_CITY_COLOR_MAP:      dict[str, PatternFill] = {}
_COMMODITY_COLOR_MAP: dict[str, PatternFill] = {}
_COLOR_MAPS_FILE = "color_maps.json"

def _hex_from_name(name: str) -> str:
    return hashlib.md5(name.encode()).hexdigest()[:6].upper()

def _get_fill(name: str, registry: dict) -> PatternFill:
    if name not in registry:
        c = _hex_from_name(name)
        registry[name] = PatternFill(start_color=c, end_color=c, fill_type="solid")
    return registry[name]

def get_city_fill(city: str) -> PatternFill:      return _get_fill(city,      _CITY_COLOR_MAP)
def get_commodity_fill(c: str) -> PatternFill:    return _get_fill(c,         _COMMODITY_COLOR_MAP)

def save_color_maps():
    data = {"cities":      {k: v.start_color.rgb for k, v in _CITY_COLOR_MAP.items()},
            "commodities": {k: v.start_color.rgb for k, v in _COMMODITY_COLOR_MAP.items()}}
    with open(_COLOR_MAPS_FILE, "w") as fh:
        json.dump(data, fh)

def load_color_maps():
    try:
        with open(_COLOR_MAPS_FILE) as fh:
            data = json.load(fh)
        for k, v in data.get("cities",      {}).items():
            _CITY_COLOR_MAP[k]      = PatternFill(start_color=v, end_color=v, fill_type="solid")
        for k, v in data.get("commodities", {}).items():
            _COMMODITY_COLOR_MAP[k] = PatternFill(start_color=v, end_color=v, fill_type="solid")
    except (FileNotFoundError, json.JSONDecodeError):
        pass

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — EXCEL STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _contrast(hex_color: str) -> str:
    hx = hex_color.strip().lstrip("#").ljust(6, "0")[:6]
    try:
        r, g, b = int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16)
    except ValueError:
        return "000000"
    return "000000" if (r*299 + g*587 + b*114)/1000 > 160 else "FFFFFF"

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(cells):
    for cell in cells:
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

def _style_section_title(cells):
    for cell in cells:
        cell.font      = Font(bold=True, size=14)
        cell.fill      = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

def _style_subsection_title(cells, color="E6E6FA"):
    for cell in cells:
        cell.font      = Font(bold=True, size=12)
        cell.fill      = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = _thin_border()

def _style_commodity_separator(cells, color):
    thick = Side(style="thick")
    for cell in cells:
        cell.fill   = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=thick, bottom=thick)

def _auto_size_columns(ws, min_width=10, max_width=50):
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        if isinstance(col_cells[0], MergedCell):
            continue
        max_len = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_width, max(min_width, max_len+2))

def _apply_roi_color(cell, roi_pct: float):
    if roi_pct > 100:   fill, fc = "00B050", "FFFFFF"
    elif roi_pct >= 50: fill, fc = "70AD47", "FFFFFF"
    elif roi_pct >= 20: fill, fc = "FFD700", "000000"
    elif roi_pct >= 0:  fill, fc = "FFA500", "FFFFFF"
    else:               fill, fc = "FF0000", "FFFFFF"
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.font = Font(color=fc, bold=True)

def sanitize_sheet_name(name: str) -> str:
    return "".join("-" if ch in r'\/?*[]:"' else ch for ch in name)[:31]

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — SHEET WRITERS
# ─────────────────────────────────────────────────────────────────────────────

def _write_cities_sheet(wb, city_snapshots):
    ws = wb.create_sheet("Cities")

    headers = ["City", "Commodity", "Quantity", "Reserve", "Sell Price (You Pay)", "Buy Price (City Pays)", "Max"]
    ws.append(headers)
    _style_header_row(ws[1])

    for snap in city_snapshots:
        city_name = snap.city_name
        city_dict = {c.commodity: c for c in snap.commodities}

        for commodity in COMMODITY_ORDER:
            if commodity not in city_dict:
                continue

            row = city_dict[commodity]

            # Formateamos los precios para añadir (Locked) solo en la vista de Excel
            display_sell = f"{row.sell_price} (Locked)" if row.sell_locked and row.sell_price else row.sell_price
            display_buy = f"{row.buy_price} (Locked)" if row.buy_locked and row.buy_price else row.buy_price

            ws.append([
                city_name,
                commodity,
                row.quantity_mt,
                row.reserve_mt,
                display_sell,
                display_buy,
                row.maximum_mt,
            ])

    _auto_size_columns(ws)


def _write_opportunities_sheet(ws, opportunities: list[dict],
                                is_rental: bool, rental_cost_per_day: Optional[int],
                                ship_capacity: int):
    """
    FIXED: Column headers now clearly indicate the correct buy/sell direction
    """
    cols = [
        "Grade", "Commodity",
        "Source City", "Buy From Source (CR/MT)",    # What YOU pay to buy
        "Destination City", "Sell To Dest (CR/MT)",  # What destination pays YOU
        "Profit/MT (CR)",
        "Src Stock (MT)", "Dst Capacity (MT)",
        "MT Loaded", "Trip Cost (CR)", "Trip Profit (CR)", "Trip ROI (%)",
        "Travel Time (min)",
    ]
    if is_rental and rental_cost_per_day:
        cols += ["Rental Cost (CR)", "Covers Rental?"]

    ws.append(cols)
    _style_header_row(ws[1])

    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    roi_col    = cols.index("Trip ROI (%)") + 1

    for op in opportunities:
        grade = op.get("grade", "D")
        row   = [
            grade,
            op["commodity"],
            op["source"],
            op["source_selling"],      # You BUY at this price
            op["destination"],
            op["destination_buying"],  # Destination PAYS you this
            op["profit_per_mt"],
            op["source_available"],
            op["destination_capacity"],
            op.get("_qty_trip", 0),
            op.get("_cost_trip", 0),
            op.get("_profit_trip", 0),
            round(op.get("_roi", 0), 2),
            op["travel_time"],
        ]
        if is_rental and rental_cost_per_day:
            covers = op.get("_profit_trip", 0) > rental_cost_per_day
            row += [rental_cost_per_day, "Yes" if covers else "No"]

        ws.append(row)
        cur = ws.max_row

        # Grade cell
        gs = GRADE_STYLES[grade]
        gc = ws.cell(row=cur, column=1)
        gc.fill = PatternFill(start_color=gs["fill"], end_color=gs["fill"], fill_type="solid")
        gc.font = Font(color=gs["font"], bold=True)
        gc.alignment = Alignment(horizontal="center")

        # Commodity
        cf = get_commodity_fill(op["commodity"])
        cc = ws.cell(row=cur, column=2)
        cc.fill = cf
        cc.font = Font(color=_contrast(cf.start_color.rgb), bold=True)

        # Source city
        sf = get_city_fill(op["source"])
        sc = ws.cell(row=cur, column=3)
        sc.fill = sf
        sc.font = Font(color=_contrast(sf.start_color.rgb), bold=True)

        # Destination city
        df = get_city_fill(op["destination"])
        dc = ws.cell(row=cur, column=5)
        dc.fill = df
        dc.font = Font(color=_contrast(df.start_color.rgb), bold=True)

        _apply_roi_color(ws.cell(row=cur, column=roi_col), op.get("_roi", 0))

        if is_rental and rental_cost_per_day:
            rc = ws.cell(row=cur, column=len(cols))
            rc.fill = green_fill if rc.value == "Yes" else red_fill
            rc.font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "B2"
    _auto_size_columns(ws)


def _write_trade_routes_sheet(ws, routes: list, origin: str):
    alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    if not routes:
        ws.append(["No profitable routes found from", origin])
        return
    for rank, route in enumerate(routes, 1):
        summary = (f"Route #{rank}  |  {' → '.join(route['stops'])}  |  "
                   f"Profit: {route['total_profit']:,.0f} CR  |  ROI: {route['roi']:.1f}%  |  "
                   f"Time: {route['total_time']} min  |  Efficiency: {route['cr_per_hour']:,.0f} CR/h")
        ws.append([summary])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        c = ws.cell(row=r, column=1)
        c.font      = Font(bold=True, size=12, color="FFFFFF")
        c.fill      = get_city_fill(origin)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.append(["Hop", "From", "To", "Commodity", "Buy (CR/MT)", "Sell (CR/MT)",
                   "Profit/MT", "MT", "Leg Cost (CR)", "Leg Profit (CR)", "Flight (min)", "Leg ROI (%)"])
        _style_header_row(ws[ws.max_row])
        for hop_i, leg in enumerate(route["legs"], 1):
            ws.append([hop_i, leg["src"], leg["dst"], leg["commodity"],
                       leg["buy_price"], leg["sell_price"], leg["profit_per_mt"],
                       leg["qty"], leg["cost"], leg["profit"], leg["flight_min"], round(leg["roi"], 2)])
            cur = ws.max_row
            for col, city_key in ((2, leg["src"]), (3, leg["dst"])):
                fill = get_city_fill(city_key)
                cell = ws.cell(row=cur, column=col)
                cell.fill = fill
                cell.font = Font(color=_contrast(fill.start_color.rgb), bold=True)
            _apply_roi_color(ws.cell(row=cur, column=12), leg["roi"])
            if hop_i % 2 == 0:
                for col in (1, 4, 5, 6, 7, 8, 9, 10, 11):
                    ws.cell(row=cur, column=col).fill = alt_fill
        ws.append(["", "", "", "TOTAL", "", "", "", "",
                   route["total_cost"], route["total_profit"], route["total_time"], round(route["roi"], 2)])
        tot = ws.max_row
        for col in range(1, 13):
            ws.cell(row=tot, column=col).font = Font(bold=True)
            ws.cell(row=tot, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _apply_roi_color(ws.cell(row=tot, column=12), route["roi"])
        ws.append([])
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def _write_macro_sheet(wb, macro_data: dict):
    alt2 = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    ws   = wb.create_sheet("MACRO")

    def _section(title):
        ws.append([title])
        _style_section_title(ws[ws.max_row])

    _section("SUMMARY BY CITY")
    ws.append(["City", "Total Potential Profit (CR)", "Num Commodities", "Num Lucrative"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["city_summary"], start=ws.max_row+1):
        ws.append([item["city"], item["total_profit"], item["num_commodities"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("TOP 10 CITIES BY TOTAL PROFIT")
    ws.append(["City", "Total Profit (CR)", "Num Lucrative"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["top_cities_by_profit"], start=ws.max_row+1):
        ws.append([item["city"], item["total_profit"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("TOP 10 CITIES BY LUCRATIVE COMMODITIES")
    ws.append(["City", "Num Lucrative", "Total Profit (CR)"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["cities_by_lucrative_commodities"], start=ws.max_row+1):
        ws.append([item["city"], item["num_lucrative"], item["total_profit"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("BEST SELLERS AND BUYERS BY COMMODITY")
    for commodity in COMMODITY_CATEGORIES:
        color_fill = get_commodity_fill(commodity)
        color = color_fill.start_color.rgb if color_fill.start_color else "E6E6FA"
        ws.append([])
        ws.append([f"--- {commodity.upper()} ---"])
        _style_commodity_separator(ws[ws.max_row], color)
        for section_label, key in (
            ("Best Sellers (lowest price — where to buy)",  "commodity_best_sellers"),
            ("Best Buyers  (highest price — where to sell)", "commodity_best_buyers"),
        ):
            ws.append([])
            ws.append([f"{commodity} — {section_label}"])
            _style_subsection_title(ws[ws.max_row], color)
            ws.append(["City",
                       "Selling Price (CR/MT)" if "Seller" in section_label else "Buying Price (CR/MT)",
                       "Capacity (MT)"])
            _style_header_row(ws[ws.max_row])
            for i, entry in enumerate(macro_data[key].get(commodity, []), start=ws.max_row+1):
                ws.append([entry["city"], entry["price"], entry["capacity"]])
                if i % 2 == 0:
                    for cell in ws[i]: cell.fill = alt2

    _auto_size_columns(ws)

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — MAIN SAVE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def save_to_excel(
    city_snapshots:      list[CitySnapshot],
    easydock_rows:       list[dict],
    opportunities:       list[dict],
    catalog:             dict,
    selected_ship:       str,
    ship_capacity:       int,
    output_file:         str,
    is_rental:           bool,
    rental_cost_per_day: Optional[int],
    origin:              str,
    budget:              Optional[int],
    containers_used:     Optional[int],
    mode:                str,
    trade_route_params:  Optional[dict],
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Config ────────────────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Parameter", "Value"])
    _style_header_row(ws_cfg[1])
    ship_data, ship_cat = None, ""
    for cat, ships in SHIPS.items():
        if selected_ship in ships:
            ship_data, ship_cat = ships[selected_ship], cat
            break
    max_c   = ship_data["max_containers"] if ship_data else 0
    c_used  = containers_used if containers_used is not None else max_c
    c_mt    = ship_data["container_mt"] if ship_data else 17
    cargo_b = ship_data["cargo_base"]   if ship_data else 0
    config_rows = [
        ("Ship",                  selected_ship),
        ("Category",              ship_cat),
        ("Base cargo (fixed MT)", cargo_b),
        ("Containers carried",    f"{c_used} / {max_c} max"),
        ("MT per container",      c_mt),
        ("Total Capacity (MT)",   f"{cargo_b} + {c_used} x {c_mt} = {ship_capacity}"),
        ("Origin",                origin),
        ("OS",                    platform.system()),
        ("Status",                "Rented" if is_rental else "Purchased"),
    ]
    if is_rental and rental_cost_per_day:
        config_rows += [
            ("Rental cost/day",                f"{rental_cost_per_day} CR/day (14 h)"),
            ("Min profit/day to cover rental", f"{rental_cost_per_day} CR"),
            ("Min profit/hour",                f"{rental_cost_per_day/14:.2f} CR/h"),
        ]
    if budget:
        config_rows.append(("Initial Budget", f"{budget} CR"))
    if trade_route_params:
        config_rows += [
            ("Mode",               "Trade Route"),
            ("Max hops",           trade_route_params.get("max_hops", 5)),
            ("Allowed commodities", ", ".join(trade_route_params.get("allowed_commodities", ["ALL"]))),
        ]
    for row in config_rows:
        ws_cfg.append(list(row))
    _auto_size_columns(ws_cfg)

    # ── Cities ────────────────────────────────────────────────────────────────
    if city_snapshots:
        _write_cities_sheet(wb, city_snapshots)

    # ── EasyDock ──────────────────────────────────────────────────────────────
    if easydock_rows:
        ws_ed = wb.create_sheet("EasyDock")
        ws_ed.append(["Location", "Category", "Name", "MT",
                      "Buying MT", "Buying CR", "Selling MT", "Selling CR"])
        _style_header_row(ws_ed[1])
        alt = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        for i, row in enumerate(easydock_rows, start=2):
            ws_ed.append([row["location"], row["category"], row["name"],
                          row["mt"], row["buying_mt"], row["buying_cr"],
                          row["selling_mt"], row["selling_cr"]])
            if i % 2 == 0:
                for cell in ws_ed[i]: cell.fill = alt
        _auto_size_columns(ws_ed)

    # ── Global Opportunities (all cities) ─────────────────────────────────────
    if opportunities:
        ws_op = wb.create_sheet("Opportunities")
        _write_opportunities_sheet(ws_op, opportunities, is_rental, rental_cost_per_day, ship_capacity)

        # City-mode filtered sub-sheet
        if mode == "city" and origin:
            city_opps = [o for o in opportunities if o["source"] == origin]
            if city_opps:
                ws_city = wb.create_sheet(sanitize_sheet_name(f"From {origin}"))
                _write_opportunities_sheet(ws_city, city_opps, is_rental,
                                           rental_cost_per_day, ship_capacity)

    # ── Trade Routes ──────────────────────────────────────────────────────────
    if mode == "route" and trade_route_params:
        allowed  = trade_route_params.get("allowed_commodities", COMMODITY_CATEGORIES)
        max_hops = trade_route_params.get("max_hops", 3)
        print(f"\n  Computing trade routes from {origin} (max {max_hops} hops)…")
        routes = compute_trade_routes(catalog, origin, allowed, ship_capacity, budget, max_hops, top_n=20)
        ws_routes = wb.create_sheet(sanitize_sheet_name(f"Trade Routes {origin}"))
        _write_trade_routes_sheet(ws_routes, routes, origin)
        print(f"  Found {len(routes)} profitable routes.")
        if routes:
            best = routes[0]
            print(f"  Best: {' → '.join(best['stops'])} | "
                  f"Profit: {best['total_profit']:,.0f} CR | ROI: {best['roi']:.1f}%")

    # ── MACRO ─────────────────────────────────────────────────────────────────
    macro_data = calculate_macro_data(catalog, opportunities or [])
    _write_macro_sheet(wb, macro_data)

    # ── Save ──────────────────────────────────────────────────────────────────
    save_color_maps()
    try:
        wb.save(output_file)
        print(f"\n✔  Saved: {output_file}")
        print(f"   Ship: {selected_ship} | Capacity: {ship_capacity} MT | Origin: {origin}")
    except PermissionError:
        ts = int(time.time())
        base, ext = os.path.splitext(output_file)
        alt_path  = f"{base}_{ts}{ext}"
        wb.save(alt_path)
        print(f"⚠  '{output_file}' was in use — saved as '{alt_path}'")

# ─────────────────────────────────────────────────────────────────────────────
# PROMPT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _prompt_city(prompt_text: str) -> str:
    print(f"\n{prompt_text}")
    for i, city in enumerate(CITY_LIST, 1):
        print(f"  {i:>2}. {city}")
    while True:
        try:
            choice = int(input("City number: ").strip())
            if 1 <= choice <= len(CITY_LIST):
                return CITY_LIST[choice - 1]
        except ValueError:
            pass
        print(f"  Enter a number between 1 and {len(CITY_LIST)}.")


def _prompt_commodities() -> list:
    print("\n  Select commodities to trade (comma-separated numbers, or 0 for all):")
    print("   0. ALL commodities")
    for i, cat in enumerate(COMMODITY_CATEGORIES, 1):
        print(f"  {i:>2}. {cat}")
    while True:
        raw = input("  Choice: ").strip()
        if raw == "0":
            return list(COMMODITY_CATEGORIES)
        try:
            indices  = [int(p) for p in raw.split(",") if p.strip()]
            selected = [COMMODITY_CATEGORIES[i-1] for i in indices
                        if 1 <= i <= len(COMMODITY_CATEGORIES)]
            if selected:
                print(f"  Selected: {', '.join(selected)}")
                return selected
        except (ValueError, IndexError):
            pass
        print("  Enter valid comma-separated numbers or 0.")


def _prompt_max_hops() -> int:
    print("\n  Max intermediate stops? (2–5)")
    while True:
        try:
            n = int(input("  Max stops (2-5): ").strip())
            if 2 <= n <= 5:
                return n
        except ValueError:
            pass
        print("  Enter a number between 2 and 5.")

# ─────────────────────────────────────────────────────────────────────────────
# CALIBRATION TOOL
# ─────────────────────────────────────────────────────────────────────────────

def calibrate_regions(image_path: str):
    """
    Herramienta interactiva para calibrar las regiones de extracción OCR.
    Muestra la imagen y permite ajustar las coordenadas visualmente.
    """
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    
    img = cv2.imread(image_path)
    if img is None:
        print(f"Error: No se pudo cargar {image_path}")
        return
    
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    h, w = img.shape[:2]
    
    print(f"\nImagen cargada: {w}x{h} píxeles")
    print("\nEsta herramienta te ayudará a identificar las coordenadas correctas.")
    print("Observa la imagen y las regiones marcadas.\n")
    
    # Coordenadas de ejemplo (ajusta estos valores)
    regions = {
        "Sell Price": {"x_pct": 0.42, "width_pct": 0.08},
        "Buy Price": {"x_pct": 0.54, "width_pct": 0.08},
        "Quantity": {"x_pct": 0.62, "width_pct": 0.08},
        "Maximum": {"x_pct": 0.70, "width_pct": 0.08},
    }
    
    fig, ax = plt.subplots(1, 1, figsize=(15, 10))
    ax.imshow(img_rgb)
    
    # Dibujar rectángulos para cada región
    colors = ['red', 'green', 'blue', 'yellow']
    y_start_pct = 0.25
    row_height_pct = 0.025
    
    for i, commodity in enumerate(COMMODITY_CATEGORIES):
        y = int(h * (y_start_pct + i * row_height_pct * 1.2))
        
        for (name, region), color in zip(regions.items(), colors):
            x = int(w * region['x_pct']) - int(w * region['width_pct'] / 2)
            width = int(w * region['width_pct'])
            height = int(h * row_height_pct)
            
            rect = patches.Rectangle(
                (x, y), width, height,
                linewidth=2, edgecolor=color, facecolor='none'
            )
            ax.add_patch(rect)
            
            if i == 0:  # Solo etiquetar la primera fila
                ax.text(x + width/2, y - 10, name, 
                       color=color, ha='center', fontsize=10, weight='bold')
    
    # Añadir etiquetas de commodities
    for i, commodity in enumerate(COMMODITY_CATEGORIES):
        y = int(h * (y_start_pct + i * row_height_pct * 1.2))
        ax.text(10, y + int(h * row_height_pct / 2), commodity,
               color='white', fontsize=9, weight='bold',
               bbox=dict(boxstyle='round', facecolor='black', alpha=0.7))
    
    plt.title("Regiones de extracción OCR - Verifica si coinciden con los datos", fontsize=14)
    plt.axis('off')
    plt.tight_layout()
    plt.savefig('calibration_preview.png', dpi=150, bbox_inches='tight')
    print("\n✓ Imagen de calibración guardada como 'calibration_preview.png'")
    print("\nRevisa el archivo y ajusta los valores en el código si es necesario:")
    print(f"  - col_sell_price_pct: {regions['Sell Price']['x_pct']}")
    print(f"  - col_buy_price_pct: {regions['Buy Price']['x_pct']}")
    print(f"  - col_quantity_pct: {regions['Quantity']['x_pct']}")
    print(f"  - col_maximum_pct: {regions['Maximum']['x_pct']}")
    print(f"  - start_y_pct: {y_start_pct}")
    print(f"  - row_spacing (cell_height + offset): {row_height_pct}")
    
    plt.show()

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(image_folder, selected_ship=None, output_file="final_trade.xlsx",
         budget=None, containers_used=None, origin=None, mode="regular"):

    print("\n====================================")
    print("  RISE TVP — ship configuration")
    print(f"  OS: {platform.system()}")
    print("====================================")

    # 1. Category
    print("\nCategory:\n  1. AIR AND SPACE\n  2. ONLY AIR")
    while True:
        try:
            choice = int(input("Category number: ").strip())
            if choice in (1, 2):
                category = ["AIR AND SPACE", "ONLY AIR"][choice - 1]; break
        except ValueError:
            pass
        print("  Enter 1 or 2.")

    # 2. Ship model
    ships     = SHIPS[category]
    ship_list = list(ships.keys())
    print(f"\nAvailable ships in {category}:")
    for i, s in enumerate(ship_list, 1):
        d   = ships[s]
        cap = d["cargo_base"] + d["max_containers"] * d["container_mt"]
        print(f"  {i}. {s}  —  {d['cargo_base']} + {d['max_containers']}×{d['container_mt']} MT  (max {cap} MT)")
    while True:
        try:
            choice = int(input("Model number: ").strip())
            if 1 <= choice <= len(ship_list):
                selected_ship = ship_list[choice - 1]; break
        except ValueError:
            pass
        print(f"  Enter 1–{len(ship_list)}.")

    ship_data = ships[selected_ship]
    max_c     = ship_data["max_containers"]

    # 3. Containers
    if containers_used is None:
        if max_c == 0:
            print(f"\n  {selected_ship} has no containers. Fixed: {ship_data['cargo_base']} MT.")
            containers_used = 0
        else:
            print(f"\n  {selected_ship} supports up to {max_c} container(s) × {ship_data['container_mt']} MT each.")
            while True:
                try:
                    n = int(input(f"  Containers today (0–{max_c}): ").strip())
                    if 0 <= n <= max_c:
                        containers_used = n; break
                except ValueError:
                    pass
                print(f"  Enter 0–{max_c}.")

    ship_capacity = ship_data["cargo_base"] + containers_used * ship_data["container_mt"]
    print(f"  Capacity: {ship_data['cargo_base']} + {containers_used}×{ship_data['container_mt']} = {ship_capacity} MT")

    # 4. Rental
    is_rental, ship_rental_cost = False, ship_data.get("rental_cost_per_day")
    print("\n  Rented (A) or Purchased (C)?")
    while True:
        r = input("  Choice: ").strip().upper()
        if r == "A":
            is_rental = True
            if ship_rental_cost is None:
                while True:
                    try:
                        ship_rental_cost = int(input("  Rental cost per day (CR): ").strip()); break
                    except ValueError:
                        print("  Enter a valid number.")
            break
        elif r == "C":
            break
        print("  Enter A or C.")

    # 5. Origin
    if origin is None:
        origin = _prompt_city("SELECT ORIGIN CITY:")

    # 6. Budget
    if budget is None:
        b_input = input("\n  Initial budget CR (Enter to skip): ").strip()
        if b_input:
            try:
                budget = int(b_input)
            except ValueError:
                print("  Invalid — skipped.")
    if budget:
        print(f"  Budget: {budget:,} CR")

    # 7. Trade Route extra prompts
    trade_route_params = None
    if mode == "route":
        trade_route_params = {
            "allowed_commodities": _prompt_commodities(),
            "max_hops":            _prompt_max_hops(),
        }

    # ── Process images ────────────────────────────────────────────────────────
    load_color_maps()

    if not os.path.exists(image_folder):
        print(f"\n✗  Folder '{image_folder}' not found.")
        return

    easydock_rows = []
    city_snapshots_dict: dict[str, CitySnapshot] = {}    
    
    print(f"\n  Scanning folder: {image_folder}...")

    for filename in sorted(os.listdir(image_folder)):
        # 1. Filtro de extensión
        if not filename.lower().endswith((".png", ".jpg", ".jpeg")):
            continue

        # --- NUEVO FILTRO DE PREFIJO ---
        # Ignoramos archivos 'FULL' para evitar ruido en el OCR y priorizamos 'END'
        if filename.upper().startswith("FULL_"):
            logger.info(f"Omitiendo archivo de inventario completo: {filename}")
            continue
        # -------------------------------

        image_path = os.path.join(image_folder, filename)
        img_cv = cv2.imread(image_path)
        
        if img_cv is None:
            continue

        target_city = None
        base_name = filename.upper()
        
        # Lógica simplificada: Detectar ciudad solo si el archivo es END_ o directo
        # Buscamos el patrón "END_NombreCiudad_1.png"
        match = re.search(r"END_([A-Za-z0-9\s]+)_\d+", filename, re.IGNORECASE)
        if match:
            city_candidate = match.group(1).strip()
            for city in CITY_LIST:
                if city.upper() == city_candidate.upper():
                    target_city = city
                    break
        
        # Si no tiene el formato END_, buscamos la palabra clave en el nombre
        if not target_city:
            for city in CITY_LIST:
                if city.upper() in base_name:
                    target_city = city
                    break
        
        if not target_city:
            logger.warning(f"No se pudo identificar la ciudad en el archivo: {filename}")
            continue

        # Proceder con el procesamiento...
        snap = parse_city_image(img_cv, target_city)
        
        if snap.commodities:
            if target_city in city_snapshots_dict:
                # Merge commodities if city already exists (updates existing entries)
                existing_snap = city_snapshots_dict[target_city]
                existing_commodities = {c.commodity: c for c in existing_snap.commodities}
                for new_comm in snap.commodities:
                    existing_commodities[new_comm.commodity] = new_comm # Overwrite with newest
                existing_snap.commodities = list(existing_commodities.values())
            else:
                city_snapshots_dict[target_city] = snap
        
        footer_info = ""
        if snap.footer:
            f = snap.footer
            mt_free = f.mt_available - f.mt_occupied
            footer_info = (f" | MT: {f.mt_occupied:,}/{f.mt_available:,}"
                           f" (free: {mt_free:,}) | CR: {f.cr_balance:,}")
        
        print(f"  ✔ {filename} → {target_city}  ({len(snap.commodities)} commodities{footer_info})")

    city_snapshots = list(city_snapshots_dict.values())

    if not city_snapshots and not easydock_rows:
        print("\n  No data found to save.")
        return

    # ── Layer 2: build catalog ─────────────────────────────────────────────────
    catalog = build_trade_catalog(city_snapshots, easydock_rows)

    # ── Layer 3: global opportunities + grading ───────────────────────────────
    raw_opps    = find_trade_opportunities(catalog, ship_capacity, budget)
    graded_opps = assign_grades(
        raw_opps, 
        ship_capacity=ship_capacity, 
        budget=budget,
        is_rental=is_rental, 
        rental_cost_per_day=ship_rental_cost
    )
    graded_opps.sort(key=lambda x: x.get("_profit_trip", 0), reverse=True)

    print(f"\n{'='*60}")
    print(f"   Total opportunities found: {len(graded_opps)}")
    if origin:
        from_origin = sum(1 for o in graded_opps if o["source"] == origin)
        print(f"   From {origin}: {from_origin}")
    print(f"{'='*60}")

    # ── Layer 4: export ───────────────────────────────────────────────────────
    save_to_excel(
        city_snapshots      = city_snapshots,
        easydock_rows       = easydock_rows,
        opportunities       = graded_opps,
        catalog             = catalog,
        selected_ship       = selected_ship,
        ship_capacity       = ship_capacity,
        output_file         = output_file,
        is_rental           = is_rental,
        rental_cost_per_day = ship_rental_cost,
        origin              = origin,
        budget              = budget,
        containers_used     = containers_used,
        mode                = mode,
        trade_route_params  = trade_route_params,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rise TVP Trade Route Optimizer")
    parser.add_argument("--images",     default="images",           help="Image folder path")
    parser.add_argument("--ship",       default=None,               help="Ship name")
    parser.add_argument("--output",     default="final_trade.xlsx", help="Output Excel file")
    parser.add_argument("--budget",     default=None, type=int,     help="Initial budget in CR")
    parser.add_argument("--containers", default=None, type=int,     help="Number of cargo containers")
    parser.add_argument("--origin",     default=None,               help="Origin city")
    parser.add_argument("--mode",       default="regular",
                        choices=["regular", "city", "route"],       help="Analysis mode")
    parser.add_argument("--calibrate",  default=None,               help="Path to image for calibration")
    
    args = parser.parse_args()

    # Modo calibración
    if args.calibrate:
        print("\n=== MODO CALIBRACIÓN ===")
        calibrate_regions(args.calibrate)
        sys.exit(0)

    # Modo normal
    try:
        logger.info("--- Iniciando optimizador Rise TVP (Versión Mejorada) ---")
        
        main(
            args.images, 
            selected_ship=args.ship, 
            output_file=args.output,
            budget=args.budget, 
            containers_used=args.containers,
            origin=args.origin, 
            mode=args.mode
        )
        
        logger.info("--- Proceso finalizado con éxito ---")

    except Exception as e:
        logger.error(f"FALLO CRÍTICO EN EL SCRIPT: {str(e)}", exc_info=True)
        print("\n" + "="*50)
        print(" SE HA DETECTADO UN ERROR. REVISA 'debug_log.txt'")
        print("="*50)
        input("\nPresiona Enter para cerrar...")