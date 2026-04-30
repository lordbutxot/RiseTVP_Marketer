"""
ocr_to_excel.py — Rise TVP Trade Route Optimizer
Extracts commodity data from screenshots via OCR and builds a trade analysis Excel workbook.
"""

import argparse
import itertools
import json
import math
import os
import platform
import re
import time
import openpyxl
import pytesseract
from pytesseract import Output
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
import cv2
import numpy as np

# ─────────────────────────────────────────────────────────────────────────────
# TESSERACT — auto-configure based on OS
# ─────────────────────────────────────────────────────────────────────────────

if platform.system() == "Windows":
    _TESSERACT_CANDIDATES = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]
    for _candidate in _TESSERACT_CANDIDATES:
        if os.path.isfile(_candidate):
            pytesseract.pytesseract.tesseract_cmd = _candidate
            break

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Images whose base name starts with any of these prefixes are skipped (not cities).
SKIP_IMAGE_PREFIXES = ("tvi",)

COMMODITY_CATEGORIES = [
    "Rare/Precious",
    "Foodstuffs",
    "Natural Materials",
    "Fuel Ore",
    "Consumer Goods",
    "Fabricated Material",
    "Refined Fuel",
]

SHIPS = {
    "AIR AND SPACE": {
        "E-10 Saint":   {"cargo_base": 7,  "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "E-11 Saint":   {"cargo_base": 7,  "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "P-13 Prowler": {"cargo_base": 1,  "max_containers": 0, "container_mt": 17, "rental_cost_per_day": None},
        "W-6 Manx":     {"cargo_base": 7,  "max_containers": 3, "container_mt": 17, "rental_cost_per_day": None},
    },
    "ONLY AIR": {
        "A-4 Wanderer":      {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
        "T-19 Stratomaster": {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
    },
}

CITIES = {
    "Alphaville":  {"x": 420, "y": 180},
    "Comstock":    {"x": 380, "y": 210},
    "Deadwood":    {"x": 340, "y": 230},
    "Ederar":      {"x": 300, "y": 200},
    "Erie":        {"x": 460, "y": 250},
    "Freedom":     {"x": 200, "y": 320},
    "Gettysburg":  {"x": 400, "y": 290},
    "Kansas":      {"x": 250, "y": 300},
    "Lancaster":   {"x": 350, "y": 350},
    "Pimli":       {"x": 270, "y": 280},
    "SovietUnion": {"x": 500, "y": 150},
    "Terrazul":    {"x": 480, "y": 320},
    "Sharney 1":   {"x": 320, "y": 400},
    "Sharney 2":   {"x": 360, "y": 450},
    "Sharney 3":   {"x": 400, "y": 500},
    "Delois Spot": {"x": 310, "y": 260},
}

CITY_LIST = sorted(CITIES.keys())

TAKEOFF_TIME = 5
LANDING_TIME = 10

FLIGHT_TIMES_EXPLICIT = {
    ("Delois Spot", "Alphaville"):  60,
    ("Delois Spot", "Comstock"):    55,
    ("Delois Spot", "Deadwood"):    60,
    ("Delois Spot", "Ederar"):      60,
    ("Delois Spot", "Erie"):        60,
    ("Delois Spot", "Freedom"):    150,
    ("Delois Spot", "Gettysburg"):  60,
    ("Delois Spot", "Kansas"):     150,
    ("Delois Spot", "Lancaster"):  120,
    ("Delois Spot", "Pimli"):       35,
    ("Delois Spot", "Sharney 1"):   60,
    ("Delois Spot", "Sharney 2"):  120,
    ("Delois Spot", "Sharney 3"):  180,
    ("Delois Spot", "SovietUnion"): 60,
    ("Delois Spot", "Terrazul"):    60,
    ("Kansas", "Alphaville"):  35,
    ("Kansas", "Comstock"):    30,
    ("Kansas", "Deadwood"):    25,
    ("Kansas", "Ederar"):      20,
    ("Kansas", "Erie"):        45,
    ("Kansas", "Freedom"):     15,
    ("Kansas", "Gettysburg"):  40,
    ("Kansas", "Lancaster"):   50,
    ("Kansas", "Pimli"):       10,
    ("Kansas", "Sharney 1"):   30,
    ("Kansas", "Sharney 2"):   60,
    ("Kansas", "Sharney 3"):   90,
    ("Kansas", "SovietUnion"): 65,
    ("Kansas", "Terrazul"):    60,
}

# Coordinates for Haversine-based flight time estimation
PLANET_RADIUS   = 2898.805  # km (0.455 × Earth radius)
FLIGHT_SPEED    = 10.0      # km/min
MIN_TRAVEL_TIME = 5.0       # minutes

CITY_COORDINATES = {
    "Alphaville":  {"lat": -4.426,  "lon": -22.115},
    "Deadwood":    {"lat": -11.077, "lon": -26.543},
    "Freedom":     {"lat": -12.812, "lon":  15.938},
    "Gettysburg":  {"lat": -2.616,  "lon": -35.244},
    "Kansas":      {"lat": -0.988,  "lon":  22.243},
    "SovietUnion": {"lat": -8.060,  "lon": -23.212},
    "Delois Spot": {"lat": -4.222,  "lon": -26.066},
    # Coordinates not yet confirmed — fall back to map-pixel distance:
    "Comstock":  None,
    "Ederar":    None,
    "Erie":      None,
    "Lancaster": None,
    "Pimli":     None,
    "Terrazul":  None,
}

# Derive pixels-per-minute speed from known Kansas routes
_ref_speeds = []
for (_src, _dst), _mins in FLIGHT_TIMES_EXPLICIT.items():
    if _src == "Kansas" and _dst in CITIES and _mins > 0:
        dx = CITIES["Kansas"]["x"] - CITIES[_dst]["x"]
        dy = CITIES["Kansas"]["y"] - CITIES[_dst]["y"]
        _ref_speeds.append(math.sqrt(dx * dx + dy * dy) / _mins)
SPEED_UNITS_PER_MIN = sum(_ref_speeds) / len(_ref_speeds) if _ref_speeds else 3.0

COMMODITY_COLORS = {
    "Rare/Precious":       "FFD700",
    "Foodstuffs":          "32CD32",
    "Natural Materials":   "1E90FF",
    "Fuel Ore":            "708090",
    "Consumer Goods":      "DC143C",
    "Fabricated Material": "8A2BE2",
    "Refined Fuel":        "FFA500",
}

GRADE_STYLES = {
    "A": {"fill": "00B050", "font": "FFFFFF"},
    "B": {"fill": "70AD47", "font": "FFFFFF"},
    "C": {"fill": "FFD700", "font": "000000"},
    "D": {"fill": "FFA500", "font": "FFFFFF"},
}




def is_area_red(image_cv, x, y, w, h, threshold=0.3):
    """
    Checks if a specific rectangular area in the image is red.
    'threshold' is the percentage of pixels that must be red to trigger.
    """
    # Crop the image to the bounding box of the text
    roi = image_cv[y:y+h, x:x+w]
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)

    # Define the range for red color in HSV
    # Red wraps around the 0-180 scale, so we check both ends
    lower_red1 = np.array([0, 70, 50])
    upper_red1 = np.array([10, 255, 255])
    lower_red2 = np.array([160, 70, 50])
    upper_red2 = np.array([180, 255, 255])

    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    full_mask = mask1 + mask2

    # Calculate the percentage of red pixels
    red_pixel_count = np.count_nonzero(full_mask)
    total_pixels = w * h
    return (red_pixel_count / total_pixels) > threshold

# ─────────────────────────────────────────────────────────────────────────────
# COLOR REGISTRY — deterministic per city / commodity name
# ─────────────────────────────────────────────────────────────────────────────

import hashlib

_CITY_COLOR_MAP: dict[str, PatternFill] = {}
_COMMODITY_COLOR_MAP: dict[str, PatternFill] = {}
_COLOR_MAPS_FILE = "color_maps.json"


def _hex_from_name(name: str) -> str:
    return hashlib.md5(name.encode()).hexdigest()[:6].upper()


def _get_fill(name: str, registry: dict) -> PatternFill:
    if name not in registry:
        color = _hex_from_name(name)
        registry[name] = PatternFill(start_color=color, end_color=color, fill_type="solid")
    return registry[name]


def get_city_fill(city: str) -> PatternFill:
    return _get_fill(city, _CITY_COLOR_MAP)


def get_commodity_fill(commodity: str) -> PatternFill:
    return _get_fill(commodity, _COMMODITY_COLOR_MAP)


def save_color_maps():
    data = {
        "cities":      {k: v.start_color.rgb for k, v in _CITY_COLOR_MAP.items()},
        "commodities": {k: v.start_color.rgb for k, v in _COMMODITY_COLOR_MAP.items()},
    }
    with open(_COLOR_MAPS_FILE, "w") as fh:
        json.dump(data, fh)


def load_color_maps():
    try:
        with open(_COLOR_MAPS_FILE) as fh:
            data = json.load(fh)
        for k, v in data.get("cities", {}).items():
            _CITY_COLOR_MAP[k] = PatternFill(start_color=v, end_color=v, fill_type="solid")
        for k, v in data.get("commodities", {}).items():
            _COMMODITY_COLOR_MAP[k] = PatternFill(start_color=v, end_color=v, fill_type="solid")
    except (FileNotFoundError, json.JSONDecodeError):
        pass

# ─────────────────────────────────────────────────────────────────────────────
# FLIGHT TIME
# ─────────────────────────────────────────────────────────────────────────────


def _haversine(lat1, lon1, lat2, lon2) -> float:
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlon / 2) ** 2
    return PLANET_RADIUS * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _pixel_distance(city_a: str, city_b: str) -> float:
    a, b = CITIES.get(city_a), CITIES.get(city_b)
    if not a or not b:
        return 0.0
    return math.sqrt((a["x"] - b["x"]) ** 2 + (a["y"] - b["y"]) ** 2)


def get_flight_time(origin: str, destination: str) -> float:
    """Return total travel time in minutes (takeoff + flight + landing)."""
    if origin == destination:
        return 0

    explicit = FLIGHT_TIMES_EXPLICIT.get((origin, destination)) or FLIGHT_TIMES_EXPLICIT.get((destination, origin))
    if explicit is not None:
        return TAKEOFF_TIME + explicit + LANDING_TIME

    c1, c2 = CITY_COORDINATES.get(origin), CITY_COORDINATES.get(destination)
    if c1 and c2:
        km = _haversine(c1["lat"], c1["lon"], c2["lat"], c2["lon"])
        return round(km / FLIGHT_SPEED + MIN_TRAVEL_TIME, 1)

    dist = _pixel_distance(origin, destination)
    if dist > 0:
        return round(TAKEOFF_TIME + dist / SPEED_UNITS_PER_MIN + LANDING_TIME, 1)

    return 120  # fallback

# ─────────────────────────────────────────────────────────────────────────────
# OCR
# ─────────────────────────────────────────────────────────────────────────────

def extract_text_from_image(image_path: str) -> str:
    """Extrae texto plano de una imagen usando Tesseract."""
    try:
        return pytesseract.image_to_string(Image.open(image_path)).strip()
    except Exception as exc:
        print(f"  ✗ OCR error on {image_path}: {exc}")
        return ""

def is_red(cv_img, x, y, w, h, threshold=0.15):
    """
    Detecta si un área específica de la imagen es roja.
    """
    pad = 2
    roi = cv_img[max(0, y-pad):y+h+pad, max(0, x-pad):x+w+pad]
    if roi.size == 0: return False
    
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    lower_red1, upper_red1 = np.array([0, 50, 50]), np.array([10, 255, 255])
    lower_red2, upper_red2 = np.array([160, 50, 50]), np.array([180, 255, 255])
    
    mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
    mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
    red_mask = cv2.bitwise_or(mask1, mask2)
    
    red_pixel_ratio = np.count_nonzero(red_mask) / (roi.shape[0] * roi.shape[1])
    return red_pixel_ratio > threshold

def parse_with_color(image_path):
    """Analiza la imagen detectando coordenadas y colores."""
    img = cv2.imread(image_path)
    if img is None:
        return []
        
    data = pytesseract.image_to_data(img, output_type=Output.DICT)
    rows = []
    n_boxes = len(data['text'])
    
    for i in range(n_boxes):
        text = data['text'][i].strip()
        if not text: continue
        
        x, y, w, h = data['left'][i], data['top'][i], data['width'][i], data['height'][i]
        
        if is_red(img, x, y, w, h):
            # Aquí puedes añadir lógica si detectas algo rojo fuera de las filas
            pass
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# PARSING
# ─────────────────────────────────────────────────────────────────────────────

# City screenshot: exactly 7 data rows, always in this fixed order.
# Positional AND keyword matching — dual strategy for maximum accuracy.
CITY_ROW_ORDER = [
    "Rare/Precious",
    "Foodstuffs",
    "Natural Materials",
    "Fuel Ore",
    "Consumer Goods",
    "Fabricated Material",
    "Refined Fuel",
]
assert CITY_ROW_ORDER == COMMODITY_CATEGORIES, "Row order must match COMMODITY_CATEGORIES"

# Keywords used to identify each commodity row even if position drifts.
# Every entry must be unique enough not to false-match another row.
CITY_ROW_KEYWORDS = {
    "Rare/Precious":       ("rare", "precious", "preci"),
    "Foodstuffs":          ("food", "stuff"),
    "Natural Materials":   ("natural", "mater"),
    "Fuel Ore":            ("fuel ore", "ore"),
    "Consumer Goods":      ("consumer", "goods"),
    "Fabricated Material": ("fabricated", "fabric"),
    "Refined Fuel":        ("refined", "fuel"),
}

# City column layout — numbers extracted after the leading text label:
#   col 0 → Quantity MT
#   col 1 → Reserve MT
#   col 2 → Selling CR/MT
#   col 3 → Buying CR/MT
#   col 4 → Maximum MT
CITY_COL_QTY  = 0
CITY_COL_RES  = 1
CITY_COL_SELL = 2
CITY_COL_BUY  = 3
CITY_COL_MAX  = 4

# City screen column header line — used to anchor the data block precisely.
# OCR may capitalise differently, so we match case-insensitively.
CITY_HEADER_KEYWORDS = ("commodity type", "quantity mt", "reserve mt",
                        "selling cr/mt", "buying cr/mt", "maximum mt")

# EasyDock column layout (numbers only, after name):
#   col 0 → MT (stock)
#   col 1 → Buying MT
#   col 2 → Buying CR
#   col 3 → Selling MT
#   col 4 → Selling CR
DOCK_COL_MT      = 0
DOCK_COL_BUY_MT  = 1
DOCK_COL_BUY_CR  = 2
DOCK_COL_SELL_MT = 3
DOCK_COL_SELL_CR = 4

# Lines containing these keywords are UI chrome, not data rows.
_SKIP_KEYWORDS = (
    "totals", "refresh", "cancel", "population",
    "fees", "ports staffed", "mt free", "cr free",
)


def normalize_number(token) -> int | None:
    """Convert a string/number token to int, stripping commas and noise."""
    if token is None:
        return None
    if isinstance(token, (int, float)):
        return int(token)
    clean = re.sub(r"[^0-9.]", "", str(token).replace(",", ""))
    try:
        return int(float(clean)) if clean else None
    except ValueError:
        return None


def _extract_numbers(line: str) -> list[int | None]:
    """Return all numbers found in a line, in order."""
    return [normalize_number(tok) for tok in re.findall(r"[\d,]+", line)]


def _is_skip_line(line: str) -> bool:
    low = line.lower()
    return any(kw in low for kw in _SKIP_KEYWORDS)


def _is_header_line(line: str) -> bool:
    """True if the line is the city column-header row."""
    low = line.lower()
    return sum(1 for kw in CITY_HEADER_KEYWORDS if kw in low) >= 3


def detect_layout(lines: list[str]) -> str:
    joined = " ".join(lines).lower()
    if any(k in joined for k in ["commodity type", "reserve mt", "selling cr/mt",
                                  "buying cr/mt", "maximum mt"]):
        return "city"
    if any(k in joined for k in ["buying mt", "selling mt", "buying cr", "selling cr"]):
        return "easydock"
    return "simple"


def _parse_totals_line(line: str) -> dict | None:
    """
    Parse lines like:
      'Totals  64,345 / 65,535 MT  97,339,730 CR'
    Returns {'mt_used': int, 'mt_total': int, 'cr_total': int} or None.
    """
    if "total" not in line.lower():
        return None
    numbers = _extract_numbers(line)
    if len(numbers) >= 3 and all(n is not None for n in numbers[:3]):
        return {
            "mt_used":  numbers[0],
            "mt_total": numbers[1],
            "cr_total": numbers[2],
        }
    return None


def _keyword_match_commodity(line: str) -> str | None:
    """
    Return the canonical commodity name if any keyword matches the line,
    otherwise None.  Used as a cross-check against positional assignment.
    """
    low = line.lower()
    for commodity, keywords in CITY_ROW_KEYWORDS.items():
        if any(kw in low for kw in keywords):
            return commodity
    return None


def _candidate_city_lines(lines: list[str]) -> list[str]:
    """
    Return lines that are plausible city commodity rows:
    - Not a skip/header line
    - Contains at least 2 numbers (tolerates OCR dropping zeros)
    - Has a non-empty text prefix before the first number
    Lines are returned in document order, starting from the row after the
    column-header line (anchored parsing) when the header is found.
    """
    # Try to anchor on the column-header line first
    header_idx = None
    for i, line in enumerate(lines):
        if _is_header_line(line):
            header_idx = i
            break

    search_lines = lines[header_idx + 1:] if header_idx is not None else lines

    result = []
    for line in search_lines:
        if _is_skip_line(line):
            continue
        if _is_header_line(line):
            continue
        nums = _extract_numbers(line)
        if len(nums) < 2:          # reduced threshold — zeros may be dropped by OCR
            continue
        first_num = re.search(r"[\d,]+", line)
        if not first_num:
            continue
        prefix = line[: first_num.start()].strip()
        if not prefix:
            continue
        result.append(line)
    return result


def _parse_city_rows_with_color(img_cv):
    """
    Versión con anclaje de columnas: Usa coordenadas X para que los números
    no se desplacen si el OCR falla en una columna.
    """
    d = pytesseract.image_to_data(img_cv, output_type=Output.DICT)
    img_width = img_cv.shape[1]
    
    # Mapeo de columnas por posición horizontal (%)
    # Qty (~35%), Reserve (~52%), Sell (~65%), Buy (~79%), Max (~93%)
    COL_RANGES = {
        "qty":     (0.30, 0.45),
        "reserve": (0.46, 0.58),
        "sell":    (0.59, 0.72),
        "buy":     (0.73, 0.86),
        "max":     (0.87, 1.00)
    }

    lines = {}
    for i in range(len(d['text'])):
        word = d['text'][i].strip()
        if not word: continue
        
        # Agrupar por línea física (Y)
        y_coord = d['top'][i] // 10 * 10 
        if y_coord not in lines: lines[y_coord] = []
        
        # Guardar posición X relativa (0.0 a 1.0)
        rel_x = d['left'][i] / img_width
        is_word_red = is_red(img_cv, d['left'][i], d['top'][i], d['width'][i], d['height'][i])
        lines[y_coord].append({'text': word, 'x': rel_x, 'red': is_word_red})

    results = []
    for y in sorted(lines.keys()):
        line_data = lines[y]
        line_text = " ".join([w['text'] for w in line_data]).lower()
        
        # Verificar si la línea contiene una materia prima válida
        commodity = _keyword_match_commodity(line_text)
        if not commodity: continue

        # Diccionario para colocar cada valor en su "cajón" correcto
        row_map = {"qty": 0, "reserve": 0, "sell": None, "buy": None, "max": 0}

        for w in line_data:
            # Limpiar el texto para dejar solo números
            clean = re.sub(r"[^0-9.]", "", w['text'])
            if not clean: continue
            
            try:
                val = int(float(clean))
                # Buscamos a qué columna pertenece según su X
                for col_name, (xmin, xmax) in COL_RANGES.items():
                    if xmin <= w['x'] <= xmax:
                        if w['red']:
                            # Si es rojo, cantidad/max es 0, precios son None
                            row_map[col_name] = None if col_name in ["sell", "buy"] else 0
                        else:
                            row_map[col_name] = val
                        break
            except: continue

        # Insertar los datos en el orden exacto que espera el Excel
        results.append([
            commodity, commodity,
            row_map["qty"], row_map["reserve"], row_map["sell"], 
            row_map["buy"], row_map["max"]
        ])
            
    return results, None


def _parse_easydock_rows(lines: list[str]) -> list:
    """
    EasyDock commodities are not in fixed order; use name-prefix + positional
    column numbers.  All rows with a price (sell or buy) are kept, including
    zero-stock rows, so buying prices are visible for analysis.

    Returns rows: [Category, Name, MT, Buying MT, Buying CR, Selling MT, Selling CR]
    """
    result = []
    for line in lines:
        if _is_skip_line(line):
            continue
        if _is_header_line(line):
            continue
        nums = _extract_numbers(line)
        if len(nums) < 2:          # at least a price must be present
            continue
        first_num = re.search(r"[\d,]+", line)
        if not first_num:
            continue
        name = line[: first_num.start()].strip()
        if not name:
            continue
        category = _infer_category(name)

        def _n(idx):
            return nums[idx] if len(nums) > idx and nums[idx] is not None else 0

        sell_cr = _n(DOCK_COL_SELL_CR)
        buy_cr  = _n(DOCK_COL_BUY_CR)

        # Keep row if there is any price information at all
        if sell_cr == 0 and buy_cr == 0:
            continue

        result.append([
            category, name,
            _n(DOCK_COL_MT),
            _n(DOCK_COL_BUY_MT),
            buy_cr,
            _n(DOCK_COL_SELL_MT),
            sell_cr,
        ])
    return result


def _infer_category(raw: str) -> str:
    """Match raw OCR text to the nearest known commodity category."""
    text = re.sub(r"^[^A-Za-z]*", "", str(raw or "")).strip()
    for cat in COMMODITY_CATEGORIES:
        if cat.lower() in text.lower():
            return cat
    return text


def parse_text_to_data(text: str) -> dict:
    """
    Main entry point: parse raw OCR text into structured rows.
    Returns {'header': [...], 'rows': [...], 'totals': dict|None, 'layout': str}
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    layout = detect_layout(lines)

    if layout == "city":
        header = ["Category", "Commodity Type", "Quantity MT", "Reserve MT",
                  "Selling CR/MT", "Buying CR/MT", "Maximum MT"]
        rows, totals = _parse_city_rows(lines)
        return {"header": header, "rows": rows, "totals": totals, "layout": "city"}

    if layout == "easydock":
        header = ["Category", "Name", "MT", "Buying MT", "Buying CR", "Selling MT", "Selling CR"]
        rows = _parse_easydock_rows(lines)
        return {"header": header, "rows": rows, "totals": None, "layout": "easydock"}

    # Simple key:value fallback
    header = ["Resource/Commodity", "Value"]
    rows = []
    for line in lines:
        if ":" in line:
            k, v = line.split(":", 1)
            rows.append([k.strip(), v.strip()])
        else:
            rows.append([line])
    return {"header": header, "rows": rows, "totals": None, "layout": "simple"}


def sanitize_sheet_name(name: str) -> str:
    return "".join("-" if ch in r'\/?*[]:"' else ch for ch in name)[:31]


def normalize_commodity_key(category: str, name: str = "") -> str:
    for src in (category, name):
        text = re.sub(r"^[^A-Za-z]*", "", str(src or "")).strip()
        for cat in COMMODITY_CATEGORIES:
            if cat.lower() in text.lower():
                return cat
        if text:
            return text
    return ""

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────


def _contrast(hex_color: str) -> str:
    hx = hex_color.strip().lstrip("#").ljust(6, "0")[:6]
    try:
        r, g, b = int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16)
    except ValueError:
        return "000000"
    return "000000" if (r * 299 + g * 587 + b * 114) / 1000 > 160 else "FFFFFF"


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(cells):
    for cell in cells:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _style_section_title(cells):
    for cell in cells:
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()


def _style_subsection_title(cells, color="E6E6FA"):
    for cell in cells:
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = _thin_border()


def _style_commodity_separator(cells, color):
    thick = Side(style="thick")
    for cell in cells:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=thick, bottom=thick)


def _auto_size_columns(ws, min_width=10, max_width=50):
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        if isinstance(col_cells[0], MergedCell):
            continue
        max_len = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col_cells
        )
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_width, max(min_width, max_len + 2))


def _apply_roi_color(cell, roi_pct: float):
    if roi_pct > 100:
        fill, font_color = "00B050", "FFFFFF"
    elif roi_pct >= 50:
        fill, font_color = "70AD47", "FFFFFF"
    elif roi_pct >= 20:
        fill, font_color = "FFD700", "000000"
    elif roi_pct >= 0:
        fill, font_color = "FFA500", "FFFFFF"
    else:
        fill, font_color = "FF0000", "FFFFFF"
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.font = Font(color=font_color, bold=True)

# ─────────────────────────────────────────────────────────────────────────────
# TRADE CATALOG & OPPORTUNITIES
# ─────────────────────────────────────────────────────────────────────────────


def build_trade_catalog(data_dict: dict) -> dict:
    catalog: dict = {}
    for sheet_name in ("Cities", "EasyDock"):
        if sheet_name not in data_dict:
            continue
        header = [h.lower() for h in data_dict[sheet_name]["header"]]
        for row in data_dict[sheet_name]["rows"]:
            if len(row) < 3:
                continue
            location = str(row[0]).strip()
            category = str(row[1]).strip()
            raw_name = str(row[2]).strip() if len(row) > 2 else ""
            commodity_key = normalize_commodity_key(category, raw_name)
            loc = catalog.setdefault(location, {})

            selling = buying = None
            quantity = reserve = max_accept = sell_capacity = buy_capacity = 0

            def _get(idx):
                return normalize_number(row[idx]) if len(row) > idx else None

            if sheet_name == "Cities" and "selling cr/mt" in header:
                qi, ri, si, bi, mi = (header.index(k) for k in
                                      ("quantity mt", "reserve mt", "selling cr/mt", "buying cr/mt", "maximum mt"))
                selling, buying    = _get(si), _get(bi)
                quantity, reserve  = (_get(qi) or 0), (_get(ri) or 0)
                max_accept         = _get(mi) or 0
                sell_capacity      = max(quantity - reserve, 0)
                buy_capacity       = max_accept

            elif sheet_name == "EasyDock" and "selling cr" in header:
                qi, bci, sci, bi, si = (header.index(k) for k in
                                        ("mt", "buying mt", "selling mt", "buying cr", "selling cr"))
                selling, buying      = _get(si), _get(bi)
                quantity             = _get(qi) or 0
                buy_capacity         = _get(bci) or 0
                sell_capacity        = _get(sci) or 0

            loc[commodity_key] = {
                "selling":      selling,
                "buying":       buying,
                "quantity":     quantity,
                "reserve":      reserve,
                "max_accept":   max_accept,
                "sell_capacity": sell_capacity,
                "buy_capacity":  buy_capacity,
            }
    return catalog


def find_trade_opportunities(catalog, ship_capacity, budget):
    opportunities = []
    for src, commodities in catalog.items():
        for commodity, src_data in commodities.items():
            # Filtro estricto: Si no hay precio de venta o es 0, no se puede comprar
            if src_data.get("selling") is None or src_data["selling"] <= 0:
                continue
                
            for dst, dst_commodities in catalog.items():
                if src == dst: continue
                if commodity not in dst_commodities: continue
                
                dst_data = dst_commodities[commodity]
                # Filtro estricto: Si no hay precio de compra, no se puede vender
                if dst_data.get("buying") is None or dst_data["buying"] <= 0:
                    continue
                
                profit_per_mt = dst_data["buying"] - src_data["selling"]
                
                if profit_per_mt > 0:
                    # Aquí calculas el ROI y el beneficio real
                    qty_to_buy = min(src_data.get("qty", 0), ship_capacity)
                    total_profit = qty_to_buy * profit_per_mt
                    
                    if total_profit > 0:
                        opportunities.append({
                            "commodity": commodity,
                            "source": src,
                            "destination": dst,
                            "profit_mt": profit_per_mt,
                            "total_profit": total_profit,
                            # ... resto de datos ...
                        })
    return opportunities

# ─────────────────────────────────────────────────────────────────────────────
# GRADING
# ─────────────────────────────────────────────────────────────────────────────


def assign_grades(opportunities: list, ship_capacity: int = 109,
                  budget: int = None, is_rental: bool = False,
                  rental_cost_per_day: int = None) -> list:
    for op in opportunities:
        price        = op["source_selling"] or 1
        max_by_budget = int(budget / price) if budget else float("inf")
        qty = max(min(ship_capacity, op["source_available"], op["destination_capacity"], max_by_budget), 0)
        cost_trip    = qty * price
        profit_trip  = qty * op["profit_per_mt"]
        roi_trip     = (profit_trip / cost_trip * 100) if cost_trip > 0 else 0

        op["_qty_trip"]    = qty
        op["_cost_trip"]   = cost_trip
        op["_profit_trip"] = profit_trip
        op["_roi"]         = roi_trip
        op["_affordable"]  = qty > 0 and profit_trip > 0
        if is_rental and rental_cost_per_day:
            op["covers_rental"] = profit_trip >= rental_cost_per_day

    viable = sorted([op for op in opportunities if op["_affordable"]],
                    key=lambda x: x["_profit_trip"], reverse=True)
    n = len(viable)
    for idx, op in enumerate(viable):
        pct = idx / n if n else 1
        op["grade"] = "A" if pct < 0.25 else "B" if pct < 0.50 else "C" if pct < 0.75 else "D"
    for op in opportunities:
        if "grade" not in op:
            op["grade"] = "D"
    return opportunities

# ─────────────────────────────────────────────────────────────────────────────
# TRADE ROUTE ENGINE
# ─────────────────────────────────────────────────────────────────────────────


def _best_leg(catalog, src, dst, commodities, ship_capacity, budget):
    travel_time = get_flight_time(src, dst)
    best = None
    for commodity in commodities:
        s = catalog.get(src, {}).get(commodity)
        d = catalog.get(dst, {}).get(commodity)
        if not s or not d or s["selling"] is None or d["buying"] is None:
            continue
        profit_per_mt = d["buying"] - s["selling"]
        available = s["quantity"] - s["reserve"]
        capacity  = d["buy_capacity"]
        if profit_per_mt <= 0 or available <= 0 or capacity <= 0:
            continue
        max_by_budget = int(budget / s["selling"]) if budget and s["selling"] > 0 else float("inf")
        qty = max(min(ship_capacity, available, capacity, max_by_budget), 0)
        if qty <= 0:
            continue
        cost   = qty * s["selling"]
        profit = qty * profit_per_mt
        if best is None or profit > best["profit"]:
            best = {
                "commodity":     commodity,
                "src": src, "dst": dst,
                "buy_price":     s["selling"],
                "sell_price":    d["buying"],
                "profit_per_mt": profit_per_mt,
                "qty": qty, "cost": cost, "profit": profit,
                "roi":           (profit / cost * 100) if cost > 0 else 0,
                "flight_min":    travel_time,
            }
    if best is None:
        return {"commodity": "— empty —", "src": src, "dst": dst,
                "buy_price": 0, "sell_price": 0, "profit_per_mt": 0,
                "qty": 0, "cost": 0, "profit": 0, "roi": 0, "flight_min": travel_time}
    return best


def compute_trade_routes(catalog, origin, allowed_commodities, ship_capacity,
                         budget, max_hops=5, top_n=20) -> list:
    locations = [loc for loc in catalog if loc != origin]
    all_routes, seen = [], set()

    for n_stops in range(1, max_hops + 1):
        for stops in itertools.permutations(locations, n_stops):
            seq = [origin] + list(stops) + [origin]
            legs, total_profit, total_cost, total_time = [], 0, 0, 0
            remaining = budget
            valid = True

            for i in range(len(seq) - 1):
                leg = _best_leg(catalog, seq[i], seq[i + 1], allowed_commodities, ship_capacity, remaining)
                if leg["profit"] == 0 and i < len(seq) - 2:
                    valid = False
                    break
                legs.append(leg)
                total_profit += leg["profit"]
                total_cost   += leg["cost"]
                total_time   += leg["flight_min"]
                if remaining is not None:
                    remaining = remaining - leg["cost"] + leg["profit"]

            if not valid or total_profit <= 0:
                continue

            roi         = (total_profit / total_cost * 100) if total_cost > 0 else 0
            cr_per_hour = (total_profit / total_time * 60) if total_time > 0 else 0
            key = tuple(seq)
            if key not in seen:
                seen.add(key)
                all_routes.append({
                    "stops": seq, "legs": legs, "n_hops": len(legs),
                    "total_profit": total_profit, "total_cost": total_cost,
                    "total_time": total_time, "roi": roi, "cr_per_hour": cr_per_hour,
                })
            if len(all_routes) >= top_n * 5:
                break

    return sorted(all_routes, key=lambda x: x["total_profit"], reverse=True)[:top_n]

# ─────────────────────────────────────────────────────────────────────────────
# SHEET WRITERS
# ─────────────────────────────────────────────────────────────────────────────


def _write_trade_routes_sheet(ws, routes: list, origin: str):
    alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    if not routes:
        ws.append(["No profitable routes found from", origin])
        return

    for rank, route in enumerate(routes, 1):
        summary = (
            f"Route #{rank}  |  {' → '.join(route['stops'])}  |  "
            f"Profit: {route['total_profit']:,.0f} CR  |  ROI: {route['roi']:.1f}%  |  "
            f"Time: {route['total_time']} min  |  Efficiency: {route['cr_per_hour']:,.0f} CR/h"
        )
        ws.append([summary])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        c = ws.cell(row=r, column=1)
        c.font      = Font(bold=True, size=12, color="FFFFFF")
        c.fill      = get_city_fill(origin)
        c.alignment = Alignment(horizontal="left", vertical="center")

        ws.append(["Hop", "From", "To", "Commodity", "Buy (CR/MT)", "Sell (CR/MT)",
                   "Profit/MT", "MT", "Leg Cost (CR)", "Leg Profit (CR)", "Flight (min)", "Leg ROI (%)"])
        _style_header_row(ws[ws.max_row])

        for hop_i, leg in enumerate(route["legs"], 1):
            ws.append([hop_i, leg["src"], leg["dst"], leg["commodity"],
                       leg["buy_price"], leg["sell_price"], leg["profit_per_mt"],
                       leg["qty"], leg["cost"], leg["profit"], leg["flight_min"], round(leg["roi"], 2)])
            cur = ws.max_row
            for col, city_key in ((2, leg["src"]), (3, leg["dst"])):
                fill = get_city_fill(city_key)
                cell = ws.cell(row=cur, column=col)
                cell.fill = fill
                cell.font = Font(color=_contrast(fill.start_color.rgb), bold=True)
            _apply_roi_color(ws.cell(row=cur, column=12), leg["roi"])
            if hop_i % 2 == 0:
                for col in (1, 4, 5, 6, 7, 8, 9, 10, 11):
                    ws.cell(row=cur, column=col).fill = alt_fill

        ws.append(["", "", "", "TOTAL", "", "", "", "",
                   route["total_cost"], route["total_profit"], route["total_time"], round(route["roi"], 2)])
        tot = ws.max_row
        for col in range(1, 13):
            ws.cell(row=tot, column=col).font = Font(bold=True)
            ws.cell(row=tot, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _apply_roi_color(ws.cell(row=tot, column=12), route["roi"])
        ws.append([])

    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def _build_opp_columns(is_rental, rental_cost_per_day) -> list:
    cols = ["Grade", "Commodity", "Source", "Buy Price (CR/MT)",
            "Destination", "Sell Price (CR/MT)", "Profit/MT (CR)",
            "Src Stock (MT)", "Dst Capacity (MT)", "MT loaded per trip",
            "Trip Cost (CR)", "Trip Profit (CR)", "Trip ROI (%)", "Travel Time (min)"]
    if is_rental and rental_cost_per_day:
        cols += ["Rental Cost (CR)", "Covers Rental?"]
    return cols


def _write_opportunities_sheet(ws, opportunities, columns, ship_capacity,
                                is_rental, rental_cost_per_day, origin, grade_col, roi_col):
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    ws.append(columns)
    _style_header_row(ws[1])

    for op in opportunities:
        grade, qty_trip = op["grade"], op["_qty_trip"]
        cost_trip, profit_trip, roi_pct = op["_cost_trip"], op["_profit_trip"], op["_roi"]
        travel_time = get_flight_time(op["source"], op["destination"])

        row = [grade, op["commodity"], op["source"], op["source_selling"],
               op["destination"], op["destination_buying"], op["profit_per_mt"],
               op["source_available"], op["destination_capacity"],
               qty_trip, cost_trip, profit_trip, roi_pct, travel_time]
        if is_rental and rental_cost_per_day:
            row += [rental_cost_per_day, "Yes" if profit_trip > rental_cost_per_day else "No"]

        ws.append(row)
        cur = ws.max_row

        # Grade cell
        gs = GRADE_STYLES[grade]
        gc = ws.cell(row=cur, column=grade_col)
        gc.fill      = PatternFill(start_color=gs["fill"], end_color=gs["fill"], fill_type="solid")
        gc.font      = Font(color=gs["font"], bold=True)
        gc.alignment = Alignment(horizontal="center")

        # Commodity cell
        cf = get_commodity_fill(op["commodity"])
        cc = ws.cell(row=cur, column=2)
        cc.fill = cf
        cc.font = Font(color=_contrast(cf.start_color.rgb), bold=True)

        # Source city cell
        sf = get_city_fill(op["source"])
        sc = ws.cell(row=cur, column=3)
        sc.fill = sf
        sc.font = Font(color=_contrast(sf.start_color.rgb), bold=True)

        # Destination city cell
        df = get_city_fill(op["destination"])
        dc = ws.cell(row=cur, column=5)
        dc.fill = df
        dc.font = Font(color=_contrast(df.start_color.rgb), bold=True)

        _apply_roi_color(ws.cell(row=cur, column=roi_col), roi_pct)

        if is_rental and rental_cost_per_day:
            rc = ws.cell(row=cur, column=len(columns))
            rc.fill = green_fill if rc.value == "Yes" else red_fill
            rc.font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "B2"
    _auto_size_columns(ws)

# ─────────────────────────────────────────────────────────────────────────────
# MACRO DATA
# ─────────────────────────────────────────────────────────────────────────────


def calculate_macro_data(catalog: dict, opportunities: list) -> dict:
    city_profits, city_commodities, city_lucrative = {}, {}, {}
    for op in opportunities:
        src = op["source"]
        city_profits.setdefault(src, 0)
        city_commodities.setdefault(src, set())
        city_lucrative.setdefault(src, set())
        city_profits[src] += op["total_profit"]
        city_commodities[src].add(op["commodity"])
        if op["total_profit"] > 0:
            city_lucrative[src].add(op["commodity"])

    city_summary = sorted(
        [{"city": c, "total_profit": city_profits.get(c, 0),
          "num_commodities": len(city_commodities.get(c, set())),
          "num_lucrative": len(city_lucrative.get(c, set()))}
         for c in catalog],
        key=lambda x: x["total_profit"], reverse=True,
    )

    commodity_best_sellers, commodity_best_buyers = {}, {}
    for commodity in COMMODITY_CATEGORIES:
        sellers, buyers = [], []
        for city, data in catalog.items():
            if commodity not in data:
                continue
            item = data[commodity]
            if item["selling"] is not None and item["sell_capacity"] > 0:
                sellers.append({"city": city, "price": item["selling"], "capacity": item["sell_capacity"]})
            if item["buying"] is not None and item["buy_capacity"] > 0:
                buyers.append({"city": city, "price": item["buying"], "capacity": item["buy_capacity"]})
        commodity_best_sellers[commodity] = sorted(sellers, key=lambda x: x["price"])[:5]
        commodity_best_buyers[commodity]  = sorted(buyers, key=lambda x: x["price"], reverse=True)[:5]

    return {
        "city_summary":                city_summary,
        "commodity_best_sellers":      commodity_best_sellers,
        "commodity_best_buyers":       commodity_best_buyers,
        "top_cities_by_profit":        city_summary[:10],
        "cities_by_lucrative_commodities": sorted(city_summary, key=lambda x: x["num_lucrative"], reverse=True)[:10],
    }

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────


def save_to_excel(data_dict, selected_ship, ship_capacity, output_file, 
                  is_rental, rental_cost_per_day, origin, budget, 
                  containers_used, mode, trade_route_params, all_opportunities=None):

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    alt_fill  = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    alt_fill2 = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # ── Config sheet ──────────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Parameter", "Value"])
    _style_header_row(ws_cfg[1])

    ship_data, ship_category = None, ""
    for cat, ships in SHIPS.items():
        if selected_ship in ships:
            ship_data, ship_category = ships[selected_ship], cat
            break

    max_c   = ship_data["max_containers"] if ship_data else 0
    c_used  = containers_used if containers_used is not None else max_c
    c_mt    = ship_data["container_mt"] if ship_data else 17
    cargo_b = ship_data["cargo_base"] if ship_data else 0

    config_rows = [
        ("Ship",                  selected_ship),
        ("Category",              ship_category),
        ("Base cargo (fixed MT)", cargo_b),
        ("Containers carried",    f"{c_used} / {max_c} max"),
        ("MT per container",      c_mt),
        ("Total Capacity (MT)",   f"{cargo_b} + {c_used} x {c_mt} = {ship_capacity}"),
        ("Origin",                origin),
        ("OS",                    platform.system()),
        ("Status",                "Rented" if is_rental else "Purchased"),
    ]
    if is_rental and rental_cost_per_day:
        config_rows += [
            ("Rental cost/day",               f"{rental_cost_per_day} CR/day (14 h)"),
            ("Min profit/day to cover rental", f"{rental_cost_per_day} CR"),
            ("Min profit/hour to cover rental", f"{rental_cost_per_day / 14:.2f} CR/h"),
        ]
    if budget:
        config_rows.append(("Initial Budget", f"{budget} CR"))
    if trade_route_params:
        config_rows += [
            ("Mode",               "Trade Route"),
            ("Max hops",          trade_route_params.get("max_hops", 5)),
            ("Allowed commodities", ", ".join(trade_route_params.get("allowed_commodities", ["ALL"]))),
        ]
    for row in config_rows:
        ws_cfg.append(list(row))
    _auto_size_columns(ws_cfg)

    # ── Cities / EasyDock sheets ──────────────────────────────────────────────
    city_totals = data_dict.get("_totals", {})
    for sheet_name in ("Cities", "EasyDock"):
        if data_dict.get(sheet_name, {}).get("rows"):
            ws = wb.create_sheet(sheet_name)
            ws.append(data_dict[sheet_name]["header"])
            _style_header_row(ws[1])
            for i, row in enumerate(data_dict[sheet_name]["rows"], start=2):
                ws.append(row)
                if i % 2 == 0:
                    for cell in ws[i]:
                        cell.fill = alt_fill
            _auto_size_columns(ws)

            # Append a Totals summary block for city screenshots
            if sheet_name == "Cities" and city_totals:
                ws.append([])
                summary_header = ["City", "MT Used", "MT Total", "MT Free", "CR in City"]
                ws.append(summary_header)
                _style_header_row(ws[ws.max_row])
                for city, t in sorted(city_totals.items()):
                    mt_free = (t["mt_total"] - t["mt_used"]) if t["mt_total"] else None
                    ws.append([city, t["mt_used"], t["mt_total"], mt_free, t["cr_total"]])
                    cur = ws.max_row
                    # Colour the city name cell
                    cf = get_city_fill(city)
                    cc = ws.cell(row=cur, column=1)
                    cc.fill = cf
                    cc.font = Font(color=_contrast(cf.start_color.rgb), bold=True)
                    # Highlight if city is almost full (> 90 % used)
                    if t["mt_total"] and t["mt_used"] / t["mt_total"] > 0.9:
                        for col in range(2, 5):
                            ws.cell(row=cur, column=col).fill = PatternFill(
                                start_color="FF0000", end_color="FF0000", fill_type="solid")
                            ws.cell(row=cur, column=col).font = Font(color="FFFFFF", bold=True)
                _auto_size_columns(ws)

    # ── Opportunities ─────────────────────────────────────────────────────────
    catalog = build_trade_catalog(data_dict)
    opportunities = find_trade_opportunities(catalog, ship_capacity, budget)

    if all_opportunities:
        all_opportunities = assign_grades(
            all_opportunities, ship_capacity=ship_capacity,
            budget=budget, is_rental=is_rental, rental_cost_per_day=rental_cost_per_day,
        )
        all_opportunities.sort(key=lambda x: x["_profit_trip"], reverse=True)

        columns   = _build_opp_columns(is_rental, rental_cost_per_day)
        grade_col = 1
        roi_col   = columns.index("Trip ROI (%)") + 1

        ws_op = wb.create_sheet("Opportunities")
        _write_opportunities_sheet(ws_op, all_opportunities, columns, ship_capacity,
                                   is_rental, rental_cost_per_day, origin, grade_col, roi_col)

        if mode == "city":
            city_opps = [op for op in all_opportunities if op["source"] == origin]
            if city_opps:
                ws_city = wb.create_sheet(sanitize_sheet_name(f"From {origin}"))
                _write_opportunities_sheet(ws_city, city_opps, columns, ship_capacity,
                                           is_rental, rental_cost_per_day, origin, grade_col, roi_col)

    # ── Trade Routes ──────────────────────────────────────────────────────────
    if mode == "route" and trade_route_params:
        allowed  = trade_route_params.get("allowed_commodities", COMMODITY_CATEGORIES)
        max_hops = trade_route_params.get("max_hops", 3)
        print(f"\n  Computing trade routes from {origin} (max {max_hops} hops)…")
        routes = compute_trade_routes(catalog, origin, allowed, ship_capacity, budget, max_hops, top_n=20)
        ws_routes = wb.create_sheet(sanitize_sheet_name(f"Trade Routes {origin}"))
        _write_trade_routes_sheet(ws_routes, routes, origin)
        print(f"  Found {len(routes)} profitable routes.")
        if routes:
            best = routes[0]
            print(f"  Best: {' → '.join(best['stops'])} | "
                  f"Profit: {best['total_profit']:,.0f} CR | ROI: {best['roi']:.1f}%")

    # ── MACRO sheet ───────────────────────────────────────────────────────────
    macro_data = calculate_macro_data(catalog, all_opportunities if all_opportunities else [])
    ws_m = wb.create_sheet("MACRO")

    def _macro_section(title):
        ws_m.append([title])
        _style_section_title(ws_m[ws_m.max_row])

    _macro_section("SUMMARY BY CITY")
    ws_m.append(["City", "Total Potential Profit (CR)", "Num Commodities", "Num Lucrative Commodities"])
    _style_header_row(ws_m[ws_m.max_row])
    for i, item in enumerate(macro_data["city_summary"], start=ws_m.max_row + 1):
        ws_m.append([item["city"], item["total_profit"], item["num_commodities"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws_m[i]: cell.fill = alt_fill2

    ws_m.append([])
    _macro_section("TOP 10 CITIES BY TOTAL PROFIT")
    ws_m.append(["City", "Total Profit (CR)", "Num Lucrative Commodities"])
    _style_header_row(ws_m[ws_m.max_row])
    for i, item in enumerate(macro_data["top_cities_by_profit"], start=ws_m.max_row + 1):
        ws_m.append([item["city"], item["total_profit"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws_m[i]: cell.fill = alt_fill2

    ws_m.append([])
    _macro_section("TOP 10 CITIES BY NUMBER OF LUCRATIVE COMMODITIES")
    ws_m.append(["City", "Num Lucrative Commodities", "Total Profit (CR)"])
    _style_header_row(ws_m[ws_m.max_row])
    for i, item in enumerate(macro_data["cities_by_lucrative_commodities"], start=ws_m.max_row + 1):
        ws_m.append([item["city"], item["num_lucrative"], item["total_profit"]])
        if i % 2 == 0:
            for cell in ws_m[i]: cell.fill = alt_fill2

    ws_m.append([])
    _macro_section("BEST SELLERS AND BUYERS BY COMMODITY")
    for commodity in COMMODITY_CATEGORIES:
        color = COMMODITY_COLORS.get(commodity, "E6E6FA")
        ws_m.append([])
        ws_m.append([f"--- {commodity.upper()} ---"])
        _style_commodity_separator(ws_m[ws_m.max_row], color)
        for section, key in (("Best Sellers (lowest price)", "commodity_best_sellers"),
                              ("Best Buyers (highest price)", "commodity_best_buyers")):
            ws_m.append([])
            ws_m.append([f"{commodity} - {section}"])
            _style_subsection_title(ws_m[ws_m.max_row], color)
            ws_m.append(["City",
                         "Selling Price (CR/MT)" if "Seller" in section else "Buying Price (CR/MT)",
                         "Capacity (MT)"])
            _style_header_row(ws_m[ws_m.max_row])
            for i, entry in enumerate(macro_data[key].get(commodity, []), start=ws_m.max_row + 1):
                ws_m.append([entry["city"], entry["price"], entry["capacity"]])
                if i % 2 == 0:
                    for cell in ws_m[i]: cell.fill = alt_fill2

    _auto_size_columns(ws_m)

    # ── Save ──────────────────────────────────────────────────────────────────
    save_color_maps()
    try:
        wb.save(output_file)
        print(f"\n✔  Saved: {output_file}")
        print(f"   Ship: {selected_ship} | Capacity: {ship_capacity} MT | Origin: {origin}")
    except PermissionError:
        ts = int(time.time())
        base, ext = os.path.splitext(output_file)
        alt_path = f"{base}_{ts}{ext}"
        wb.save(alt_path)
        print(f"⚠  '{output_file}' was in use — saved as '{alt_path}'")

# ─────────────────────────────────────────────────────────────────────────────
# PROMPT HELPERS
# ─────────────────────────────────────────────────────────────────────────────


def _prompt_city(prompt_text: str) -> str:
    print(f"\n{prompt_text}")
    for i, city in enumerate(CITY_LIST, 1):
        print(f"  {i:>2}. {city}")
    while True:
        try:
            choice = int(input("City number: ").strip())
            if 1 <= choice <= len(CITY_LIST):
                return CITY_LIST[choice - 1]
        except ValueError:
            pass
        print(f"  Enter a number between 1 and {len(CITY_LIST)}.")


def _prompt_commodities() -> list:
    print("\n  Select commodities to trade (comma-separated numbers, or 0 for all):")
    print("   0. ALL commodities")
    for i, cat in enumerate(COMMODITY_CATEGORIES, 1):
        print(f"  {i:>2}. {cat}")
    while True:
        raw = input("  Choice: ").strip()
        if raw == "0":
            return list(COMMODITY_CATEGORIES)
        try:
            indices = [int(p) for p in raw.split(",") if p.strip()]
            selected = [COMMODITY_CATEGORIES[i - 1] for i in indices if 1 <= i <= len(COMMODITY_CATEGORIES)]
            if selected:
                print(f"  Selected: {', '.join(selected)}")
                return selected
        except (ValueError, IndexError):
            pass
        print("  Enter valid comma-separated numbers or 0.")


def _prompt_max_hops() -> int:
    print("\n  Max intermediate stops? (2–5 | more = slower calculation)")
    while True:
        try:
            n = int(input("  Max stops (2-5): ").strip())
            if 2 <= n <= 5:
                return n
        except ValueError:
            pass
        print("  Enter a number between 2 and 5.")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────


def main(image_folder, selected_ship=None, output_file="final_trade.xlsx",
         budget=None, containers_used=None, origin=None, mode="regular"):

    print("\n====================================")
    print("  RISE TVP — ship configuration")
    print(f"  OS: {platform.system()}")
    print("====================================")

    # 1. Category
    print("\nCategory:\n  1. AIR AND SPACE\n  2. ONLY AIR")
    while True:
        try:
            choice = int(input("Category number: ").strip())
            if choice in (1, 2):
                category = ["AIR AND SPACE", "ONLY AIR"][choice - 1]
                break
        except ValueError:
            pass
        print("  Enter 1 or 2.")

    # 2. Ship model
    ships = SHIPS[category]
    ship_list = list(ships.keys())
    print(f"\nAvailable ships in {category}:")
    for i, s in enumerate(ship_list, 1):
        d = ships[s]
        cap = d["cargo_base"] + d["max_containers"] * d["container_mt"]
        print(f"  {i}. {s}  —  {d['cargo_base']} + {d['max_containers']}×{d['container_mt']} MT  (max {cap} MT)")
    while True:
        try:
            choice = int(input("Model number: ").strip())
            if 1 <= choice <= len(ship_list):
                selected_ship = ship_list[choice - 1]
                break
        except ValueError:
            pass
        print(f"  Enter 1–{len(ship_list)}.")

    ship_data = ships[selected_ship]
    max_c     = ship_data["max_containers"]

    # 3. Containers
    if containers_used is None:
        if max_c == 0:
            print(f"\n  {selected_ship} has no containers. Fixed: {ship_data['cargo_base']} MT.")
            containers_used = 0
        else:
            print(f"\n  {selected_ship} supports up to {max_c} container(s) × {ship_data['container_mt']} MT each.")
            while True:
                try:
                    n = int(input(f"  Containers today (0–{max_c}): ").strip())
                    if 0 <= n <= max_c:
                        containers_used = n
                        break
                except ValueError:
                    pass
                print(f"  Enter 0–{max_c}.")

    ship_capacity = ship_data["cargo_base"] + containers_used * ship_data["container_mt"]
    print(f"  Capacity: {ship_data['cargo_base']} + {containers_used}×{ship_data['container_mt']} = {ship_capacity} MT")

    # 4. Rental
    is_rental         = False
    ship_rental_cost  = ship_data.get("rental_cost_per_day")
    print(f"\n  Rented (A) or Purchased (C)?")
    while True:
        r = input("  Choice: ").strip().upper()
        if r == "A":
            is_rental = True
            if ship_rental_cost is None:
                while True:
                    try:
                        ship_rental_cost = int(input("  Rental cost per day (CR): ").strip())
                        break
                    except ValueError:
                        print("  Enter a valid number.")
            break
        elif r == "C":
            break
        print("  Enter A or C.")

    # 5. Origin
    if origin is None:
        origin = _prompt_city("SELECT ORIGIN CITY:")

    # 6. Budget
    if budget is None:
        b_input = input("\n  Initial budget CR (Enter to skip): ").strip()
        if b_input:
            try:
                budget = int(b_input)
            except ValueError:
                print("  Invalid — skipped.")
    if budget:
        print(f"  Budget: {budget:,} CR")

    # 7. Trade Route extra prompts
    trade_route_params = None
    if mode == "route":
        trade_route_params = {
            "allowed_commodities": _prompt_commodities(),
            "max_hops":            _prompt_max_hops(),
        }

    # ── Process images ────────────────────────────────────────────────────────
    load_color_maps()

    if not os.path.exists(image_folder):
        print(f"\n✗  Folder '{image_folder}' not found.")
        return

    data_dict = {
        "Cities": {
            "header": ["Location", "Category", "Commodity Type", "Quantity MT",
                       "Reserve MT", "Selling CR/MT", "Buying CR/MT", "Maximum MT"],
            "rows": [],
        },
        "EasyDock": {
            "header": ["Location", "Category", "Name", "MT", "Buying MT",
                       "Buying CR", "Selling MT", "Selling CR"],
            "rows": [],
        },
        "_totals": {},   # location → {"mt_used", "mt_total", "cr_total"}
    }

    for filename in sorted(os.listdir(image_folder)):
        if not filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff")):
            continue

        base_name = filename.rsplit("_", 1)[0]

        # Skip non-city images (e.g. TVI newspaper header images)
        if base_name.lower().startswith(SKIP_IMAGE_PREFIXES):
            print(f"  — Skipped (non-city image): {filename}")
            continue

        sheet_key  = "EasyDock" if base_name.lower() == "easydock" else "Cities"
        image_path = os.path.join(image_folder, filename)
        
        # 1. Cargamos la imagen con OpenCV para poder analizar colores
        img_cv = cv2.imread(image_path)
        
        if img_cv is not None:
            if sheet_key == "Cities":
                # USAR EL NUEVO PROCESADOR COLOR-AWARE PARA CIUDADES
                # Esto detecta rojos y usa palabras clave para no saltarse filas (como en Deois)
                rows, totals = _parse_city_rows_with_color(img_cv)
                parsed = {"rows": rows, "totals": totals, "layout": "city"}
            else:
                # PARA EASYDOCK: Seguimos usando el OCR tradicional por ahora
                text = extract_text_from_image(image_path)
                parsed = parse_text_to_data(text) if text else None

            if parsed and parsed["rows"]:
                rows_added = 0
                for row in parsed["rows"]:
                    row.insert(0, base_name)
                    rows_added += 1
                data_dict[sheet_key]["rows"].extend(parsed["rows"])

                # Guardar totales (solo para el layout de ciudad)
                if parsed.get("totals") and sheet_key == "Cities":
                    data_dict["_totals"][base_name] = parsed["totals"]
                    t = parsed["totals"]
                    print(f"  ✔ {filename} → {sheet_key}  "
                          f"({rows_added} filas | "
                          f"MT: {t['mt_used']:,}/{t['mt_total']:,} | "
                          f"CR: {t['cr_total']:,})")
                else:
                    print(f"  ✔ {filename} → {sheet_key}  ({rows_added} filas)")
            else:
                print(f"  ⚠ No se pudieron procesar datos de: {filename}")
        else:
            print(f"  ✗ No se pudo abrir la imagen: {filename}")

# --- CÁLCULO Y FILTRADO ---
    if data_dict["Cities"]["rows"] or ( "EasyDock" in data_dict and data_dict["EasyDock"]["rows"] ):
        catalog = build_trade_catalog(data_dict)
        
        # Calculamos todas las rutas posibles primero
        raw_opps = find_trade_opportunities(catalog, ship_capacity, budget)
        
        # Filtramos por ciudad si es el modo 2
        if mode == "city" and origin:
            opportunities = [o for o in raw_opps if o['source'] == origin]
        else:
            opportunities = raw_opps

        print(f"\n{'='*60}")
        print(f"   Opportunities found: {len(opportunities)}")
        print(f"{'='*60}")

        # Llamada a la función de guardado
        save_to_excel(
            data_dict, selected_ship, ship_capacity,
            output_file=output_file, is_rental=is_rental,
            rental_cost_per_day=ship_rental_cost, origin=origin,
            budget=budget, containers_used=containers_used,
            mode=mode, trade_route_params=trade_route_params,
            all_opportunities=opportunities  # Ahora la función sí aceptará esto
        )
    else:
        print("\n  No data found to save.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rise TVP Trade Route Optimizer")
    parser.add_argument("--images",     default="images",           help="Image folder path")
    parser.add_argument("--ship",       default=None,               help="Ship name")
    parser.add_argument("--output",     default="final_trade.xlsx", help="Output Excel file")
    parser.add_argument("--budget",     default=None, type=int,     help="Initial budget in CR")
    parser.add_argument("--containers", default=None, type=int,     help="Number of cargo containers")
    parser.add_argument("--origin",     default=None,               help="Origin city")
    parser.add_argument("--mode",       default="regular",
                        choices=["regular", "city", "route"],       help="Analysis mode")
    args = parser.parse_args()
    main(args.images, selected_ship=args.ship, output_file=args.output,
         budget=args.budget, containers_used=args.containers,
         origin=args.origin, mode=args.mode)
