"""
ocr_to_excel.py — Rise TVP Trade Route Optimizer  (v2 — refactored)

Architecture
────────────
  Layer 1 – OCR / Parsing     parse_city_image()  → CitySnapshot
  Layer 2 – Catalog builder   build_trade_catalog() → dict
  Layer 3 – Optimizer         find_trade_opportunities(), assign_grades()
  Layer 4 – Excel export      save_to_excel()

Single source of truth for commodities: COMMODITY_CATEGORIES list.
No legacy Tax / Fee / Distance columns in output.
"""

import argparse
import itertools
import json
import math
import os
import platform
import re
import time
from dataclasses import dataclass, field
from typing import Optional

import cv2
import numpy as np
import openpyxl
import pytesseract
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image
from pytesseract import Output

# ─────────────────────────────────────────────────────────────────────────────
# TESSERACT — auto-configure
# ─────────────────────────────────────────────────────────────────────────────

if platform.system() == "Windows":
    for _cand in [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    ]:
        if os.path.isfile(_cand):
            pytesseract.pytesseract.tesseract_cmd = _cand
            break

# ─────────────────────────────────────────────────────────────────────────────
# CENTRAL CONSTANTS  ← single source of truth
# ─────────────────────────────────────────────────────────────────────────────

# To add a new commodity, append it here — nowhere else.
COMMODITY_CATEGORIES: list[str] = [
    "Rare/Precious",
    "Foodstuffs",
    "Natural Materials",
    "Fuel Ore",
    "Consumer Goods",
    "Fabricated Material",
    "Refined Fuel",
]

SKIP_IMAGE_PREFIXES: tuple[str, ...] = ("tvi",)

SHIPS: dict = {
    "AIR AND SPACE": {
        "E-10 Saint":        {"cargo_base": 7, "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "E-11 Saint":        {"cargo_base": 7, "max_containers": 6, "container_mt": 17, "rental_cost_per_day": None},
        "P-13 Prowler":      {"cargo_base": 1, "max_containers": 0, "container_mt": 17, "rental_cost_per_day": None},
        "W-6 Manx":          {"cargo_base": 7, "max_containers": 3, "container_mt": 17, "rental_cost_per_day": None},
    },
    "ONLY AIR": {
        "A-4 Wanderer":      {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
        "T-19 Stratomaster": {"cargo_base": 1, "max_containers": 1, "container_mt": 17, "rental_cost_per_day": None},
    },
}

CITIES: dict = {
    "Alphaville":  {"x": 420, "y": 180},
    "Comstock":    {"x": 380, "y": 210},
    "Deadwood":    {"x": 340, "y": 230},
    "Ederar":      {"x": 300, "y": 200},
    "Erie":        {"x": 460, "y": 250},
    "Freedom":     {"x": 200, "y": 320},
    "Gettysburg":  {"x": 400, "y": 290},
    "Kansas":      {"x": 250, "y": 300},
    "Lancaster":   {"x": 350, "y": 350},
    "Pimli":       {"x": 270, "y": 280},
    "SovietUnion": {"x": 500, "y": 150},
    "Terrazul":    {"x": 480, "y": 320},
    "Sharney 1":   {"x": 320, "y": 400},
    "Sharney 2":   {"x": 360, "y": 450},
    "Sharney 3":   {"x": 400, "y": 500},
    "Delois Spot": {"x": 310, "y": 260},
}
CITY_LIST: list[str] = sorted(CITIES.keys())

TAKEOFF_TIME  = 5
LANDING_TIME  = 10
PLANET_RADIUS = 2898.805   # km
FLIGHT_SPEED  = 10.0       # km/min
MIN_TRAVEL_TIME = 5.0

FLIGHT_TIMES_EXPLICIT: dict[tuple[str, str], int] = {
    ("Delois Spot", "Alphaville"):  60, ("Delois Spot", "Comstock"):    55,
    ("Delois Spot", "Deadwood"):    60, ("Delois Spot", "Ederar"):      60,
    ("Delois Spot", "Erie"):        60, ("Delois Spot", "Freedom"):    150,
    ("Delois Spot", "Gettysburg"):  60, ("Delois Spot", "Kansas"):     150,
    ("Delois Spot", "Lancaster"):  120, ("Delois Spot", "Pimli"):       35,
    ("Delois Spot", "Sharney 1"):   60, ("Delois Spot", "Sharney 2"):  120,
    ("Delois Spot", "Sharney 3"):  180, ("Delois Spot", "SovietUnion"): 60,
    ("Delois Spot", "Terrazul"):    60,
    ("Kansas", "Alphaville"):  35, ("Kansas", "Comstock"):    30,
    ("Kansas", "Deadwood"):    25, ("Kansas", "Ederar"):      20,
    ("Kansas", "Erie"):        45, ("Kansas", "Freedom"):     15,
    ("Kansas", "Gettysburg"):  40, ("Kansas", "Lancaster"):   50,
    ("Kansas", "Pimli"):       10, ("Kansas", "Sharney 1"):   30,
    ("Kansas", "Sharney 2"):   60, ("Kansas", "Sharney 3"):   90,
    ("Kansas", "SovietUnion"): 65, ("Kansas", "Terrazul"):    60,
}

CITY_COORDINATES: dict[str, Optional[dict]] = {
    "Alphaville":  {"lat": -4.426,  "lon": -22.115},
    "Deadwood":    {"lat": -11.077, "lon": -26.543},
    "Freedom":     {"lat": -12.812, "lon":  15.938},
    "Gettysburg":  {"lat": -2.616,  "lon": -35.244},
    "Kansas":      {"lat": -0.988,  "lon":  22.243},
    "SovietUnion": {"lat": -8.060,  "lon": -23.212},
    "Delois Spot": {"lat": -4.222,  "lon": -26.066},
    "Comstock": None, "Ederar": None, "Erie": None,
    "Lancaster": None, "Pimli": None, "Terrazul": None,
}

_ref_speeds = []
for (_s, _d), _m in FLIGHT_TIMES_EXPLICIT.items():
    if _s == "Kansas" and _d in CITIES and _m > 0:
        dx = CITIES["Kansas"]["x"] - CITIES[_d]["x"]
        dy = CITIES["Kansas"]["y"] - CITIES[_d]["y"]
        _ref_speeds.append(math.sqrt(dx*dx + dy*dy) / _m)
SPEED_UNITS_PER_MIN = sum(_ref_speeds) / len(_ref_speeds) if _ref_speeds else 3.0

COMMODITY_COLORS: dict[str, str] = {
    "Rare/Precious":       "FFD700",
    "Foodstuffs":          "32CD32",
    "Natural Materials":   "1E90FF",
    "Fuel Ore":            "708090",
    "Consumer Goods":      "DC143C",
    "Fabricated Material": "8A2BE2",
    "Refined Fuel":        "FFA500",
}

GRADE_STYLES: dict[str, dict] = {
    "A": {"fill": "00B050", "font": "FFFFFF"},
    "B": {"fill": "70AD47", "font": "FFFFFF"},
    "C": {"fill": "FFD700", "font": "000000"},
    "D": {"fill": "FFA500", "font": "FFFFFF"},
}

# Keywords used to keyword-match each commodity row in OCR output.
# Derived automatically from COMMODITY_CATEGORIES so there's one place to edit.
CITY_ROW_KEYWORDS: dict[str, tuple] = {
    "Rare/Precious":       ("rare", "precious", "preci"),
    "Foodstuffs":          ("food", "stuff"),
    "Natural Materials":   ("natural", "mater"),
    "Fuel Ore":            ("fuel ore", "ore"),
    "Consumer Goods":      ("consumer", "goods"),
    "Fabricated Material": ("fabricated", "fabric"),
    "Refined Fuel":        ("refined", "fuel"),
}
assert set(CITY_ROW_KEYWORDS.keys()) == set(COMMODITY_CATEGORIES), \
    "CITY_ROW_KEYWORDS must cover every entry in COMMODITY_CATEGORIES"

_SKIP_KEYWORDS = (
    "totals", "refresh", "cancel", "population",
    "fees", "ports staffed", "mt free", "cr free",
)

CITY_HEADER_KEYWORDS = (
    "commodity type", "quantity mt", "reserve mt",
    "selling cr/mt", "buying cr/mt", "maximum mt",
)

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 1 — DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class CommodityRow:
    """One commodity row as read from a city screenshot."""
    commodity:    str
    quantity_mt:  int   = 0
    reserve_mt:   int   = 0
    sell_price:   Optional[int] = None   # Price market SELLS to us (our buy cost)
    buy_price:    Optional[int] = None   # Price market BUYS from us (our sell revenue)
    maximum_mt:   int   = 0
    sell_locked:  bool  = False          # True if sell_price was displayed in red
    buy_locked:   bool  = False          # True if buy_price was displayed in red

    @property
    def sell_capacity(self) -> int:
        """How much we can actually buy from this city."""
        return max(self.quantity_mt - self.reserve_mt, 0)

    @property
    def buy_capacity(self) -> int:
        """How much this city will accept from us."""
        return self.maximum_mt


@dataclass
class MarketFooter:
    """The 'Totals' row parsed from the bottom of a city screen."""
    mt_occupied:  int = 0
    mt_available: int = 0
    cr_balance:   int = 0


@dataclass
class CitySnapshot:
    """Complete parsed data from one city screenshot."""
    city_name:   str
    commodities: list[CommodityRow]     = field(default_factory=list)
    footer:      Optional[MarketFooter] = None

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 1 — OCR / PARSING  (image → CitySnapshot)
# ─────────────────────────────────────────────────────────────────────────────

def _is_red(img_cv, x: int, y: int, w: int, h: int, threshold: float = 0.15) -> bool:
    """Return True if the given bounding box contains predominantly red pixels."""
    pad = 2
    roi = img_cv[max(0, y-pad):y+h+pad, max(0, x-pad):x+w+pad]
    if roi.size == 0:
        return False
    hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    m1 = cv2.inRange(hsv, np.array([0,   50, 50]), np.array([10,  255, 255]))
    m2 = cv2.inRange(hsv, np.array([160, 50, 50]), np.array([180, 255, 255]))
    ratio = np.count_nonzero(cv2.bitwise_or(m1, m2)) / (roi.shape[0] * roi.shape[1])
    return ratio > threshold


def _keyword_match_commodity(text: str) -> Optional[str]:
    """Return canonical commodity name matching any keyword in text, else None."""
    low = text.lower()
    for commodity, keywords in CITY_ROW_KEYWORDS.items():
        if any(kw in low for kw in keywords):
            return commodity
    return None


def _parse_footer(text: str) -> Optional[MarketFooter]:
    """
    Parse the Totals footer line, e.g.:
      'Totals  64,330 / 65,535 MT  97,122,100 CR'

    Strategy: find all integers in the line; the first two are MT values,
    the third is the CR balance.  Returns None if not enough numbers found.
    """
    if "total" not in text.lower():
        return None
    numbers = [int(n.replace(",", "")) for n in re.findall(r"[\d,]+", text) if n.replace(",", "").isdigit()]
    if len(numbers) >= 3:
        return MarketFooter(mt_occupied=numbers[0], mt_available=numbers[1], cr_balance=numbers[2])
    return None


def parse_city_image(img_cv, city_name: str) -> CitySnapshot:
    """
    Layer 1 entry point — pure OCR/parsing, no trade logic.

    Uses coordinate-anchored column detection so values land in the right
    column even when OCR drops a cell.  Red prices are flagged as locked
    and excluded from trade calculations.

    Column layout (% of image width):
        Qty     30–45%
        Reserve 46–58%
        Sell    59–72%   ← price market sells TO us  (our purchase cost)
        Buy     73–86%   ← price market pays US       (our sale revenue)
        Max     87–100%
    """
    snapshot = CitySnapshot(city_name=city_name)

    d = pytesseract.image_to_data(img_cv, output_type=Output.DICT)
    img_width = img_cv.shape[1]

    # --- UPDATE INSIDE parse_city_image ---
    COL_RANGES = {
        "qty":     (0.25, 0.45), # Widened
        "reserve": (0.45, 0.58),
        "sell":    (0.58, 0.72), 
        "buy":     (0.72, 0.87), # Shifted slightly right
        "max":     (0.85, 1.00), # Overlap with 'buy' is okay
    }
    # Group words by physical scan-line (bucket top-coord to nearest 10px)
    lines: dict[int, list[dict]] = {}
    for i, text in enumerate(d["text"]):
        word = text.strip()
        if not word:
            continue
        y_bucket = d["top"][i] // 10 * 10
        lines.setdefault(y_bucket, []).append({
            "text":  word,
            "x":     d["left"][i] / img_width,
            "red":   _is_red(img_cv, d["left"][i], d["top"][i], d["width"][i], d["height"][i]),
            "raw_x": d["left"][i],
            "raw_y": d["top"][i],
        })

    footer_found = False
    for y in sorted(lines.keys()):
        line_words = lines[y]
        line_text  = " ".join(w["text"] for w in line_words)

        # ── Footer / Totals row — parse separately, stop commodity collection ──
        if "total" in line_text.lower():
            footer = _parse_footer(line_text)
            if footer:
                snapshot.footer = footer
            footer_found = True
            continue

        # ── Skip header and chrome lines ──
        low = line_text.lower()
        if any(kw in low for kw in _SKIP_KEYWORDS):
            continue
        if sum(1 for kw in CITY_HEADER_KEYWORDS if kw in low) >= 2:
            continue

        # ── Identify commodity by keyword ──
        commodity = _keyword_match_commodity(line_text)
        if not commodity:
            continue

        # ── Place each numeric word into the correct column bucket ──
        col_vals:  dict[str, Optional[int]] = {k: None for k in COL_RANGES}
        col_red:   dict[str, bool]          = {k: False for k in COL_RANGES}

        for w in line_words:
            clean = re.sub(r"[^0-9]", "", w["text"])
            if not clean: continue
            val = int(clean)
            
            # New logic: Check from right-to-left to ensure Max MT 
            # doesn't get stolen by the Buy Price bucket
            if w["x"] >= 0.85:
                col_vals["max"] = val
            elif 0.72 <= w["x"] < 0.85:
                col_vals["buy"] = val
                col_red["buy"]  = w["red"]
                break

        def _iv(key: str, default: int = 0) -> int:
            v = col_vals.get(key)
            return v if v is not None else default

        # Sell/Buy are None when red (locked) or not detected
        sell = None if col_red["sell"] else col_vals.get("sell")
        buy  = None if col_red["buy"]  else col_vals.get("buy")

        row = CommodityRow(
            commodity   = commodity,
            quantity_mt = _iv("qty"),
            reserve_mt  = _iv("reserve"),
            sell_price  = sell,
            buy_price   = buy,
            maximum_mt  = _iv("max"),
            sell_locked = col_red["sell"],
            buy_locked  = col_red["buy"],
        )
        snapshot.commodities.append(row)

    return snapshot


def parse_easydock_image(image_path: str, location_name: str) -> list[dict]:
    """
    EasyDock parser — commodity names not fixed-order; uses text prefix + positional columns.
    Returns raw row dicts for backward compatibility with build_trade_catalog.
    """
    text = pytesseract.image_to_string(Image.open(image_path)).strip()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rows = []

    DOCK_COL_MT = 0; DOCK_COL_BUY_MT = 1; DOCK_COL_BUY_CR = 2
    DOCK_COL_SELL_MT = 3; DOCK_COL_SELL_CR = 4

    def _extract_numbers(line):
        return [int(n.replace(",","")) for n in re.findall(r"[\d,]+", line)
                if n.replace(",","").isdigit()]

    for line in lines:
        low = line.lower()
        if any(kw in low for kw in _SKIP_KEYWORDS):
            continue
        nums = _extract_numbers(line)
        if len(nums) < 2:
            continue
        first_num = re.search(r"[\d,]+", line)
        if not first_num:
            continue
        name = line[:first_num.start()].strip()
        if not name:
            continue

        def _n(idx):
            return nums[idx] if len(nums) > idx else 0

        sell_cr = _n(DOCK_COL_SELL_CR)
        buy_cr  = _n(DOCK_COL_BUY_CR)
        if sell_cr == 0 and buy_cr == 0:
            continue

        rows.append({
            "location":   location_name,
            "category":   _infer_category(name),
            "name":       name,
            "mt":         _n(DOCK_COL_MT),
            "buying_mt":  _n(DOCK_COL_BUY_MT),
            "buying_cr":  buy_cr,
            "selling_mt": _n(DOCK_COL_SELL_MT),
            "selling_cr": sell_cr,
        })
    return rows


def _infer_category(raw: str) -> str:
    text = re.sub(r"^[^A-Za-z]*", "", str(raw or "")).strip()
    for cat in COMMODITY_CATEGORIES:
        if cat.lower() in text.lower():
            return cat
    return text

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 2 — CATALOG BUILDER  (snapshots → unified price dict)
# ─────────────────────────────────────────────────────────────────────────────

def build_trade_catalog(
    city_snapshots: list[CitySnapshot],
    easydock_rows:  list[dict],
) -> dict[str, dict[str, dict]]:
    """
    Returns:
        catalog[city_name][commodity] = {
            "sell_price":    int | None,  # price WE pay to buy from this city
            "buy_price":     int | None,  # price city pays US when we sell
            "sell_capacity": int,         # MT available for us to purchase
            "buy_capacity":  int,         # MT city will accept from us
        }

    Mercantile Loop:
        Source city  → use sell_price   (Selling CR/MT column)
        Destination  → use buy_price    (Buying CR/MT column)
        Profit/MT    = destination.buy_price - source.sell_price
    """
    catalog: dict = {}

    # ── City screenshots ──────────────────────────────────────────────────────
    for snap in city_snapshots:
        city = catalog.setdefault(snap.city_name, {})
        for row in snap.commodities:
            city[row.commodity] = {
                "sell_price": {"value": row.sell_price, "is_locked": row.sell_locked},
                "buy_price":  {"value": row.buy_price,  "is_locked": row.buy_locked},
                "sell_capacity": row.sell_capacity,
                "buy_capacity":  row.buy_capacity,
            }

    # ── EasyDock rows ─────────────────────────────────────────────────────────
    for row in easydock_rows:
        location  = row["location"]
        commodity = row["category"]
        city = catalog.setdefault(location, {})
        city[commodity] = {
            "sell_price":    row.get("selling_cr") or None,
            "buy_price":     row.get("buying_cr")  or None,
            "sell_capacity": row.get("selling_mt", 0),
            "buy_capacity":  row.get("buying_mt",  0),
        }

    return catalog

# ─────────────────────────────────────────────────────────────────────────────
# FLIGHT TIME
# ─────────────────────────────────────────────────────────────────────────────

def _haversine(lat1, lon1, lat2, lon2) -> float:
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a  = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return PLANET_RADIUS * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def _pixel_distance(a: str, b: str) -> float:
    ca, cb = CITIES.get(a), CITIES.get(b)
    if not ca or not cb:
        return 0.0
    return math.sqrt((ca["x"]-cb["x"])**2 + (ca["y"]-cb["y"])**2)


def get_flight_time(origin: str, destination: str) -> float:
    if origin == destination:
        return 0.0
    explicit = (FLIGHT_TIMES_EXPLICIT.get((origin, destination))
             or FLIGHT_TIMES_EXPLICIT.get((destination, origin)))
    if explicit is not None:
        return TAKEOFF_TIME + explicit + LANDING_TIME
    c1, c2 = CITY_COORDINATES.get(origin), CITY_COORDINATES.get(destination)
    if c1 and c2:
        km = _haversine(c1["lat"], c1["lon"], c2["lat"], c2["lon"])
        return round(km / FLIGHT_SPEED + MIN_TRAVEL_TIME, 1)
    dist = _pixel_distance(origin, destination)
    if dist > 0:
        return round(TAKEOFF_TIME + dist / SPEED_UNITS_PER_MIN + LANDING_TIME, 1)
    return 120.0

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 3 — OPTIMIZER
# ─────────────────────────────────────────────────────────────────────────────

def find_trade_opportunities(catalog, ship_capacity, budget):
    opportunities = []
    
    for src_city, src_goods in catalog.items():
        for dst_city, dst_goods in catalog.items():
            if src_city == dst_city: continue
                
            for commodity, src_info in src_goods.items():
                if commodity in dst_goods:
                    dst_info = dst_goods[commodity]
                    
                    # New logic to handle the dictionary-based price data
                    s_price = src_info["sell_price"]["value"]
                    s_locked = src_info["sell_price"]["is_locked"]
                    
                    b_price = dst_info["buy_price"]["value"]
                    b_locked = dst_info["buy_price"]["is_locked"]
                    
                    if s_price and b_price:
                        profit_per_mt = b_price - s_price
                        
                        if profit_per_mt > 0:
                            # Assign Status Emoji
                            if s_locked and b_locked:
                                status = "🔴 FULLY LOCKED"
                            elif s_locked:
                                status = "🟠 SOURCE LOCKED"
                            elif b_locked:
                                status = "🟡 DEST LOCKED"
                            else:
                                status = "🟢 OPEN"

                            opportunities.append({
                                "commodity": commodity,
                                "source": src_city,
                                "destination": dst_city,
                                "profit_per_mt": profit_per_mt,
                                "status": status,
                                "source_selling": s_price,
                                "destination_buying": b_price,
                                "source_available": src_info["sell_capacity"],
                                "destination_capacity": dst_info["buy_capacity"],
                                "travel_time": get_flight_time(src_city, dst_city)
                            })
    
    return sorted(opportunities, key=lambda x: x["profit_per_mt"], reverse=True)


def assign_grades(
    opportunities:       list[dict],
    ship_capacity:       int,
    budget:              Optional[int] = None,
    is_rental:           bool  = False,
    rental_cost_per_day: Optional[int] = None,
) -> list[dict]:
    """Add grade, trip quantities, cost, profit, and ROI to each opportunity."""
    for op in opportunities:
        price         = op["source_selling"] or 1
        max_by_budget = int(budget / price) if budget else float("inf")
        qty  = max(min(ship_capacity,
                       op["source_available"],
                       op["destination_capacity"],
                       max_by_budget), 0)
        cost   = qty * price
        profit = qty * op["profit_per_mt"]
        roi    = (profit / cost * 100) if cost > 0 else 0.0

        op["_qty_trip"]    = qty
        op["_cost_trip"]   = cost
        op["_profit_trip"] = profit # Keep for backward compatibility
        op["total_profit"] = profit # ADD THIS LINE to fix the KeyError
        op["_roi"]         = roi
        op["_affordable"]  = qty > 0 and profit > 0
        if is_rental and rental_cost_per_day:
            op["covers_rental"] = profit >= rental_cost_per_day

    viable = sorted([o for o in opportunities if o["_affordable"]],
                    key=lambda x: x["_profit_trip"], reverse=True)
    n = len(viable)
    for idx, op in enumerate(viable):
        pct = idx / n if n else 1
        op["grade"] = "A" if pct < 0.25 else "B" if pct < 0.50 else "C" if pct < 0.75 else "D"
    for op in opportunities:
        if "grade" not in op:
            op["grade"] = "D"
    return opportunities


def compute_trade_routes(catalog, origin, allowed_commodities,
                         ship_capacity, budget, max_hops=5, top_n=20) -> list:
    """Multi-hop chained route optimizer."""
    locations = [loc for loc in catalog if loc != origin]
    all_routes, seen = [], set()

    def _best_leg(src, dst):
        travel_time = get_flight_time(src, dst)
        best = None
        for commodity in allowed_commodities:
            s = catalog.get(src, {}).get(commodity)
            d = catalog.get(dst, {}).get(commodity)
            if not s or not d:
                continue
            sp = s.get("sell_price") or 0
            dp = d.get("buy_price")  or 0
            if sp <= 0 or dp <= 0:
                continue
            ppm = dp - sp
            if ppm <= 0:
                continue
            avail = s.get("sell_capacity", 0)
            cap   = d.get("buy_capacity",  0)
            if avail <= 0 or cap <= 0:
                continue
            max_by_budget = int(budget / sp) if budget and sp > 0 else float("inf")
            qty    = max(min(ship_capacity, avail, cap, max_by_budget), 0)
            if qty <= 0:
                continue
            cost   = qty * sp
            profit = qty * ppm
            if best is None or profit > best["profit"]:
                best = dict(commodity=commodity, src=src, dst=dst,
                            buy_price=sp, sell_price=dp, profit_per_mt=ppm,
                            qty=qty, cost=cost, profit=profit,
                            roi=(profit/cost*100) if cost else 0,
                            flight_min=travel_time)
        return best or dict(commodity="— empty —", src=src, dst=dst,
                            buy_price=0, sell_price=0, profit_per_mt=0,
                            qty=0, cost=0, profit=0, roi=0, flight_min=travel_time)

    for n_stops in range(1, max_hops+1):
        for stops in itertools.permutations(locations, n_stops):
            seq = [origin] + list(stops) + [origin]
            legs, total_profit, total_cost, total_time = [], 0, 0, 0
            remaining, valid = budget, True
            for i in range(len(seq)-1):
                leg = _best_leg(seq[i], seq[i+1])
                if leg["profit"] == 0 and i < len(seq)-2:
                    valid = False; break
                legs.append(leg)
                total_profit += leg["profit"]
                total_cost   += leg["cost"]
                total_time   += leg["flight_min"]
                if remaining is not None:
                    remaining = remaining - leg["cost"] + leg["profit"]
            if not valid or total_profit <= 0:
                continue
            roi         = (total_profit / total_cost * 100) if total_cost else 0
            cr_per_hour = (total_profit / total_time * 60) if total_time else 0
            key = tuple(seq)
            if key not in seen:
                seen.add(key)
                all_routes.append(dict(stops=seq, legs=legs, n_hops=len(legs),
                                       total_profit=total_profit, total_cost=total_cost,
                                       total_time=total_time, roi=roi, cr_per_hour=cr_per_hour))
            if len(all_routes) >= top_n * 5:
                break
    return sorted(all_routes, key=lambda x: x["total_profit"], reverse=True)[:top_n]

# ─────────────────────────────────────────────────────────────────────────────
# MACRO DATA
# ─────────────────────────────────────────────────────────────────────────────

def calculate_macro_data(catalog: dict, opportunities: list) -> dict:
    city_profits, city_commodities, city_lucrative = {}, {}, {}
    for op in opportunities:
        src = op["source"]
        # Ensure we don't crash if total_profit is missing or 0
        p = op.get("total_profit", 0) 
        
        city_profits.setdefault(src, 0)
        city_commodities.setdefault(src, set())
        city_lucrative.setdefault(src, set())
        
        city_profits[src] += p
        city_commodities[src].add(op["commodity"])
        if p > 0:
            city_lucrative[src].add(op["commodity"])

    city_summary = sorted(
        [{"city": c,
          "total_profit":    city_profits.get(c, 0),
          "num_commodities": len(city_commodities.get(c, set())),
          "num_lucrative":   len(city_lucrative.get(c, set()))}
         for c in catalog],
        key=lambda x: x["total_profit"], reverse=True,
    )

    commodity_best_sellers, commodity_best_buyers = {}, {}
    for commodity in COMMODITY_CATEGORIES:
        sellers, buyers = [], []
        for city, data in catalog.items():
            if commodity not in data:
                continue
            item = data[commodity]
            
            # EXTRACT THE VALUE from the price dictionary
            sp_data = item.get("sell_price", {})
            bp_data = item.get("buy_price", {})
            
            # Handle both formats (dict from screenshots or int from EasyDock)
            sp = sp_data.get("value") if isinstance(sp_data, dict) else sp_data
            bp = bp_data.get("value") if isinstance(bp_data, dict) else bp_data
            
            if sp and sp > 0 and item.get("sell_capacity", 0) > 0:
                sellers.append({"city": city, "price": sp, "capacity": item["sell_capacity"]})
            if bp and bp > 0 and item.get("buy_capacity", 0) > 0:
                buyers.append({"city": city, "price": bp, "capacity": item["buy_capacity"]})
        
        commodity_best_sellers[commodity] = sorted(sellers, key=lambda x: x["price"])[:5]
        commodity_best_buyers[commodity]  = sorted(buyers,  key=lambda x: x["price"], reverse=True)[:5]

    return {
        "city_summary":                    city_summary,
        "commodity_best_sellers":          commodity_best_sellers,
        "commodity_best_buyers":           commodity_best_buyers,
        "top_cities_by_profit":            city_summary[:10],
        "cities_by_lucrative_commodities": sorted(city_summary,
                                                   key=lambda x: x["num_lucrative"],
                                                   reverse=True)[:10],
    }

# ─────────────────────────────────────────────────────────────────────────────
# COLOR REGISTRY
# ─────────────────────────────────────────────────────────────────────────────

import hashlib
_CITY_COLOR_MAP:      dict[str, PatternFill] = {}
_COMMODITY_COLOR_MAP: dict[str, PatternFill] = {}
_COLOR_MAPS_FILE = "color_maps.json"

def _hex_from_name(name: str) -> str:
    return hashlib.md5(name.encode()).hexdigest()[:6].upper()

def _get_fill(name: str, registry: dict) -> PatternFill:
    if name not in registry:
        c = _hex_from_name(name)
        registry[name] = PatternFill(start_color=c, end_color=c, fill_type="solid")
    return registry[name]

def get_city_fill(city: str) -> PatternFill:      return _get_fill(city,      _CITY_COLOR_MAP)
def get_commodity_fill(c: str) -> PatternFill:    return _get_fill(c,         _COMMODITY_COLOR_MAP)

def save_color_maps():
    data = {"cities":      {k: v.start_color.rgb for k, v in _CITY_COLOR_MAP.items()},
            "commodities": {k: v.start_color.rgb for k, v in _COMMODITY_COLOR_MAP.items()}}
    with open(_COLOR_MAPS_FILE, "w") as fh:
        json.dump(data, fh)

def load_color_maps():
    try:
        with open(_COLOR_MAPS_FILE) as fh:
            data = json.load(fh)
        for k, v in data.get("cities",      {}).items():
            _CITY_COLOR_MAP[k]      = PatternFill(start_color=v, end_color=v, fill_type="solid")
        for k, v in data.get("commodities", {}).items():
            _COMMODITY_COLOR_MAP[k] = PatternFill(start_color=v, end_color=v, fill_type="solid")
    except (FileNotFoundError, json.JSONDecodeError):
        pass

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — EXCEL STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _contrast(hex_color: str) -> str:
    hx = hex_color.strip().lstrip("#").ljust(6, "0")[:6]
    try:
        r, g, b = int(hx[0:2], 16), int(hx[2:4], 16), int(hx[4:6], 16)
    except ValueError:
        return "000000"
    return "000000" if (r*299 + g*587 + b*114)/1000 > 160 else "FFFFFF"

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(cells):
    for cell in cells:
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

def _style_section_title(cells):
    for cell in cells:
        cell.font      = Font(bold=True, size=14)
        cell.fill      = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _thin_border()

def _style_subsection_title(cells, color="E6E6FA"):
    for cell in cells:
        cell.font      = Font(bold=True, size=12)
        cell.fill      = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border    = _thin_border()

def _style_commodity_separator(cells, color):
    thick = Side(style="thick")
    for cell in cells:
        cell.fill   = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=thick, bottom=thick)

def _auto_size_columns(ws, min_width=10, max_width=50):
    from openpyxl.cell.cell import MergedCell
    for col_cells in ws.columns:
        if isinstance(col_cells[0], MergedCell):
            continue
        max_len = max(
            (len(str(c.value)) if c.value is not None and not isinstance(c, MergedCell) else 0)
            for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_width, max(min_width, max_len+2))

def _apply_roi_color(cell, roi_pct: float):
    if roi_pct > 100:   fill, fc = "00B050", "FFFFFF"
    elif roi_pct >= 50: fill, fc = "70AD47", "FFFFFF"
    elif roi_pct >= 20: fill, fc = "FFD700", "000000"
    elif roi_pct >= 0:  fill, fc = "FFA500", "FFFFFF"
    else:               fill, fc = "FF0000", "FFFFFF"
    cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    cell.font = Font(color=fc, bold=True)

def sanitize_sheet_name(name: str) -> str:
    return "".join("-" if ch in r'\/?*[]:"' else ch for ch in name)[:31]

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — SHEET WRITERS
# ─────────────────────────────────────────────────────────────────────────────

def _write_cities_sheet(wb, city_snapshots: list[CitySnapshot]):
    """
    Cities sheet layout:
      Columns: Commodity | Sell Price (our cost) | Buy Price (our revenue)
               | Sell Capacity | Buy Capacity | City

    Market Summary sub-table at bottom for footer data (MT/CR totals).
    No legacy Tax / Fee / Distance columns.
    """
    ws = wb.create_sheet("Cities")
    headers = ["City", "Commodity", "Sell Price (CR/MT)", "Buy Price (CR/MT)",
               "Sell Capacity (MT)", "Buy Capacity (MT)", "Sell Locked", "Buy Locked"]
    ws.append(headers)
    _style_header_row(ws[1])

    alt = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    row_i = 2
    for snap in city_snapshots:
        for cr in snap.commodities:
            ws.append([
                snap.city_name,
                cr.commodity,
                cr.sell_price,
                cr.buy_price,
                cr.sell_capacity,
                cr.buy_capacity,
                "🔒" if cr.sell_locked else "",
                "🔒" if cr.buy_locked  else "",
            ])
            if row_i % 2 == 0:
                for cell in ws[row_i]:
                    cell.fill = alt
            row_i += 1

    # ── Market Summary sub-table ──────────────────────────────────────────────
    footers = [(snap.city_name, snap.footer) for snap in city_snapshots if snap.footer]
    if footers:
        ws.append([])
        ws.append(["── Market Summary (Totals) ──"])
        _style_section_title(ws[ws.max_row])
        ws.append(["City", "MT Occupied", "MT Available", "MT Free", "CR Balance"])
        _style_header_row(ws[ws.max_row])
        for city_name, f in footers:
            mt_free = f.mt_available - f.mt_occupied
            ws.append([city_name, f.mt_occupied, f.mt_available, mt_free, f.cr_balance])
            cur = ws.max_row
            cf  = get_city_fill(city_name)
            c   = ws.cell(row=cur, column=1)
            c.fill = cf
            c.font = Font(color=_contrast(cf.start_color.rgb), bold=True)
            if f.mt_available and f.mt_occupied / f.mt_available > 0.9:
                for col in range(2, 5):
                    ws.cell(row=cur, column=col).fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid")
                    ws.cell(row=cur, column=col).font = Font(color="FFFFFF", bold=True)

    _auto_size_columns(ws)


def _write_opportunities_sheet(ws, opportunities: list[dict],
                                is_rental: bool, rental_cost_per_day: Optional[int],
                                ship_capacity: int):
    """
    Global Opportunities tab — every profitable A→B route, already graded.
    No Tax / Fee / Distance ghost columns.
    """
    cols = [
        "Grade", "Commodity",
        "Source City", "Buy Price (CR/MT)",          # sell_price = our cost
        "Destination City", "Sell Price (CR/MT)",     # buy_price  = our revenue
        "Profit/MT (CR)",
        "Src Stock (MT)", "Dst Capacity (MT)",
        "MT Loaded", "Trip Cost (CR)", "Trip Profit (CR)", "Trip ROI (%)",
        "Travel Time (min)",
    ]
    if is_rental and rental_cost_per_day:
        cols += ["Rental Cost (CR)", "Covers Rental?"]

    ws.append(cols)
    _style_header_row(ws[1])

    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    red_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    roi_col    = cols.index("Trip ROI (%)") + 1

    for op in opportunities:
        grade = op.get("grade", "D")
        row   = [
            grade,
            op["commodity"],
            op["source"],
            op["source_selling"],
            op["destination"],
            op["destination_buying"],
            op["profit_per_mt"],
            op["source_available"],
            op["destination_capacity"],
            op.get("_qty_trip", 0),
            op.get("_cost_trip", 0),
            op.get("_profit_trip", 0),
            round(op.get("_roi", 0), 2),
            op["travel_time"],
        ]
        if is_rental and rental_cost_per_day:
            covers = op.get("_profit_trip", 0) > rental_cost_per_day
            row += [rental_cost_per_day, "Yes" if covers else "No"]

        ws.append(row)
        cur = ws.max_row

        # Grade cell
        gs = GRADE_STYLES[grade]
        gc = ws.cell(row=cur, column=1)
        gc.fill = PatternFill(start_color=gs["fill"], end_color=gs["fill"], fill_type="solid")
        gc.font = Font(color=gs["font"], bold=True)
        gc.alignment = Alignment(horizontal="center")

        # Commodity
        cf = get_commodity_fill(op["commodity"])
        cc = ws.cell(row=cur, column=2)
        cc.fill = cf
        cc.font = Font(color=_contrast(cf.start_color.rgb), bold=True)

        # Source city
        sf = get_city_fill(op["source"])
        sc = ws.cell(row=cur, column=3)
        sc.fill = sf
        sc.font = Font(color=_contrast(sf.start_color.rgb), bold=True)

        # Destination city
        df = get_city_fill(op["destination"])
        dc = ws.cell(row=cur, column=5)
        dc.fill = df
        dc.font = Font(color=_contrast(df.start_color.rgb), bold=True)

        _apply_roi_color(ws.cell(row=cur, column=roi_col), op.get("_roi", 0))

        if is_rental and rental_cost_per_day:
            rc = ws.cell(row=cur, column=len(cols))
            rc.fill = green_fill if rc.value == "Yes" else red_fill
            rc.font = Font(color="FFFFFF", bold=True)

    ws.freeze_panes = "B2"
    _auto_size_columns(ws)


def _write_trade_routes_sheet(ws, routes: list, origin: str):
    alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    if not routes:
        ws.append(["No profitable routes found from", origin])
        return
    for rank, route in enumerate(routes, 1):
        summary = (f"Route #{rank}  |  {' → '.join(route['stops'])}  |  "
                   f"Profit: {route['total_profit']:,.0f} CR  |  ROI: {route['roi']:.1f}%  |  "
                   f"Time: {route['total_time']} min  |  Efficiency: {route['cr_per_hour']:,.0f} CR/h")
        ws.append([summary])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=12)
        c = ws.cell(row=r, column=1)
        c.font      = Font(bold=True, size=12, color="FFFFFF")
        c.fill      = get_city_fill(origin)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.append(["Hop", "From", "To", "Commodity", "Buy (CR/MT)", "Sell (CR/MT)",
                   "Profit/MT", "MT", "Leg Cost (CR)", "Leg Profit (CR)", "Flight (min)", "Leg ROI (%)"])
        _style_header_row(ws[ws.max_row])
        for hop_i, leg in enumerate(route["legs"], 1):
            ws.append([hop_i, leg["src"], leg["dst"], leg["commodity"],
                       leg["buy_price"], leg["sell_price"], leg["profit_per_mt"],
                       leg["qty"], leg["cost"], leg["profit"], leg["flight_min"], round(leg["roi"], 2)])
            cur = ws.max_row
            for col, city_key in ((2, leg["src"]), (3, leg["dst"])):
                fill = get_city_fill(city_key)
                cell = ws.cell(row=cur, column=col)
                cell.fill = fill
                cell.font = Font(color=_contrast(fill.start_color.rgb), bold=True)
            _apply_roi_color(ws.cell(row=cur, column=12), leg["roi"])
            if hop_i % 2 == 0:
                for col in (1, 4, 5, 6, 7, 8, 9, 10, 11):
                    ws.cell(row=cur, column=col).fill = alt_fill
        ws.append(["", "", "", "TOTAL", "", "", "", "",
                   route["total_cost"], route["total_profit"], route["total_time"], round(route["roi"], 2)])
        tot = ws.max_row
        for col in range(1, 13):
            ws.cell(row=tot, column=col).font = Font(bold=True)
            ws.cell(row=tot, column=col).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _apply_roi_color(ws.cell(row=tot, column=12), route["roi"])
        ws.append([])
    ws.freeze_panes = "A2"
    _auto_size_columns(ws)


def _write_macro_sheet(wb, macro_data: dict):
    alt2 = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    ws   = wb.create_sheet("MACRO")

    def _section(title):
        ws.append([title])
        _style_section_title(ws[ws.max_row])

    _section("SUMMARY BY CITY")
    ws.append(["City", "Total Potential Profit (CR)", "Num Commodities", "Num Lucrative"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["city_summary"], start=ws.max_row+1):
        ws.append([item["city"], item["total_profit"], item["num_commodities"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("TOP 10 CITIES BY TOTAL PROFIT")
    ws.append(["City", "Total Profit (CR)", "Num Lucrative"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["top_cities_by_profit"], start=ws.max_row+1):
        ws.append([item["city"], item["total_profit"], item["num_lucrative"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("TOP 10 CITIES BY LUCRATIVE COMMODITIES")
    ws.append(["City", "Num Lucrative", "Total Profit (CR)"])
    _style_header_row(ws[ws.max_row])
    for i, item in enumerate(macro_data["cities_by_lucrative_commodities"], start=ws.max_row+1):
        ws.append([item["city"], item["num_lucrative"], item["total_profit"]])
        if i % 2 == 0:
            for cell in ws[i]: cell.fill = alt2

    ws.append([])
    _section("BEST SELLERS AND BUYERS BY COMMODITY")
    for commodity in COMMODITY_CATEGORIES:   # ← driven by single constant
        color = COMMODITY_COLORS.get(commodity, "E6E6FA")
        ws.append([])
        ws.append([f"--- {commodity.upper()} ---"])
        _style_commodity_separator(ws[ws.max_row], color)
        for section_label, key in (
            ("Best Sellers (lowest price — where to buy)",  "commodity_best_sellers"),
            ("Best Buyers  (highest price — where to sell)", "commodity_best_buyers"),
        ):
            ws.append([])
            ws.append([f"{commodity} — {section_label}"])
            _style_subsection_title(ws[ws.max_row], color)
            ws.append(["City",
                       "Selling Price (CR/MT)" if "Seller" in section_label else "Buying Price (CR/MT)",
                       "Capacity (MT)"])
            _style_header_row(ws[ws.max_row])
            for i, entry in enumerate(macro_data[key].get(commodity, []), start=ws.max_row+1):
                ws.append([entry["city"], entry["price"], entry["capacity"]])
                if i % 2 == 0:
                    for cell in ws[i]: cell.fill = alt2

    _auto_size_columns(ws)

# ─────────────────────────────────────────────────────────────────────────────
# LAYER 4 — MAIN SAVE FUNCTION
# ─────────────────────────────────────────────────────────────────────────────

def save_to_excel(
    city_snapshots:      list[CitySnapshot],
    easydock_rows:       list[dict],
    opportunities:       list[dict],
    catalog:             dict,
    selected_ship:       str,
    ship_capacity:       int,
    output_file:         str,
    is_rental:           bool,
    rental_cost_per_day: Optional[int],
    origin:              str,
    budget:              Optional[int],
    containers_used:     Optional[int],
    mode:                str,
    trade_route_params:  Optional[dict],
):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Config ────────────────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Parameter", "Value"])
    _style_header_row(ws_cfg[1])
    ship_data, ship_cat = None, ""
    for cat, ships in SHIPS.items():
        if selected_ship in ships:
            ship_data, ship_cat = ships[selected_ship], cat
            break
    max_c   = ship_data["max_containers"] if ship_data else 0
    c_used  = containers_used if containers_used is not None else max_c
    c_mt    = ship_data["container_mt"] if ship_data else 17
    cargo_b = ship_data["cargo_base"]   if ship_data else 0
    config_rows = [
        ("Ship",                  selected_ship),
        ("Category",              ship_cat),
        ("Base cargo (fixed MT)", cargo_b),
        ("Containers carried",    f"{c_used} / {max_c} max"),
        ("MT per container",      c_mt),
        ("Total Capacity (MT)",   f"{cargo_b} + {c_used} x {c_mt} = {ship_capacity}"),
        ("Origin",                origin),
        ("OS",                    platform.system()),
        ("Status",                "Rented" if is_rental else "Purchased"),
    ]
    if is_rental and rental_cost_per_day:
        config_rows += [
            ("Rental cost/day",                f"{rental_cost_per_day} CR/day (14 h)"),
            ("Min profit/day to cover rental", f"{rental_cost_per_day} CR"),
            ("Min profit/hour",                f"{rental_cost_per_day/14:.2f} CR/h"),
        ]
    if budget:
        config_rows.append(("Initial Budget", f"{budget} CR"))
    if trade_route_params:
        config_rows += [
            ("Mode",               "Trade Route"),
            ("Max hops",           trade_route_params.get("max_hops", 5)),
            ("Allowed commodities", ", ".join(trade_route_params.get("allowed_commodities", ["ALL"]))),
        ]
    for row in config_rows:
        ws_cfg.append(list(row))
    _auto_size_columns(ws_cfg)

    # ── Cities ────────────────────────────────────────────────────────────────
    if city_snapshots:
        _write_cities_sheet(wb, city_snapshots)

    # ── EasyDock ──────────────────────────────────────────────────────────────
    if easydock_rows:
        ws_ed = wb.create_sheet("EasyDock")
        ws_ed.append(["Location", "Category", "Name", "MT",
                      "Buying MT", "Buying CR", "Selling MT", "Selling CR"])
        _style_header_row(ws_ed[1])
        alt = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        for i, row in enumerate(easydock_rows, start=2):
            ws_ed.append([row["location"], row["category"], row["name"],
                          row["mt"], row["buying_mt"], row["buying_cr"],
                          row["selling_mt"], row["selling_cr"]])
            if i % 2 == 0:
                for cell in ws_ed[i]: cell.fill = alt
        _auto_size_columns(ws_ed)

    # ── Global Opportunities (all cities) ─────────────────────────────────────
    if opportunities:
        ws_op = wb.create_sheet("Opportunities")
        _write_opportunities_sheet(ws_op, opportunities, is_rental, rental_cost_per_day, ship_capacity)

        # City-mode filtered sub-sheet
        if mode == "city" and origin:
            city_opps = [o for o in opportunities if o["source"] == origin]
            if city_opps:
                ws_city = wb.create_sheet(sanitize_sheet_name(f"From {origin}"))
                _write_opportunities_sheet(ws_city, city_opps, is_rental,
                                           rental_cost_per_day, ship_capacity)

    # ── Trade Routes ──────────────────────────────────────────────────────────
    if mode == "route" and trade_route_params:
        allowed  = trade_route_params.get("allowed_commodities", COMMODITY_CATEGORIES)
        max_hops = trade_route_params.get("max_hops", 3)
        print(f"\n  Computing trade routes from {origin} (max {max_hops} hops)…")
        routes = compute_trade_routes(catalog, origin, allowed, ship_capacity, budget, max_hops, top_n=20)
        ws_routes = wb.create_sheet(sanitize_sheet_name(f"Trade Routes {origin}"))
        _write_trade_routes_sheet(ws_routes, routes, origin)
        print(f"  Found {len(routes)} profitable routes.")
        if routes:
            best = routes[0]
            print(f"  Best: {' → '.join(best['stops'])} | "
                  f"Profit: {best['total_profit']:,.0f} CR | ROI: {best['roi']:.1f}%")

    # ── MACRO ─────────────────────────────────────────────────────────────────
    macro_data = calculate_macro_data(catalog, opportunities or [])
    _write_macro_sheet(wb, macro_data)

    # ── Save ──────────────────────────────────────────────────────────────────
    save_color_maps()
    try:
        wb.save(output_file)
        print(f"\n✔  Saved: {output_file}")
        print(f"   Ship: {selected_ship} | Capacity: {ship_capacity} MT | Origin: {origin}")
    except PermissionError:
        ts = int(time.time())
        base, ext = os.path.splitext(output_file)
        alt_path  = f"{base}_{ts}{ext}"
        wb.save(alt_path)
        print(f"⚠  '{output_file}' was in use — saved as '{alt_path}'")

# ─────────────────────────────────────────────────────────────────────────────
# PROMPT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _prompt_city(prompt_text: str) -> str:
    print(f"\n{prompt_text}")
    for i, city in enumerate(CITY_LIST, 1):
        print(f"  {i:>2}. {city}")
    while True:
        try:
            choice = int(input("City number: ").strip())
            if 1 <= choice <= len(CITY_LIST):
                return CITY_LIST[choice - 1]
        except ValueError:
            pass
        print(f"  Enter a number between 1 and {len(CITY_LIST)}.")


def _prompt_commodities() -> list:
    print("\n  Select commodities to trade (comma-separated numbers, or 0 for all):")
    print("   0. ALL commodities")
    for i, cat in enumerate(COMMODITY_CATEGORIES, 1):
        print(f"  {i:>2}. {cat}")
    while True:
        raw = input("  Choice: ").strip()
        if raw == "0":
            return list(COMMODITY_CATEGORIES)
        try:
            indices  = [int(p) for p in raw.split(",") if p.strip()]
            selected = [COMMODITY_CATEGORIES[i-1] for i in indices
                        if 1 <= i <= len(COMMODITY_CATEGORIES)]
            if selected:
                print(f"  Selected: {', '.join(selected)}")
                return selected
        except (ValueError, IndexError):
            pass
        print("  Enter valid comma-separated numbers or 0.")


def _prompt_max_hops() -> int:
    print("\n  Max intermediate stops? (2–5)")
    while True:
        try:
            n = int(input("  Max stops (2-5): ").strip())
            if 2 <= n <= 5:
                return n
        except ValueError:
            pass
        print("  Enter a number between 2 and 5.")

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main(image_folder, selected_ship=None, output_file="final_trade.xlsx",
         budget=None, containers_used=None, origin=None, mode="regular"):

    print("\n====================================")
    print("  RISE TVP — ship configuration")
    print(f"  OS: {platform.system()}")
    print("====================================")

    # 1. Category
    print("\nCategory:\n  1. AIR AND SPACE\n  2. ONLY AIR")
    while True:
        try:
            choice = int(input("Category number: ").strip())
            if choice in (1, 2):
                category = ["AIR AND SPACE", "ONLY AIR"][choice - 1]; break
        except ValueError:
            pass
        print("  Enter 1 or 2.")

    # 2. Ship model
    ships     = SHIPS[category]
    ship_list = list(ships.keys())
    print(f"\nAvailable ships in {category}:")
    for i, s in enumerate(ship_list, 1):
        d   = ships[s]
        cap = d["cargo_base"] + d["max_containers"] * d["container_mt"]
        print(f"  {i}. {s}  —  {d['cargo_base']} + {d['max_containers']}×{d['container_mt']} MT  (max {cap} MT)")
    while True:
        try:
            choice = int(input("Model number: ").strip())
            if 1 <= choice <= len(ship_list):
                selected_ship = ship_list[choice - 1]; break
        except ValueError:
            pass
        print(f"  Enter 1–{len(ship_list)}.")

    ship_data = ships[selected_ship]
    max_c     = ship_data["max_containers"]

    # 3. Containers
    if containers_used is None:
        if max_c == 0:
            print(f"\n  {selected_ship} has no containers. Fixed: {ship_data['cargo_base']} MT.")
            containers_used = 0
        else:
            print(f"\n  {selected_ship} supports up to {max_c} container(s) × {ship_data['container_mt']} MT each.")
            while True:
                try:
                    n = int(input(f"  Containers today (0–{max_c}): ").strip())
                    if 0 <= n <= max_c:
                        containers_used = n; break
                except ValueError:
                    pass
                print(f"  Enter 0–{max_c}.")

    ship_capacity = ship_data["cargo_base"] + containers_used * ship_data["container_mt"]
    print(f"  Capacity: {ship_data['cargo_base']} + {containers_used}×{ship_data['container_mt']} = {ship_capacity} MT")

    # 4. Rental
    is_rental, ship_rental_cost = False, ship_data.get("rental_cost_per_day")
    print("\n  Rented (A) or Purchased (C)?")
    while True:
        r = input("  Choice: ").strip().upper()
        if r == "A":
            is_rental = True
            if ship_rental_cost is None:
                while True:
                    try:
                        ship_rental_cost = int(input("  Rental cost per day (CR): ").strip()); break
                    except ValueError:
                        print("  Enter a valid number.")
            break
        elif r == "C":
            break
        print("  Enter A or C.")

    # 5. Origin
    if origin is None:
        origin = _prompt_city("SELECT ORIGIN CITY:")

    # 6. Budget
    if budget is None:
        b_input = input("\n  Initial budget CR (Enter to skip): ").strip()
        if b_input:
            try:
                budget = int(b_input)
            except ValueError:
                print("  Invalid — skipped.")
    if budget:
        print(f"  Budget: {budget:,} CR")

    # 7. Trade Route extra prompts
    trade_route_params = None
    if mode == "route":
        trade_route_params = {
            "allowed_commodities": _prompt_commodities(),
            "max_hops":            _prompt_max_hops(),
        }

    # ── Process images ────────────────────────────────────────────────────────
    load_color_maps()

    if not os.path.exists(image_folder):
        print(f"\n✗  Folder '{image_folder}' not found.")
        return

    city_snapshots: list[CitySnapshot] = []
    easydock_rows:  list[dict]          = []

    for filename in sorted(os.listdir(image_folder)):
        if not filename.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff")):
            continue

        base_name  = filename.rsplit("_", 1)[0]
        image_path = os.path.join(image_folder, filename)

        # Skip non-city assets (e.g. TVI newspaper images)
        if base_name.lower().startswith(SKIP_IMAGE_PREFIXES):
            print(f"  — Skipped: {filename}")
            continue

        img_cv = cv2.imread(image_path)
        if img_cv is None:
            print(f"  ✗ Could not open image: {filename}")
            continue

        if base_name.lower() == "easydock":
            rows = parse_easydock_image(image_path, base_name)
            easydock_rows.extend(rows)
            print(f"  ✔ {filename} → EasyDock  ({len(rows)} rows)")
        else:
            # ── Layer 1: pure OCR → CitySnapshot ──────────────────────────────
            snap = parse_city_image(img_cv, base_name)
            city_snapshots.append(snap)
            footer_info = ""
            if snap.footer:
                f = snap.footer
                mt_free     = f.mt_available - f.mt_occupied
                footer_info = (f" | MT: {f.mt_occupied:,}/{f.mt_available:,}"
                               f" (free: {mt_free:,}) | CR: {f.cr_balance:,}")
            print(f"  ✔ {filename} → Cities  ({len(snap.commodities)} commodities{footer_info})")

    if not city_snapshots and not easydock_rows:
        print("\n  No data found to save.")
        return

    # ── Layer 2: build catalog ─────────────────────────────────────────────────
    catalog = build_trade_catalog(city_snapshots, easydock_rows)

    # ── Layer 3: global opportunities + grading ───────────────────────────────
    raw_opps    = find_trade_opportunities(catalog, ship_capacity, budget)
    graded_opps = assign_grades(raw_opps, ship_capacity=ship_capacity, budget=budget,
                                is_rental=is_rental, rental_cost_per_day=ship_rental_cost)
    graded_opps.sort(key=lambda x: x.get("_profit_trip", 0), reverse=True)

    print(f"\n{'='*60}")
    print(f"   Total opportunities found: {len(graded_opps)}")
    if origin:
        from_origin = sum(1 for o in graded_opps if o["source"] == origin)
        print(f"   From {origin}: {from_origin}")
    print(f"{'='*60}")

    # ── Layer 4: export ───────────────────────────────────────────────────────
    save_to_excel(
        city_snapshots      = city_snapshots,
        easydock_rows       = easydock_rows,
        opportunities       = graded_opps,       # global — all cities
        catalog             = catalog,
        selected_ship       = selected_ship,
        ship_capacity       = ship_capacity,
        output_file         = output_file,
        is_rental           = is_rental,
        rental_cost_per_day = ship_rental_cost,
        origin              = origin,
        budget              = budget,
        containers_used     = containers_used,
        mode                = mode,
        trade_route_params  = trade_route_params,
    )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Rise TVP Trade Route Optimizer")
    parser.add_argument("--images",     default="images",           help="Image folder path")
    parser.add_argument("--ship",       default=None,               help="Ship name")
    parser.add_argument("--output",     default="final_trade.xlsx", help="Output Excel file")
    parser.add_argument("--budget",     default=None, type=int,     help="Initial budget in CR")
    parser.add_argument("--containers", default=None, type=int,     help="Number of cargo containers")
    parser.add_argument("--origin",     default=None,               help="Origin city")
    parser.add_argument("--mode",       default="regular",
                        choices=["regular", "city", "route"],       help="Analysis mode")
    args = parser.parse_args()
    main(args.images, selected_ship=args.ship, output_file=args.output,
         budget=args.budget, containers_used=args.containers,
         origin=args.origin, mode=args.mode)
